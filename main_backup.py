import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import os
import pandas as pd
from typing import Dict
from datetime import datetime
import json
import qrcode
from PIL import Image, ImageTk
import io
import re


# Helper function untuk mendapatkan path AppData
def get_appdata_path():
    """Get AppData Local path untuk menyimpan database dan export files"""
    appdata = os.getenv('LOCALAPPDATA')  # C:\Users\Username\AppData\Local
    if not appdata:
        appdata = os.path.expanduser('~\\AppData\\Local')
    
    # Buat folder khusus aplikasi
    app_folder = os.path.join(appdata, 'ArsipDigitalOwnCloud')
    if not os.path.exists(app_folder):
        os.makedirs(app_folder)
    
    return app_folder


def get_database_path():
    """Get full path untuk database.xlsx di AppData"""
    return os.path.join(get_appdata_path(), 'database.xlsx')


def get_export_path():
    """Get full path untuk file_export.xlsx di AppData"""
    return os.path.join(get_appdata_path(), 'file_export.xlsx')

# Import business logic modules
from arsip_logic import ArsipProcessor, FileManager, AnggotaFolderReader
from web_server import get_web_server_manager

# Import untuk PDF dan OCR
try:
    import pytesseract
    from pdf2image import convert_from_path
    OCR_AVAILABLE = True
except ImportError:
    OCR_AVAILABLE = False
    pytesseract = None
    convert_from_path = None


def get_responsive_dimensions(base_width, base_height, screen_width, screen_height):
    if screen_width >= 1920:  # Large screens (4K, etc)
        width = base_width
        height = base_height
        padding = 30
        fonts = {'title': 18, 'subtitle': 11, 'normal': 10, 'small': 9}
    elif screen_width >= 1366:  # Medium screens (standard laptop)
        width = int(base_width * 0.95)
        height = int(base_height * 0.95)
        padding = 25
        fonts = {'title': 16, 'subtitle': 10, 'normal': 9, 'small': 8}
    elif screen_width >= 1024:  # Small laptop
        width = int(base_width * 0.85)
        height = int(base_height * 0.85)
        padding = 20
        fonts = {'title': 14, 'subtitle': 9, 'normal': 8, 'small': 7}
    else:  # Very small screens
        width = int(base_width * 0.75)
        height = int(base_height * 0.75)
        padding = 15
        fonts = {'title': 12, 'subtitle': 8, 'normal': 7, 'small': 6}
    
    # Ensure window doesn't exceed 85% of screen size
    max_width = int(screen_width * 0.85)
    max_height = int(screen_height * 0.85)
    width = min(width, max_width)
    height = min(height, max_height)
    
    return width, height, padding, fonts


# ConfigManager untuk mengelola settings
class ConfigManager:
    """Manager untuk menyimpan dan membaca konfigurasi aplikasi"""
    
    def __init__(self):
        self.config_file = "app_config.json"
        self.default_config = {
            "default_folder": "",
            "web_server_enabled": False,
            "web_server_port": 1212
        }
        self.config = self.load_config()
    
    def load_config(self):
        """Load konfigurasi dari file"""
        try:
            if os.path.exists(self.config_file):
                with open(self.config_file, 'r', encoding='utf-8') as f:
                    return json.load(f)
            else:
                return self.default_config.copy()
        except Exception as e:
            print(f"Error loading config: {e}")
            return self.default_config.copy()
    
    def save_config(self):
        """Simpan konfigurasi ke file"""
        try:
            with open(self.config_file, 'w', encoding='utf-8') as f:
                json.dump(self.config, f, indent=4, ensure_ascii=False)
            return True
        except Exception as e:
            print(f"Error saving config: {e}")
            return False
    
    def get_default_folder(self):
        """Get default folder path"""
        return self.config.get("default_folder", "")
    
    def set_default_folder(self, folder_path):
        """Set default folder path"""
        self.config["default_folder"] = folder_path
        return self.save_config()
    
    def get_web_server_enabled(self):
        """Get web server enabled status"""
        return self.config.get("web_server_enabled", False)
    
    def set_web_server_enabled(self, enabled):
        """Set web server enabled status"""
        self.config["web_server_enabled"] = enabled
        return self.save_config()
    
    def get_web_server_port(self):
        """Get web server port"""
        return self.config.get("web_server_port", 1212)
    
    def set_web_server_port(self, port):
        """Set web server port"""
        self.config["web_server_port"] = port
        return self.save_config()


# Global config manager instance
config_manager = ConfigManager()


# Global web server manager instance (from web_server module)
web_server_manager = get_web_server_manager()


# Jalankan GUI jika file ini dijalankan langsung
if __name__ == "__main__":
    import tkinter as tk
    from main import MainMenu  # pastikan MainMenu terimport jika run as script
    
    # Hapus file_export.xlsx jika ada (di AppData)
    export_path = get_export_path()
    if os.path.exists(export_path):
        try:
            os.remove(export_path)
            print(f"File {export_path} dihapus")
        except Exception as e:
            print(f"Gagal menghapus {export_path}: {e}")
    
    root = tk.Tk()
    app = MainMenu(root)
    root.mainloop()


# Import business logic modules
from arsip_logic import ArsipProcessor, FileManager, AnggotaFolderReader


class MainMenu:
    def __init__(self, root):
        self.root = root
        self.setup_window()
        self.create_menu_widgets()
    
    def setup_window(self):
        """Setup window utama untuk menu"""
        self.root.title("Aplikasi Arsip Digital - Menu Utama")
        
        # Get screen dimensions
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        
        # Get responsive dimensions
        width, height, self.padding, self.fonts = get_responsive_dimensions(
            600, 800, screen_width, screen_height
        )
        
        self.root.geometry(f"{width}x{height}")
        self.root.resizable(True, True)
        
        # Set minimum size to prevent too small windows
        self.root.minsize(350, 450)
        
        # Center window
        self.center_window()
        
        # Set window icon (optional)
        try:
            self.root.iconbitmap("icon.ico")  # Jika ada file icon
        except:
            pass
    
    def center_window(self):
        """Center window di layar"""
        self.root.update_idletasks()
        width = self.root.winfo_width()
        height = self.root.winfo_height()
        x = (self.root.winfo_screenwidth() // 2) - (width // 2)
        y = (self.root.winfo_screenheight() // 2) - (height // 2)
        self.root.geometry(f'{width}x{height}+{x}+{y}')
    
    def create_menu_widgets(self):
        """Membuat widget menu utama dengan tampilan modern dan grid horizontal"""
        from tkinter import ttk

        # Main frame
        main_frame = ttk.Frame(self.root, padding=str(self.padding))
        main_frame.grid(row=0, column=0, sticky="nsew")
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)

        # Logo dan Title
        header_frame = ttk.Frame(main_frame)
        header_frame.grid(row=0, column=0, columnspan=2, pady=(0, 20))

        logo_label = ttk.Label(
            header_frame, text="📁", font=("Arial", max(32, self.fonts['title'] * 2))
        )
        logo_label.grid(row=0, column=0, padx=10)

        title_label = ttk.Label(
            header_frame, text="APLIKASI ARSIP DIGITAL",
            font=("Arial", self.fonts['title'], "bold")
        )
        title_label.grid(row=0, column=1, sticky="w")

        subtitle_label = ttk.Label(
            header_frame,
            text="Sistem Manajemen Arsip Digital",
            font=("Arial", self.fonts['subtitle']),
            foreground="gray"
        )
        subtitle_label.grid(row=1, column=0, columnspan=2)

        # Style tombol modern
        style = ttk.Style()
        button_padding = (max(12, self.padding-8), max(10, self.padding-12))
        style.configure(
            "Menu.TButton",
            padding=button_padding,
            font=("Arial", self.fonts['normal']),
            relief="flat"
        )

        # Hover effect
        style.map(
            "Menu.TButton",
            background=[("active", "#e6e6e6")],
            relief=[("pressed", "sunken")]
        )

        buttons = [
            ("📋 Cek Arsip Digital", self.open_cek_arsip),
            ("📂 Scan Folder Arsip Digital", self.open_scan_folder),
            ("🌐 Universal Scan Database", self.open_universal_scan),
            ("📊 Scan File Besar", self.open_scan_large_files),
            ("💰 Cek Pengajuan Dana", self.open_cek_pengajuan_dana),
            ("👨‍👩‍👧‍👦 Cek NO KK", self.open_cek_no_kk),
            ("📃 PDF Tool", self.open_pdf_tool),
            ("⚙️ Pengaturan", self.open_settings)
        ]

        # Frame menu dengan layout 2 kolom
        menu_frame = ttk.Frame(main_frame)
        menu_frame.grid(row=1, column=0, columnspan=2, pady=20)

        col = 0
        row = 0
        for i, (text, command) in enumerate(buttons):
            btn = ttk.Button(menu_frame, text=text, command=command, style="Menu.TButton")
            btn.grid(row=row, column=col, sticky="nsew", padx=10, pady=10, ipadx=10, ipady=10)

            # Buat grid 2 kolom
            col += 1
            if col > 1:
                col = 0
                row += 1

        # Footer
        footer_frame = ttk.Frame(main_frame)
        footer_frame.grid(row=2, column=0, columnspan=2, pady=20)

        exit_btn = ttk.Button(footer_frame, text="Keluar", command=self.exit_app)
        exit_btn.grid(row=0, column=0, padx=20)

        version_label = ttk.Label(
            footer_frame,
            text="v1.1.6 - Developed by Riky Dwianto",
            font=("Arial", self.fonts['small']),
            foreground="gray"
        )
        version_label.grid(row=1, column=0, pady=(10, 0))

    
    def open_cek_arsip(self):
        """Membuka form Cek Arsip Digital"""
        # Hide main menu window
        self.root.withdraw()
        
        # Create new window for cek arsip
        arsip_window = tk.Toplevel(self.root)
        arsip_app = ArsipDigitalApp(arsip_window, self.root)
        
        # Handle window close to return to main menu
        def on_arsip_close():
            arsip_window.destroy()
            self.root.deiconify()  # Show main menu again
        
        arsip_window.protocol("WM_DELETE_WINDOW", on_arsip_close)
    
    def open_scan_folder(self):
        """Membuka form Scan Folder Arsip Digital"""
        # Hide main menu window
        self.root.withdraw()
        
        # Create new window for scan folder
        scan_window = tk.Toplevel(self.root)
        scan_app = ScanFolderApp(scan_window, self.root)
        
        # Handle window close to return to main menu
        def on_scan_close():
            scan_window.destroy()
            self.root.deiconify()  # Show main menu again
        
        scan_window.protocol("WM_DELETE_WINDOW", on_scan_close)
    
    def open_scan_large_files(self):
        """Membuka form Scan File Besar (>10MB)"""
        # Hide main menu window
        self.root.withdraw()
        
        # Create new window for scan large files
        scan_window = tk.Toplevel(self.root)
        scan_app = ScanLargeFilesApp(scan_window, self.root)
        
        # Handle window close to return to main menu
        def on_scan_close():
            scan_window.destroy()
            self.root.deiconify()  # Show main menu again
        
        scan_window.protocol("WM_DELETE_WINDOW", on_scan_close)
    
    def open_cek_pengajuan_dana(self):
        """Membuka form Cek Pengajuan Dana"""
        # Hide main menu window
        self.root.withdraw()
        
        # Create new window for cek pengajuan dana
        pengajuan_window = tk.Toplevel(self.root)
        pengajuan_app = CekPengajuanDanaApp(pengajuan_window, self.root)
        
        # Handle window close to return to main menu
        def on_pengajuan_close():
            pengajuan_window.destroy()
            self.root.deiconify()  # Show main menu again
        
        pengajuan_window.protocol("WM_DELETE_WINDOW", on_pengajuan_close)

    def open_pdf_tool(self):
        """Membuka form PDF Tool (merge/split/convert)"""
        # Hide main menu window
        self.root.withdraw()

        pdf_window = tk.Toplevel(self.root)
        pdf_app = PDFToolApp(pdf_window, self.root)

        def on_pdf_close():
            pdf_window.destroy()
            self.root.deiconify()

        pdf_window.protocol("WM_DELETE_WINDOW", on_pdf_close)
    
    def open_settings(self):
        """Membuka form Pengaturan"""
        # Hide main menu window
        self.root.withdraw()
        
        # Create new window for settings
        settings_window = tk.Toplevel(self.root)
        settings_app = SettingsApp(settings_window, self.root)
        
        # Handle window close to return to main menu
        def on_settings_close():
            settings_window.destroy()
            self.root.deiconify()  # Show main menu again
        
        settings_window.protocol("WM_DELETE_WINDOW", on_settings_close)
    
    def open_universal_scan(self):
        """Membuka form Universal Scan Database"""
        # Hide main menu window
        self.root.withdraw()
        
        # Create new window for universal scan
        universal_window = tk.Toplevel(self.root)
        universal_app = UniversalScanApp(universal_window, self.root)
        
        # Handle window close to return to main menu
        def on_universal_close():
            universal_window.destroy()
            self.root.deiconify()  # Show main menu again
        
        universal_window.protocol("WM_DELETE_WINDOW", on_universal_close)
    
    def open_cek_no_kk(self):
        """Membuka form Cek NO KK"""
        # Hide main menu window
        self.root.withdraw()
        
        # Create new window for cek no kk
        nokk_window = tk.Toplevel(self.root)
        nokk_app = CekNoKKApp(nokk_window, self.root)
        
        # Handle window close to return to main menu
        def on_nokk_close():
            nokk_window.destroy()
            self.root.deiconify()  # Show main menu again
        
        nokk_window.protocol("WM_DELETE_WINDOW", on_nokk_close)
    

    
    def exit_app(self):
        """Keluar dari aplikasi"""
        if messagebox.askokcancel("Keluar", "Apakah Anda yakin ingin keluar dari aplikasi?"):
            self.root.destroy()


class SettingsApp:
    """Form untuk Pengaturan Aplikasi"""
    
    def __init__(self, root, parent_window=None):
        self.root = root
        self.parent_window = parent_window
        
        self.setup_window()
        self.create_widgets()
    
    def setup_window(self):
        """Setup window pengaturan"""
        self.root.title("Pengaturan - Aplikasi Arsip Digital")
        
        # Get screen dimensions
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        
        # Get responsive dimensions
        width, height, self.padding, self.fonts = get_responsive_dimensions(
            700, 800, screen_width, screen_height
        )
        
        self.root.geometry(f"{width}x{height}")
        self.root.resizable(True, True)
        
        # Set minimum size to prevent too small windows
        self.root.minsize(480, 500)
        
        # Center window
        self.center_window()
    
    def center_window(self):
        """Center window di layar"""
        self.root.update_idletasks()
        width = self.root.winfo_width()
        height = self.root.winfo_height()
        x = (self.root.winfo_screenwidth() // 2) - (width // 2)
        y = (self.root.winfo_screenheight() // 2) - (height // 2)
        self.root.geometry(f'{width}x{height}+{x}+{y}')
    
    def create_widgets(self):
        """Membuat widget untuk pengaturan dengan scrollable canvas"""
        # Main container frame
        container = ttk.Frame(self.root)
        container.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Configure grid weights
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        container.columnconfigure(0, weight=1)
        container.rowconfigure(0, weight=1)
        
        # Create canvas with scrollbar
        canvas = tk.Canvas(container, highlightthickness=0)
        scrollbar = ttk.Scrollbar(container, orient="vertical", command=canvas.yview)
        
        # Scrollable frame inside canvas
        scrollable_frame = ttk.Frame(canvas)
        
        # Configure canvas scroll region
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        # Create window in canvas
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Grid layout
        canvas.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))
        
        # Enable mousewheel scrolling
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        
        canvas.bind_all("<MouseWheel>", _on_mousewheel)
        
        # Main frame dengan padding yang responsif
        screen_width = self.root.winfo_screenwidth()
        padding_size = 30 if screen_width >= 1366 else 20 if screen_width >= 1024 else 15
        
        main_frame = ttk.Frame(scrollable_frame, padding=str(padding_size))
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Configure main_frame
        scrollable_frame.columnconfigure(0, weight=1)
        main_frame.columnconfigure(0, weight=1)
        
        # Title dengan ukuran font responsif
        title_size = 16 if screen_width >= 1366 else 14 if screen_width >= 1024 else 12
        title_label = ttk.Label(
            main_frame, 
            text="⚙️ PENGATURAN", 
            font=("Arial", title_size, "bold")
        )
        title_label.grid(row=0, column=0, pady=(0, 10))
        
        # Subtitle
        subtitle_size = 10 if screen_width >= 1366 else 9 if screen_width >= 1024 else 8
        subtitle_label = ttk.Label(
            main_frame, 
            text="Konfigurasi default untuk aplikasi",
            font=("Arial", subtitle_size),
            foreground="gray"
        )
        subtitle_label.grid(row=1, column=0, pady=(0, 20))
        
        # Frame untuk Default Folder dengan padding responsif
        folder_padding = 15 if screen_width >= 1366 else 12 if screen_width >= 1024 else 10
        folder_frame = ttk.LabelFrame(main_frame, text="Default Folder Arsip Digital", padding=str(folder_padding))
        folder_frame.grid(row=2, column=0, sticky=(tk.W, tk.E), pady=(0, 15))
        folder_frame.columnconfigure(0, weight=1)
        
        # Info label dengan wraplength responsif
        wrap_length = 500 if screen_width >= 1366 else 400 if screen_width >= 1024 else 350
        info_label = ttk.Label(
            folder_frame,
            text="Folder ini akan digunakan sebagai default saat membuka form lain",
            font=("Arial", 9),
            foreground="gray",
            wraplength=wrap_length
        )
        info_label.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        
        # Current default folder display
        current_default = config_manager.get_default_folder()
        display_text = current_default if current_default else "Belum ada folder default yang dipilih"
        
        self.folder_var = tk.StringVar(value=display_text)
        folder_path_label = ttk.Label(
            folder_frame, 
            textvariable=self.folder_var,
            foreground="blue" if current_default else "gray",
            wraplength=wrap_length,
            font=("Arial", 9, "bold" if current_default else "normal")
        )
        folder_path_label.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=(0, 15))
        
        # Buttons frame
        btn_frame = ttk.Frame(folder_frame)
        btn_frame.grid(row=2, column=0)
        
        # Browse button
        browse_btn = ttk.Button(
            btn_frame, 
            text="📂 Pilih Folder Default", 
            command=self.browse_default_folder
        )
        browse_btn.grid(row=0, column=0, padx=(0, 10))
        
        # Clear button
        clear_btn = ttk.Button(
            btn_frame, 
            text="🗑️ Hapus Default", 
            command=self.clear_default_folder
        )
        clear_btn.grid(row=0, column=1, padx=(10, 0))
        
        # ===== WEB SERVER SECTION =====
        # Frame untuk Web Server dengan padding responsif
        webserver_frame = ttk.LabelFrame(main_frame, text="🌐 Web Server", padding=str(folder_padding))
        webserver_frame.grid(row=3, column=0, sticky=(tk.W, tk.E), pady=(0, 15))
        webserver_frame.columnconfigure(0, weight=1)
        
        # Info label
        webserver_info_label = ttk.Label(
            webserver_frame,
            text="Aktifkan web server untuk akses file arsip melalui browser",
            font=("Arial", 9),
            foreground="gray",
            wraplength=wrap_length
        )
        webserver_info_label.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        
        # Server status frame
        status_frame = ttk.Frame(webserver_frame)
        status_frame.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        
        ttk.Label(status_frame, text="Status:", font=("Arial", 9, "bold")).grid(row=0, column=0, sticky=tk.W, padx=(0, 10))
        self.server_status_var = tk.StringVar(value="⚫ Tidak Aktif")
        ttk.Label(
            status_frame, 
            textvariable=self.server_status_var,
            font=("Arial", 9)
        ).grid(row=0, column=1, sticky=tk.W)
        
        # Local IP frame
        ip_frame = ttk.Frame(webserver_frame)
        ip_frame.grid(row=2, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        
        ttk.Label(ip_frame, text="IP Lokal:", font=("Arial", 9, "bold")).grid(row=0, column=0, sticky=tk.W, padx=(0, 10))
        local_ip = web_server_manager.get_local_ip()
        self.local_ip_var = tk.StringVar(value=local_ip)
        ttk.Label(
            ip_frame, 
            textvariable=self.local_ip_var,
            font=("Arial", 9),
            foreground="blue"
        ).grid(row=0, column=1, sticky=tk.W)
        
        # Port frame
        port_frame = ttk.Frame(webserver_frame)
        port_frame.grid(row=3, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        
        ttk.Label(port_frame, text="Port:", font=("Arial", 9, "bold")).grid(row=0, column=0, sticky=tk.W, padx=(0, 10))
        self.port_var = tk.StringVar(value=str(config_manager.get_web_server_port()))
        port_entry = ttk.Entry(port_frame, textvariable=self.port_var, width=10)
        port_entry.grid(row=0, column=1, sticky=tk.W)
        
        # URL frame
        url_frame = ttk.Frame(webserver_frame)
        url_frame.grid(row=4, column=0, sticky=(tk.W, tk.E), pady=(0, 15))
        url_frame.columnconfigure(0, weight=1)
        
        ttk.Label(url_frame, text="URL Akses:", font=("Arial", 9, "bold")).grid(row=0, column=0, sticky=tk.W, pady=(0, 5))
        
        self.url_local_var = tk.StringVar(value="http://localhost:1212")
        url_local_label = ttk.Label(
            url_frame, 
            textvariable=self.url_local_var,
            font=("Arial", 8),
            foreground="blue",
            cursor="hand2"
        )
        url_local_label.grid(row=1, column=0, sticky=tk.W, padx=(0, 0))
        
        self.url_network_var = tk.StringVar(value=f"http://{local_ip}:1212")
        url_network_label = ttk.Label(
            url_frame, 
            textvariable=self.url_network_var,
            font=("Arial", 8),
            foreground="blue",
            cursor="hand2"
        )
        url_network_label.grid(row=2, column=0, sticky=tk.W, padx=(0, 0))
        
        # QR Code frame
        qr_frame = ttk.LabelFrame(webserver_frame, text="📱 QR Code untuk HP", padding="10")
        qr_frame.grid(row=5, column=0, sticky=(tk.W, tk.E), pady=(10, 10))
        qr_frame.columnconfigure(0, weight=1)
        
        # QR Code info
        qr_info_label = ttk.Label(
            qr_frame,
            text="Scan QR code ini dari HP (pastikan 1 jaringan WiFi)",
            font=("Arial", 8),
            foreground="gray"
        )
        qr_info_label.grid(row=0, column=0, pady=(0, 10))
        
        # QR Code container
        self.qr_label = ttk.Label(qr_frame, text="QR Code akan muncul saat server aktif", foreground="gray")
        self.qr_label.grid(row=1, column=0)
        
        # Refresh QR button
        refresh_qr_btn = ttk.Button(
            qr_frame,
            text="🔄 Refresh QR Code",
            command=self.refresh_qr_code
        )
        refresh_qr_btn.grid(row=2, column=0, pady=(10, 0))
        
        # Server control buttons
        server_btn_frame = ttk.Frame(webserver_frame)
        server_btn_frame.grid(row=6, column=0, pady=(10, 0))
        
        self.start_server_btn = ttk.Button(
            server_btn_frame, 
            text="▶️ Start Server", 
            command=self.start_web_server
        )
        self.start_server_btn.grid(row=0, column=0, padx=(0, 10))
        
        self.stop_server_btn = ttk.Button(
            server_btn_frame, 
            text="⏹️ Stop Server", 
            command=self.stop_web_server,
            state=tk.DISABLED
        )
        self.stop_server_btn.grid(row=0, column=1, padx=(10, 0))
        
        # Update server status saat load
        self.update_server_status()
        
        # Status label
        self.status_var = tk.StringVar(value="")
        status_label = ttk.Label(
            main_frame,
            textvariable=self.status_var,
            font=("Arial", 9),
            foreground="green"
        )
        status_label.grid(row=4, column=0, pady=(0, 15))
        
        # Footer buttons
        footer_frame = ttk.Frame(main_frame)
        footer_frame.grid(row=5, column=0, pady=(10, 0))
        
        # Back button
        if self.parent_window:
            back_btn = ttk.Button(
                footer_frame, 
                text="⬅️ Kembali ke Menu", 
                command=self.back_to_menu
            )
            back_btn.grid(row=0, column=0)
        
        # Update canvas width to match window
        def _configure_canvas(event):
            canvas.itemconfig(canvas.find_withtag("all")[0], width=event.width)
        
        canvas.bind("<Configure>", _configure_canvas)
    
    def update_server_status(self):
        """Update status web server di UI"""
        info = web_server_manager.get_server_info()
        
        if info["status"] == "Running":
            self.server_status_var.set("🟢 Aktif")
            self.start_server_btn.config(state=tk.DISABLED)
            self.stop_server_btn.config(state=tk.NORMAL)
            # Generate QR Code saat server aktif
            self.generate_qr_code(info["url_network"])
        else:
            self.server_status_var.set("⚫ Tidak Aktif")
            self.start_server_btn.config(state=tk.NORMAL)
            self.stop_server_btn.config(state=tk.DISABLED)
            # Clear QR Code saat server mati
            self.qr_label.config(image='', text="QR Code akan muncul saat server aktif", foreground="gray")
        
        self.url_local_var.set(info["url_local"])
        self.url_network_var.set(info["url_network"])
    
    def generate_qr_code(self, url):
        """Generate QR code untuk URL"""
        try:
            # Create QR code
            qr = qrcode.QRCode(
                version=1,
                error_correction=qrcode.constants.ERROR_CORRECT_L,
                box_size=10,
                border=4,
            )
            qr.add_data(url)
            qr.make(fit=True)
            
            # Create image
            img = qr.make_image(fill_color="black", back_color="white")
            
            # Resize untuk fit di window (200x200)
            img = img.resize((150, 150), Image.Resampling.LANCZOS)
            
            # Convert to PhotoImage
            photo = ImageTk.PhotoImage(img)
            
            # Store reference to prevent garbage collection
            self.qr_photo = photo
            
            # Update label
            self.qr_label.config(image=photo, text="")
            
        except Exception as e:
            self.qr_label.config(image='', text=f"Error: {str(e)}", foreground="red")
    
    def refresh_qr_code(self):
        """Refresh QR code"""
        info = web_server_manager.get_server_info()
        if info["status"] == "Running":
            self.generate_qr_code(info["url_network"])
            messagebox.showinfo("QR Code", "QR Code berhasil di-refresh!")
        else:
            messagebox.showwarning("Warning", "Server belum aktif!\nStart server terlebih dahulu.")

    
    def start_web_server(self):
        """Start web server"""
        # cek dulu apakah ada file database.xlsx di AppData
        database_path = get_database_path()
        if not os.path.exists(database_path):
            messagebox.showerror("Error", f"File database.xlsx tidak ditemukan di:\n{database_path}\n\nSilakan scan folder arsip digital terlebih dahulu lalu pilih simpan dan singkron.")
            return
        else:
            try:
                port = int(self.port_var.get())
                if port < 1024 or port > 65535:
                    messagebox.showerror("Error", "Port harus antara 1024 dan 65535")
                    return
            except ValueError:
                messagebox.showerror("Error", "Port harus berupa angka")
                return
            
            # Save port to config
            config_manager.set_web_server_port(port)
            
            success, message = web_server_manager.start_server(port)
            
            if success:
                self.status_var.set("✅ Web server berhasil dijalankan!")
                messagebox.showinfo("Server Started", message)
                self.update_server_status()
                
                # Clear status after 3 seconds
                self.root.after(3000, lambda: self.status_var.set(""))
            else:
                messagebox.showerror("Error", f"Gagal start server:\n{message}")
    
    def stop_web_server(self):
        """Stop web server"""
        success, message = web_server_manager.stop_server()
        
        if success:
            self.status_var.set("✅ Web server berhasil dihentikan!")
            messagebox.showinfo("Server Stopped", message)
            self.update_server_status()
            
            # Clear QR code
            self.qr_label.config(image='', text="QR Code akan muncul saat server aktif", foreground="gray")
            
            # Clear status after 3 seconds
            self.root.after(3000, lambda: self.status_var.set(""))
        else:
            messagebox.showerror("Error", f"Gagal stop server:\n{message}")
    
    def browse_default_folder(self):
        """Pilih folder default dan auto-generate database"""
        current_default = config_manager.get_default_folder()
        initial_dir = current_default if current_default and os.path.exists(current_default) else os.getcwd()
        
        folder_path = filedialog.askdirectory(
            title="Pilih Folder Default Arsip Digital",
            initialdir=initial_dir
        )
        
        if folder_path:
            if os.path.exists(folder_path) and os.path.isdir(folder_path):
                # Save to config
                if config_manager.set_default_folder(folder_path):
                    self.folder_var.set(folder_path)
                    self.status_var.set("✅ Folder default berhasil disimpan!")
                    
                    # Tanyakan apakah user ingin auto-generate database
                    
                    # Clear status after 3 seconds
                    self.root.after(3000, lambda: self.status_var.set(""))
                else:
                    messagebox.showerror("Error", "Gagal menyimpan konfigurasi!")
            else:
                messagebox.showerror("Error", "Folder yang dipilih tidak valid!")
    
    def scan_struktur_lengkap_simple(self, root_path, status_label, progress_window):
        """Scan struktur lengkap sederhana untuk database web"""
        result = {
            "success": True,
            "root_path": root_path,
            "root_name": os.path.basename(root_path),
            "scan_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "items": []
        }
        
        try:
            item_count = 0
            
            # Scan semua file dan folder secara rekursif
            for root, dirs, files in os.walk(root_path):
                # Update status setiap 10 item
                if item_count % 10 == 0:
                    relative_root = os.path.relpath(root, root_path)
                    status_label.config(text=f"Scanning: {relative_root}...")
                    progress_window.update()
                
                # Scan folders
                for dir_name in dirs:
                    dir_path = os.path.join(root, dir_name)
                    relative_path = os.path.relpath(dir_path, root_path)
                    
                    try:
                        # Hitung ukuran folder (total file di dalamnya)
                        folder_size = 0
                        for sub_root, sub_dirs, sub_files in os.walk(dir_path):
                            for sub_file in sub_files:
                                try:
                                    sub_file_path = os.path.join(sub_root, sub_file)
                                    folder_size += os.path.getsize(sub_file_path)
                                except:
                                    pass
                        
                        folder_size_mb = round(folder_size / (1024 * 1024), 2)
                        last_modified = datetime.fromtimestamp(os.path.getmtime(dir_path)).strftime("%Y-%m-%d %H:%M:%S")
                        
                        result["items"].append({
                            'file_path': dir_path,
                            'relative_path': relative_path,
                            'status': 'folder',
                            'ukuran_mb': folder_size_mb,
                            'last_modified': last_modified,
                            'scan_time': result["scan_time"],
                            'file_name': dir_name,
                            'directory': os.path.dirname(dir_path),
                            'extension': ''
                        })
                        
                        item_count += 1
                    
                    except Exception as e:
                        print(f"Error scanning folder {dir_path}: {e}")
                
                # Scan files
                for file_name in files:
                    file_path = os.path.join(root, file_name)
                    relative_path = os.path.relpath(file_path, root_path)
                    
                    try:
                        file_size = os.path.getsize(file_path)
                        file_size_mb = round(file_size / (1024 * 1024), 2)
                        last_modified = datetime.fromtimestamp(os.path.getmtime(file_path)).strftime("%Y-%m-%d %H:%M:%S")
                        file_extension = os.path.splitext(file_name)[1]
                        
                        result["items"].append({
                            'file_path': file_path,
                            'relative_path': relative_path,
                            'status': 'file',
                            'ukuran_mb': file_size_mb,
                            'last_modified': last_modified,
                            'scan_time': result["scan_time"],
                            'file_name': file_name,
                            'directory': os.path.dirname(file_path),
                            'extension': file_extension
                        })
                        
                        item_count += 1
                    
                    except Exception as e:
                        print(f"Error scanning file {file_path}: {e}")
            
            result["total_items"] = item_count
            return result
            
        except Exception as e:
            result["success"] = False
            result["error"] = str(e)
            return result
    
    def export_simple_database_for_web(self, scan_result, output_path, status_label, progress_window):
        """Export database sederhana untuk web server"""
        try:
            status_label.config(text="Menyimpan ke Excel...")
            progress_window.update()
            
            # Prepare data untuk DataFrame
            df_data = []
            for idx, item in enumerate(scan_result["items"], 1):
                df_data.append({
                    'ID': idx,
                    'File_Path': item['file_path'],
                    'Relative_Path': item['relative_path'],
                    'Status': item['status'],
                    'Ukuran_MB': item['ukuran_mb'],
                    'Last_Modified': item['last_modified'],
                    'Scan_Time': item['scan_time'],
                    'File_Name': item['file_name'],
                    'Directory': item['directory'],
                    'Extension': item['extension']
                })
            
            df = pd.DataFrame(df_data)
            
            # Create summary sheet
            summary_data = [{
                'Informasi': 'Folder Root',
                'Value': scan_result["root_name"]
            }, {
                'Informasi': 'Root Path',
                'Value': scan_result["root_path"]
            }, {
                'Informasi': 'Waktu Generate',
                'Value': scan_result["scan_time"]
            }, {
                'Informasi': 'Total Records',
                'Value': len(scan_result["items"])
            }, {
                'Informasi': 'Total Files',
                'Value': len([item for item in scan_result["items"] if item['status'] == 'file'])
            }, {
                'Informasi': 'Total Folders',
                'Value': len([item for item in scan_result["items"] if item['status'] == 'folder'])
            }, {
                'Informasi': 'Total Size (MB)',
                'Value': round(sum([item['ukuran_mb'] for item in scan_result["items"]]), 2)
            }, {
                'Informasi': 'Generated By',
                'Value': 'Aplikasi Arsip Digital - Auto Database Generator'
            }, {
                'Informasi': 'Purpose',
                'Value': 'Database for Web Server Access'
            }]
            
            df_summary = pd.DataFrame(summary_data)
            
            # Export to Excel with multiple sheets
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name="Database", index=False)
                df_summary.to_excel(writer, sheet_name="Summary", index=False)
            
            return {
                "success": True,
                "file_path": output_path,
                "total_records": len(df_data)
            }
            
        except Exception as e:
            return {
                "success": False,
                "error": str(e)
            }
    
    def clear_default_folder(self):
        """Hapus folder default"""
        if config_manager.get_default_folder():
            result = messagebox.askyesno(
                "Konfirmasi",
                "Apakah Anda yakin ingin menghapus folder default?\n\n"
                "Anda perlu memilih folder secara manual di setiap form."
            )
            
            if result:
                if config_manager.set_default_folder(""):
                    self.folder_var.set("Belum ada folder default yang dipilih")
                    self.status_var.set("✅ Folder default berhasil dihapus!")
                    
                    # Clear status after 3 seconds
                    self.root.after(3000, lambda: self.status_var.set(""))
                    
                    messagebox.showinfo("Berhasil", "Folder default berhasil dihapus!")
                else:
                    messagebox.showerror("Error", "Gagal menghapus konfigurasi!")
        else:
            messagebox.showinfo("Info", "Tidak ada folder default yang tersimpan.")
    
    def back_to_menu(self):
        """Kembali ke menu utama"""
        if self.parent_window:
            self.root.destroy()
            self.parent_window.deiconify()


class CekNoKKApp:
    """Form untuk Cek NO KK (Nomor Kartu Keluarga)"""
    
    def __init__(self, root, parent_window=None):
        self.root = root
        self.parent_window = parent_window
        
        # Initialize variables
        self.results = []
        self.status_var = tk.StringVar(value="✅ Ready - Klik 'PROSES CEK NO KK' untuk memulai")
        self.is_paused = False
        self.is_processing = False
        
        self.setup_window()
        self.create_widgets()
    
    def setup_window(self):
        """Setup window cek no kk"""
        self.root.title("Cek NO KK - Nomor Kartu Keluarga")
        
        # Get screen dimensions
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        
        # Get responsive dimensions
        width, height, self.padding, self.fonts = get_responsive_dimensions(
            900, 700, screen_width, screen_height
        )
        
        self.root.geometry(f"{width}x{height}")
        self.root.minsize(800, 600)
        self.root.resizable(True, True)
        
        # Center window
        self.center_window()
    
    def center_window(self):
        """Center window di layar"""
        self.root.update_idletasks()
        width = self.root.winfo_width()
        height = self.root.winfo_height()
        x = (self.root.winfo_screenwidth() // 2) - (width // 2)
        y = (self.root.winfo_screenheight() // 2) - (height // 2)
        self.root.geometry(f'{width}x{height}+{x}+{y}')
    
    def create_widgets(self):
        """Membuat widget untuk cek no kk"""
        # Main frame
        main_frame = ttk.Frame(self.root, padding=str(self.padding))
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Configure grid weights
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(4, weight=1)  # Results frame expandable
        
        # Title
        title_label = ttk.Label(
            main_frame, 
            text="👨‍👩‍👧‍👦 CEK NOMOR KK", 
            font=("Arial", self.fonts['title'], "bold")
        )
        title_label.grid(row=0, column=0, pady=(0, 5))
        
        # Subtitle
        subtitle_label = ttk.Label(
            main_frame, 
            text="Validasi dan pengecekan Nomor Kartu Keluarga dari Database",
            font=("Arial", self.fonts['subtitle']),
            foreground="gray"
        )
        subtitle_label.grid(row=1, column=0, pady=(0, 20))
        
        # Info frame
        info_frame = ttk.LabelFrame(main_frame, text="ℹ️ Informasi", padding="15")
        info_frame.grid(row=2, column=0, sticky=(tk.W, tk.E), pady=(0, 15))
        
        database_path = get_database_path()
        info_text = (
            "Proses ini akan:\n\n"
            f"1. Membaca data dari:\n   {database_path}\n"
            "   (Sheet: 02.DATA_ANGGOTA)\n"
            "2. Filter file PDF yang diawali dengan '02'\n"
            "3. Ekstrak NO KK dari PDF menggunakan OCR\n"
            "4. Validasi format NO KK (16 digit angka)\n"
            "5. Menampilkan hasil pengecekan\n"
            "6. Export hasil ke Excel"
        )
        
        info_label = ttk.Label(
            info_frame,
            text=info_text,
            font=("Arial", self.fonts['normal']),
            foreground="blue",
            justify=tk.LEFT
        )
        info_label.grid(row=0, column=0, sticky=tk.W)
        
        # Button frame
        btn_frame = ttk.Frame(main_frame)
        btn_frame.grid(row=3, column=0, pady=(0, 15))
        
        # Proses button (tombol utama)
        self.proses_btn = ttk.Button(
            btn_frame, 
            text="▶️ PROSES CEK NO KK", 
            command=self.proses_cek_nokk,
            style="Accent.TButton"
        )
        self.proses_btn.grid(row=0, column=0, padx=(0, 10), ipadx=20, ipady=10)
        
        # Pause/Resume button
        self.pause_btn = ttk.Button(
            btn_frame, 
            text="⏸️ Pause", 
            command=self.toggle_pause,
            state=tk.DISABLED
        )
        self.pause_btn.grid(row=0, column=1, padx=(10, 10))
        
        # Export button
        self.export_btn = ttk.Button(
            btn_frame, 
            text="💾 Export Hasil", 
            command=self.export_results,
            state=tk.DISABLED
        )
        self.export_btn.grid(row=0, column=2, padx=(10, 10))
        
        # Back button
        if self.parent_window:
            back_btn = ttk.Button(
                btn_frame, 
                text="⬅️ Kembali", 
                command=self.back_to_menu
            )
            back_btn.grid(row=0, column=3, padx=(10, 0))
        
        # Results frame dengan treeview
        results_frame = ttk.LabelFrame(main_frame, text="Hasil Pengecekan NO KK", padding="10")
        results_frame.grid(row=4, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(0, 10))
        results_frame.columnconfigure(0, weight=1)
        results_frame.rowconfigure(0, weight=1)
        
        # Create scrollbars
        tree_scroll_y = ttk.Scrollbar(results_frame, orient=tk.VERTICAL)
        tree_scroll_y.grid(row=0, column=1, sticky=(tk.N, tk.S))
        
        tree_scroll_x = ttk.Scrollbar(results_frame, orient=tk.HORIZONTAL)
        tree_scroll_x.grid(row=1, column=0, sticky=(tk.W, tk.E))
        
        # Treeview untuk hasil
        self.tree = ttk.Treeview(
            results_frame,
            columns=("No", "NO KK", "Status", "Panjang", "Format", "Keterangan", "Nama", "Nomor Center", "Status File", "Path"),
            show="headings",
            yscrollcommand=tree_scroll_y.set,
            xscrollcommand=tree_scroll_x.set
        )
        tree_scroll_y.config(command=self.tree.yview)
        tree_scroll_x.config(command=self.tree.xview)
        
        # Define columns
        self.tree.heading("No", text="No")
        self.tree.heading("NO KK", text="NO KK")
        self.tree.heading("Status", text="Status")
        self.tree.heading("Panjang", text="Panjang")
        self.tree.heading("Format", text="Format")
        self.tree.heading("Keterangan", text="Keterangan")
        self.tree.heading("Nama", text="Nama Anggota")
        self.tree.heading("Nomor Center", text="Nomor Center")
        self.tree.heading("Status File", text="Status File")
        self.tree.heading("Path", text="Path File")
        
        # Set column widths
        self.tree.column("No", width=50, anchor=tk.CENTER)
        self.tree.column("NO KK", width=150, anchor=tk.CENTER)
        self.tree.column("Status", width=80, anchor=tk.CENTER)
        self.tree.column("Panjang", width=70, anchor=tk.CENTER)
        self.tree.column("Format", width=100, anchor=tk.CENTER)
        self.tree.column("Keterangan", width=250, anchor=tk.W)
        self.tree.column("Nama", width=200, anchor=tk.W)
        self.tree.column("Nomor Center", width=120, anchor=tk.CENTER)
        self.tree.column("Status File", width=100, anchor=tk.CENTER)
        self.tree.column("Path", width=400, anchor=tk.W)
        self.tree.column("Nama", width=200, anchor=tk.W)
        self.tree.column("Nomor Center", width=120, anchor=tk.CENTER)
        self.tree.column("Path", width=400, anchor=tk.W)
        
        self.tree.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Status label
        status_label = ttk.Label(
            main_frame,
            textvariable=self.status_var,
            font=("Arial", self.fonts['small']),
          
        )
        status_label.grid(row=5, column=0, sticky=tk.W, pady=(5, 0))
    
    def validate_nokk(self, nokk):
        """Validasi format NO KK"""
        result = {
            "nokk": nokk,
            "valid": False,
            "panjang": len(str(nokk)) if nokk else 0,
            "format": "Invalid",
            "keterangan": ""
        }
        
        # Handle NaN or empty
        if pd.isna(nokk) or nokk == "" or nokk is None:
            result["keterangan"] = "NO KK kosong"
            result["format"] = "Kosong"
            return result
        
        # Convert to string and remove spaces
        nokk_str = str(nokk).strip().replace(" ", "")
        result["nokk"] = nokk_str
        result["panjang"] = len(nokk_str)
        
        # Check panjang
        if len(nokk_str) != 16:
            result["keterangan"] = f"Panjang tidak sesuai (harus 16 digit, saat ini {len(nokk_str)})"
            return result
        
        # Check apakah semua digit
        if not nokk_str.isdigit():
            result["format"] = "Bukan Angka"
            result["keterangan"] = "NO KK harus berisi angka saja"
            return result
        
        # Valid
        result["valid"] = True
        result["format"] = "Valid"
        result["keterangan"] = "Format NO KK valid (16 digit angka)"
        
        return result
    
    def toggle_pause(self):
        """Toggle pause/resume state"""
        self.is_paused = not self.is_paused
        
        if self.is_paused:
            self.pause_btn.config(text="▶️ Resume")
            self.status_var.set("⏸️ PAUSED - Klik 'Resume' untuk melanjutkan")
        else:
            self.pause_btn.config(text="⏸️ Pause")
            self.status_var.set("▶️ RESUMED - Melanjutkan proses...")
    
    def wait_if_paused(self):
        """Wait loop saat paused"""
        while self.is_paused and self.is_processing:
            self.root.update()
            self.root.after(100)  # Check every 100ms
    
    def deskew_image(self, image):
        """Straighten skewed/tilted image menggunakan projection profile"""
        try:
            import numpy as np
            from PIL import Image
            
            # Convert PIL Image to numpy array
            img_array = np.array(image)
            
            # Binarize dengan threshold
            threshold = 128
            binary = img_array < threshold
            
            # Hitung projection profile untuk berbagai sudut (-20 sampai +20 derajat)
            best_angle = 0
            max_variance = 0
            
            for angle in range(-20, 21, 1):
                # Rotate image
                rotated = image.rotate(angle, expand=False, fillcolor=255)
                rotated_array = np.array(rotated.convert('L'))
                rotated_binary = rotated_array < threshold
                
                # Horizontal projection (sum across rows)
                h_projection = np.sum(rotated_binary, axis=1)
                
                # Variance of projection - higher = better alignment
                variance = np.var(h_projection)
                
                if variance > max_variance:
                    max_variance = variance
                    best_angle = angle
            
            # Apply best rotation
            if best_angle != 0:
                print(f"🔄 Deskewing image: {best_angle} degrees")
                deskewed = image.rotate(best_angle, expand=True, fillcolor=255)
                return deskewed
            else:
                print("✅ Image already straight")
                return image
                
        except Exception as e:
            print(f"⚠️ Deskew failed: {str(e)}, using original image")
            return image
    
    def extract_nokk_from_pdf(self, pdf_path):
        """Ekstrak NO KK dari PDF menggunakan OCR dengan fokus ke header"""
        if not os.path.exists(pdf_path):
            return None
        
        try:
            # Auto-detect Tesseract path
            tesseract_paths = [
                r"C:\Program Files\Tesseract-OCR\tesseract.exe",
                r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe",
                r"C:\Tesseract-OCR\tesseract.exe",
            ]
            
            tesseract_found = False
            for tess_path in tesseract_paths:
                if os.path.exists(tess_path):
                    pytesseract.pytesseract.tesseract_cmd = tess_path
                    tesseract_found = True
                    break
            
            if not tesseract_found:
                print("⚠️ Tesseract OCR tidak ditemukan. Install dari: https://github.com/UB-Mannheim/tesseract/wiki")
                return None
            
            # Detect Poppler path (same as PDF Tool)
            poppler_paths = [
                "poppler-25.07.0/Library/bin",  # Portable version
                r"C:\Program Files\poppler\Library\bin",
                r"C:\poppler\Library\bin"
            ]
            
            poppler_path = None
            for path in poppler_paths:
                if os.path.exists(path):
                    poppler_path = path
                    break
            
            # Convert PDF to images (first page only)
            try:
                if poppler_path:
                    images = convert_from_path(
                        pdf_path, 
                        first_page=1, 
                        last_page=1, 
                        poppler_path=poppler_path,
                        dpi=400  # Higher DPI untuk OCR lebih akurat
                    )
                else:
                    images = convert_from_path(
                        pdf_path, 
                        first_page=1, 
                        last_page=1,
                        dpi=400
                    )
            except Exception as e:
                print(f"Error converting PDF to image: {str(e)}")
                return None
            
            if not images:
                return None
            
            # Preprocessing image untuk OCR lebih baik
            from PIL import ImageEnhance, ImageFilter, ImageOps
            import numpy as np
            
            image = images[0]
            width, height = image.size
            
            # Crop hanya bagian header (20% atas) - area NO KK biasanya di sini
            header_height = int(height * 0.2)
            image_header = image.crop((0, 0, width, header_height))
            
            # Convert to grayscale
            image_header = image_header.convert('L')
            
            # DESKEW: Straighten image jika miring
            image_header = self.deskew_image(image_header)
            
            # Enhance contrast lebih kuat
            enhancer = ImageEnhance.Contrast(image_header)
            image_header = enhancer.enhance(3.0)
            
            # Enhance sharpness
            enhancer = ImageEnhance.Sharpness(image_header)
            image_header = enhancer.enhance(2.0)
            
            # Threshold untuk binarisasi (black & white)
            image_header = ImageOps.autocontrast(image_header)
            
            # OCR dengan config optimized untuk angka
            configs = [
                '--psm 6 -c tessedit_char_whitelist=0123456789',  # Hanya angka
                '--psm 11 -c tessedit_char_whitelist=0123456789',
                '--psm 6',  # Tanpa whitelist sebagai fallback
            ]
            
            all_text = ""
            for config in configs:
                try:
                    text = pytesseract.image_to_string(image_header, lang='eng', config=config)
                    all_text += text + "\n"
                except:
                    continue
            
            if not all_text:
                return None
            
            # Debug: print extracted text
            print(f"\n📄 Header text from {os.path.basename(pdf_path)}:")
            print(all_text[:300])
            
            # Clean up common OCR errors before pattern matching
            # Ganti huruf yang mirip angka
            cleaned_text = all_text
            replacements = {
                'b': '6',  # huruf b sering dibaca untuk angka 6
                'B': '8',
                'O': '0',
                'o': '0',
                'l': '1',
                'I': '1',
                'S': '5',
                'Z': '2',
            }
            
            for old, new in replacements.items():
                cleaned_text = cleaned_text.replace(old, new)
            
            # Search for 16-digit number pattern (NO KK)
            # Pattern 1: 16 consecutive digits (dari cleaned text)
            pattern = r'\b\d{16}\b'
            matches = re.findall(pattern, cleaned_text)
            
            if matches:
                # Ambil yang pertama (biasanya NO KK di header)
                print(f"✅ Found NO KK (after cleanup): {matches[0]}")
                return matches[0]
            
            # Pattern 2: From original text (tanpa cleanup)
            matches = re.findall(pattern, all_text)
            if matches:
                print(f"✅ Found NO KK: {matches[0]}")
                return matches[0]
            
            # Pattern 3: With spaces/dots (e.g., "3302 0403 0205 2186")
            pattern_with_space = r'(\d{4}[\s\.\-]?\d{4}[\s\.\-]?\d{4}[\s\.\-]?\d{4})'
            matches = re.findall(pattern_with_space, cleaned_text)
            
            if matches:
                # Remove spaces, dots, dashes
                nokk = re.sub(r'[\s\.\-]', '', matches[0])
                if len(nokk) == 16 and nokk.isdigit():
                    print(f"✅ Found NO KK (with separators): {nokk}")
                    return nokk
            
            # Pattern 4: Cari angka 15-18 digit (kadang OCR salah)
            pattern_flexible = r'\d{15,18}'
            matches = re.findall(pattern_flexible, cleaned_text)
            
            if matches:
                # Coba berbagai kemungkinan untuk mendapat 16 digit
                for match in matches:
                    if len(match) == 16:
                        print(f"✅ Found NO KK (flexible): {match}")
                        return match
                    elif len(match) == 17:
                        # Coba ambil 16 digit pertama atau terakhir
                        candidate1 = match[:16]
                        candidate2 = match[1:]
                        # Prioritas yang dimulai dengan 33 (kode Jawa Tengah)
                        if candidate1.startswith('33'):
                            print(f"✅ Found NO KK (17→16, first): {candidate1}")
                            return candidate1
                        elif candidate2.startswith('33'):
                            print(f"✅ Found NO KK (17→16, last): {candidate2}")
                            return candidate2
                        else:
                            print(f"✅ Found NO KK (17→16): {candidate1}")
                            return candidate1
                    elif len(match) == 15:
                        # Mungkin kurang 1 digit, tapi tetap return
                        print(f"⚠️ Found 15 digits (might be incomplete): {match}")
                        # Don't return, keep searching
            
            print("❌ NO KK tidak ditemukan dalam header")
            return None
            
        except Exception as e:
            print(f"❌ Error extracting NO KK from {pdf_path}: {str(e)}")
            return None
    
    def proses_cek_nokk(self):
        """Proses cek NO KK dari database.xlsx - ekstrak dari PDF file yang dimulai dengan 02"""
        # Set processing flag
        self.is_processing = True
        self.is_paused = False
        
        # Enable pause button, disable proses button
        self.pause_btn.config(state=tk.NORMAL, text="⏸️ Pause")
        self.proses_btn.config(state=tk.DISABLED)
        
        # Check OCR availability
        if not OCR_AVAILABLE:
            messagebox.showerror(
                "OCR Not Available",
                "pytesseract atau pdf2image tidak tersedia!\n\n"
                "Install dengan: pip install pytesseract pdf2image\n\n"
                "Dan pastikan Tesseract OCR sudah terinstall di sistem."
            )
            return
        
        # Clear previous results
        for item in self.tree.get_children():
            self.tree.delete(item)
        self.results = []
        
        # Check if database.xlsx exists di AppData
        database_path = get_database_path()
        if not os.path.exists(database_path):
            messagebox.showerror(
                "File Tidak Ditemukan",
                f"File database.xlsx tidak ditemukan di:\n{database_path}\n\n"
                "Silakan scan folder arsip digital terlebih dahulu:\n"
                "Menu → Scan Folder Arsip Digital\n\n"
                "Setelah scan selesai, pilih 'Simpan dan Sinkronkan'\n"
                "untuk membuat file database.xlsx"
            )
            self.status_var.set("❌ Error: database.xlsx tidak ditemukan")
            return
        
        try:
            self.status_var.set("🔄 Membaca database.xlsx...")
            self.root.update()
            
            # Read Excel - sheet 02.DATA_ANGGOTA dari AppData
            df = pd.read_excel(database_path, sheet_name="02.DATA_ANGGOTA")
            
            # Check required columns
            required_cols = ["TYPE", "NAMA_FILE", "PATH"]
            missing_cols = [col for col in required_cols if col not in df.columns]
            if missing_cols:
                messagebox.showerror(
                    "Kolom Tidak Ditemukan",
                    f"Kolom berikut tidak ditemukan:\n{', '.join(missing_cols)}\n\n"
                    "Pastikan database.xlsx memiliki struktur yang benar."
                )
                self.status_var.set("❌ Error: Struktur database tidak sesuai")
                return
            
            # Filter: TYPE = "FILE" dan NAMA_FILE dimulai dengan "02"
            df_filtered = df[
                (df["TYPE"] == "FILE") & 
                (df["NAMA_FILE"].astype(str).str.startswith("02"))
            ].copy()
            
            total_rows = len(df_filtered)
            
            if total_rows == 0:
                messagebox.showwarning(
                    "Data Kosong",
                    "Tidak ada file yang dimulai dengan '02' di sheet 02.DATA_ANGGOTA!"
                )
                self.status_var.set("⚠️ Warning: Tidak ada data yang sesuai filter")
                return
            
            self.status_var.set(f"🔄 Ditemukan {total_rows} file PDF untuk diproses...")
            self.root.update()
            
            # Get additional columns if available
            id_nama_col = "ID_NAMA_ANGGOTA" if "ID_NAMA_ANGGOTA" in df_filtered.columns else None
            nomor_center_col = "NOMOR_CENTER" if "NOMOR_CENTER" in df_filtered.columns else None
            
            # Process each PDF file
            valid_count = 0
            invalid_count = 0
            not_found_count = 0
            
            for idx, row in df_filtered.iterrows():
                # Check if paused
                self.wait_if_paused()
                
                pdf_path = row["PATH"]
                nama_file = row["NAMA_FILE"]
                id_nama = row[id_nama_col] if id_nama_col else "-"
                nomor_center = row[nomor_center_col] if nomor_center_col else "-"
                
                # Update progress
                current = len(self.results) + 1
                if not self.is_paused:
                    self.status_var.set(f"🔄 Memproses {current}/{total_rows}: {nama_file}...")
                self.root.update()
                
                # Check if file exists
                file_exists = os.path.exists(pdf_path)
                file_status = "✅ Ada" if file_exists else "❌ Tidak Ada"
                
                # Extract NO KK from PDF only if file exists
                nokk = None
                if file_exists:
                    nokk = self.extract_nokk_from_pdf(pdf_path)
                
                if nokk:
                    # Validate
                    result = self.validate_nokk(nokk)
                    result["nama"] = id_nama
                    result["nomor_center"] = nomor_center
                    result["path"] = pdf_path
                    result["file_status"] = file_status
                    
                    # Count
                    if result["valid"]:
                        valid_count += 1
                    else:
                        invalid_count += 1
                elif not file_exists:
                    # File not found
                    not_found_count += 1
                    result = {
                        "nokk": "-",
                        "valid": False,
                        "panjang": 0,
                        "format": "-",
                        "keterangan": "File PDF tidak ditemukan",
                        "nama": id_nama,
                        "nomor_center": nomor_center,
                        "path": pdf_path,
                        "file_status": file_status
                    }
                else:
                    # File exists but NO KK not found in OCR
                    not_found_count += 1
                    result = {
                        "nokk": "-",
                        "valid": False,
                        "panjang": 0,
                        "format": "-",
                        "keterangan": "NO KK tidak ditemukan di PDF",
                        "nama": id_nama,
                        "nomor_center": nomor_center,
                        "path": pdf_path,
                        "file_status": file_status
                    }
                
                # Add to results
                self.results.append(result)
                
                # Add to treeview
                if result["nokk"] == "-":
                    status_icon = "⚠️"
                    tag = "not_found"
                elif result["valid"]:
                    status_icon = "✅"
                    tag = "valid"
                else:
                    status_icon = "❌"
                    tag = "invalid"
                
                self.tree.insert(
                    "", 
                    tk.END, 
                    values=(
                        current,
                        result["nokk"],
                        status_icon,
                        result["panjang"],
                        result["format"],
                        result["keterangan"],
                        id_nama,
                        nomor_center,
                        result.get("file_status", "-"),
                        pdf_path
                    ),
                    tags=(tag,)
                )
            
            # Configure tags untuk warna
            self.tree.tag_configure("valid", foreground="green")
            self.tree.tag_configure("invalid", foreground="red")
            self.tree.tag_configure("not_found", foreground="orange")
            
            # Enable export button
            self.export_btn.config(state=tk.NORMAL)
            
            # Update status
            self.status_var.set(
                f"✅ Selesai: {total_rows} file | ✅ Valid: {valid_count} | ❌ Invalid: {invalid_count} | ⚠️ Tidak Ditemukan: {not_found_count}"
            )
            
            # Show result
            messagebox.showinfo(
                "Proses Selesai",
                f"Pengecekan NO KK dari PDF selesai!\n\n"
                f"Total File PDF: {total_rows}\n"
                f"✅ Valid: {valid_count}\n"
                f"❌ Invalid: {invalid_count}\n"
                f"⚠️ NO KK Tidak Ditemukan: {not_found_count}\n\n"
                f"Klik 'Export Hasil' untuk menyimpan hasil pengecekan."
            )
            
        except Exception as e:
            messagebox.showerror(
                "Error",
                f"Terjadi kesalahan saat memproses data:\n\n{str(e)}"
            )
            self.status_var.set(f"❌ Error: {str(e)}")
        finally:
            # Reset processing state
            self.is_processing = False
            self.is_paused = False
            self.pause_btn.config(state=tk.DISABLED, text="⏸️ Pause")
            self.proses_btn.config(state=tk.NORMAL)
    
    def export_results(self):
        """Export hasil pengecekan ke Excel"""
        if not self.results:
            messagebox.showwarning("Tidak Ada Data", "Tidak ada hasil untuk di-export!")
            return
        
        # Ask for save location
        file_path = filedialog.asksaveasfilename(
            title="Export Hasil Pengecekan NO KK",
            defaultextension=".xlsx",
            initialfile=f"cek_nokk_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            filetypes=[
                ("Excel Files", "*.xlsx"),
                ("All Files", "*.*")
            ]
        )
        
        if not file_path:
            return
        
        try:
            self.status_var.set("🔄 Export ke Excel...")
            self.root.update()
            
            # Prepare data
            df_data = []
            for idx, result in enumerate(self.results, 1):
                df_data.append({
                    'No': idx,
                    'NO_KK': result['nokk'],
                    'Status': 'VALID' if result['valid'] else 'INVALID',
                    'Panjang': result['panjang'],
                    'Format': result['format'],
                    'Keterangan': result['keterangan'],
                    'Nama': result.get('nama', '-'),
                    'Nomor_Center': result.get('nomor_center', '-'),
                    'Status_File': result.get('file_status', '-'),
                    'Path': result.get('path', '-')
                })
            
            df = pd.DataFrame(df_data)
            
            # Summary data
            valid_count = len([r for r in self.results if r['valid']])
            invalid_count = len([r for r in self.results if not r['valid']])
            
            summary_data = [{
                'Informasi': 'Total Data',
                'Value': len(self.results)
            }, {
                'Informasi': 'Valid',
                'Value': valid_count
            }, {
                'Informasi': 'Invalid',
                'Value': invalid_count
            }, {
                'Informasi': 'Persentase Valid',
                'Value': f"{(valid_count/len(self.results)*100):.2f}%" if self.results else "0%"
            }, {
                'Informasi': 'Waktu Export',
                'Value': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            }]
            
            df_summary = pd.DataFrame(summary_data)
            
            # Export to Excel with multiple sheets
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name="Hasil Cek NO KK", index=False)
                df_summary.to_excel(writer, sheet_name="Summary", index=False)
            
            self.status_var.set(f"✅ Export berhasil: {os.path.basename(file_path)}")
            messagebox.showinfo(
                "Export Berhasil",
                f"Hasil pengecekan berhasil di-export!\n\n"
                f"File: {file_path}\n\n"
                f"Total: {len(self.results)} NO KK\n"
                f"✅ Valid: {valid_count}\n"
                f"❌ Invalid: {invalid_count}"
            )
            
        except Exception as e:
            messagebox.showerror("Error", f"Gagal export hasil:\n{str(e)}")
            self.status_var.set(f"❌ Error export")
    
    def back_to_menu(self):
        """Kembali ke menu utama"""
        if self.parent_window:
            self.root.destroy()
            self.parent_window.deiconify()


class CekPengajuanDanaApp:
    """Form untuk Cek Pengajuan Dana dari Surat Keluar"""
    
    def __init__(self, root, parent_window=None):
        self.root = root
        self.parent_window = parent_window
        
        self.setup_window()
        self.create_widgets()
    
    def setup_window(self):
        """Setup window cek pengajuan dana"""
        self.root.title("Cek Pengajuan Dana - Surat Keluar")
        
        # Get screen dimensions
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        
        # Set responsive window size (90% of screen width, 90% of height)
        window_width = int(screen_width * 0.9)
        window_height = int(screen_height * 0.9)
        
        # Minimum size constraints
        min_width = 1200
        min_height = 600
        window_width = max(window_width, min_width)
        window_height = max(window_height, min_height)
        
        self.root.geometry(f"{window_width}x{window_height}")
        self.root.minsize(min_width, min_height)
        self.root.resizable(True, True)
        
        # Center window
        self.center_window()
    
    def center_window(self):
        """Center window di layar"""
        self.root.update_idletasks()
        width = self.root.winfo_width()
        height = self.root.winfo_height()
        x = (self.root.winfo_screenwidth() // 2) - (width // 2)
        y = (self.root.winfo_screenheight() // 2) - (height // 2)
        self.root.geometry(f'{width}x{height}+{x}+{y}')
    
    def create_widgets(self):
        """Membuat widget untuk cek pengajuan dana dengan scrollable canvas"""
        # Main container frame
        container = ttk.Frame(self.root)
        container.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Configure grid weights
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        container.columnconfigure(0, weight=1)
        container.rowconfigure(0, weight=1)
        
        # Create canvas with scrollbar
        canvas = tk.Canvas(container, highlightthickness=0)
        scrollbar = ttk.Scrollbar(container, orient="vertical", command=canvas.yview)
        
        # Scrollable frame inside canvas
        scrollable_frame = ttk.Frame(canvas)
        
        # Configure canvas scroll region
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        # Create window in canvas
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Grid layout
        canvas.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))
        
        # Enable mousewheel scrolling
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        
        canvas.bind_all("<MouseWheel>", _on_mousewheel)
        
        # Main frame
        main_frame = ttk.Frame(scrollable_frame, padding="20")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Configure main_frame
        scrollable_frame.columnconfigure(0, weight=1)
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(5, weight=1)  # Row 5 untuk results_frame (treeview)
        
        # Title
        title_label = ttk.Label(
            main_frame, 
            text="💰 CEK PENGAJUAN DANA", 
            font=("Arial", 16, "bold")
        )
        title_label.grid(row=0, column=0, pady=(0, 10))
        
        # Subtitle
        subtitle_label = ttk.Label(
            main_frame, 
            text="Scan file PENGAJUAN_DANA.xlsm dari folder Surat Keluar",
            font=("Arial", 10),
            foreground="gray"
        )
        subtitle_label.grid(row=1, column=0, pady=(0, 5))
        
        # Info label
        info_label = ttk.Label(
            main_frame,
            text="💡 Tip: Double-click pada baris untuk membuka file Excel",
            font=("Arial", 9, "italic"),
            foreground="blue"
        )
        info_label.grid(row=2, column=0, pady=(0, 15))
        
        # Button frame
        btn_frame = ttk.Frame(main_frame)
        btn_frame.grid(row=3, column=0, pady=(0, 10))
        
        # Scan button
        scan_btn = ttk.Button(
            btn_frame, 
            text="🔍 Mulai Scan", 
            command=self.scan_pengajuan_dana,
            style="Accent.TButton"
        )
        scan_btn.grid(row=0, column=0, padx=(0, 10))
        
        # Analisa button
        self.analisa_btn = ttk.Button(
            btn_frame, 
            text="🔬 Analisa Data", 
            command=self.analisa_data,
            state=tk.DISABLED
        )
        self.analisa_btn.grid(row=0, column=1, padx=(10, 10))
        
        # Export button
        self.export_btn = ttk.Button(
            btn_frame, 
            text="📊 Export ke Excel", 
            command=self.export_to_excel,
            state=tk.DISABLED
        )
        self.export_btn.grid(row=0, column=2, padx=(10, 10))
        
        # Back button
        if self.parent_window:
            back_btn = ttk.Button(
                btn_frame, 
                text="⬅️ Kembali", 
                command=self.back_to_menu
            )
            back_btn.grid(row=0, column=3, padx=(10, 0))
        
        # Status info bar (di atas treeview)
        status_info_frame = ttk.Frame(main_frame)
        status_info_frame.grid(row=4, column=0, sticky=(tk.W, tk.E), pady=(5, 10))
        
        self.status_var = tk.StringVar(value="Siap untuk scan")
        status_label = ttk.Label(
            status_info_frame,
            textvariable=self.status_var,
            font=("Arial", 9),
            foreground="blue"
        )
        status_label.grid(row=0, column=0, sticky=tk.W)
        
        # Results frame dengan treeview
        results_frame = ttk.LabelFrame(main_frame, text="Hasil Scan", padding="10")
        results_frame.grid(row=2, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(0, 15))
        results_frame.columnconfigure(0, weight=1)
        results_frame.rowconfigure(0, weight=1)
        
        # Create scrollbars (vertical dan horizontal)
        tree_scroll_y = ttk.Scrollbar(results_frame, orient=tk.VERTICAL)
        tree_scroll_y.grid(row=0, column=1, sticky=(tk.N, tk.S))
        
        tree_scroll_x = ttk.Scrollbar(results_frame, orient=tk.HORIZONTAL)
        tree_scroll_x.grid(row=1, column=0, sticky=(tk.W, tk.E))
        
        self.tree = ttk.Treeview(
            results_frame,
            columns=("No", "Tahun", "Bulan", "Nomor Surat", "Nama File", "Path"),
            show="headings",
            yscrollcommand=tree_scroll_y.set,
            xscrollcommand=tree_scroll_x.set
        )
        tree_scroll_y.config(command=self.tree.yview)
        tree_scroll_x.config(command=self.tree.xview)
        
        # Define columns
        self.tree.heading("No", text="No")
        self.tree.heading("Tahun", text="Tahun")
        self.tree.heading("Bulan", text="Bulan")
        self.tree.heading("Nomor Surat", text="Nomor Surat")
        self.tree.heading("Nama File", text="Nama File")
        self.tree.heading("Path", text="Path Lengkap")
        
        # Set column widths
        self.tree.column("No", width=50, anchor=tk.CENTER)
        self.tree.column("Tahun", width=80, anchor=tk.CENTER)
        self.tree.column("Bulan", width=100, anchor=tk.CENTER)
        self.tree.column("Nomor Surat", width=100, anchor=tk.CENTER)
        self.tree.column("Nama File", width=250, anchor=tk.W)
        self.tree.column("Path", width=300, anchor=tk.W)
        
        self.tree.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Bind double-click event untuk membuka file
        self.tree.bind("<Double-Button-1>", self.open_selected_file)
        
        # Data storage
        self.scan_results = []
        
        # Update canvas width to match window
        def _configure_canvas(event):
            canvas.itemconfig(canvas.find_withtag("all")[0], width=event.width)
        
        canvas.bind("<Configure>", _configure_canvas)
    
    def scan_pengajuan_dana(self):
        """Scan folder surat keluar untuk file PENGAJUAN_DANA.xlsm"""
        # Clear previous results
        for item in self.tree.get_children():
            self.tree.delete(item)
        self.scan_results = []
        
        # Gunakan default folder dari config
        default_folder = config_manager.get_default_folder()
        
        if not default_folder or not os.path.exists(default_folder):
            messagebox.showwarning(
                "Folder Tidak Ditemukan",
                "Default folder belum diset atau tidak ditemukan.\n\n"
                "Silakan set default folder di menu Pengaturan terlebih dahulu."
            )
            return
        
        # Path ke folder surat keluar
        base_path = os.path.join(default_folder, "01.SURAT_MENYURAT", "02.SURAT_KELUAR")
        
        if not os.path.exists(base_path):
            messagebox.showerror(
                "Folder Tidak Ada",
                f"Folder Surat Keluar tidak ditemukan:\n{base_path}\n\n"
                f"Pastikan struktur folder sudah benar."
            )
            return
        
        self.status_var.set("🔄 Scanning...")
        self.root.update()
        
        # Nama-nama bulan
        bulan_names = {
            "01": "JANUARI", "02": "FEBRUARI", "03": "MARET", "04": "APRIL",
            "05": "MEI", "06": "JUNI", "07": "JULI", "08": "AGUSTUS",
            "09": "SEPTEMBER", "10": "OKTOBER", "11": "NOVEMBER", "12": "DESEMBER"
        }
        
        # Scan dari tahun 2020 sampai tahun sekarang + 1
        current_year = datetime.now().year
        found_count = 0
        
        for year in range(2020, current_year + 2):
            year_folder = os.path.join(base_path, str(year))
            
            if not os.path.exists(year_folder):
                continue
            
            # Loop untuk setiap bulan (01-12)
            for bulan_code, bulan_name in bulan_names.items():
                # Format: 01.JANUARI, 02.FEBRUARI, dst
                bulan_folder_name = f"{bulan_code}.{bulan_name}"
                bulan_folder = os.path.join(year_folder, bulan_folder_name)
                
                if not os.path.exists(bulan_folder):
                    continue
                
                # Scan semua file di folder bulan
                try:
                    for file in os.listdir(bulan_folder):
                        # Skip file temporary (dimulai dengan ~)
                        if file.startswith('~'):
                            continue
                        
                        # Cek apakah file berakhiran PENGAJUAN_DANA.xlsm
                        if file.upper().endswith("PENGAJUAN_DANA.XLSM"):
                            # Extract nomor surat (3 digit di awal)
                            nomor_surat = file[:3] if len(file) >= 3 else "???"
                            
                            file_path = os.path.join(bulan_folder, file)
                            
                            # Simpan hasil (tanpa data analisa dulu)
                            self.scan_results.append({
                                "tahun": year,
                                "bulan": bulan_name,
                                "bulan_code": bulan_code,
                                "nomor_surat": nomor_surat,
                                "nama_file": file,
                                "path": file_path,
                                "nomor_surat_f8": "",  # Akan diisi saat analisa
                                "nominal_input": "",  # Akan diisi saat analisa
                                "nominal_kebutuhan": "",  # Akan diisi saat analisa
                                "status_balance": "",  # Akan diisi saat analisa
                                "tanggal_disburse_awal": "",  # Akan diisi saat analisa
                                "tanggal_disburse_akhir": "",  # Akan diisi saat analisa
                                "nama_bm": "",  # Akan diisi saat analisa
                                "data_analisa": {}  # Untuk data analisa lainnya
                            })
                            
                            found_count += 1
                            
                            # Insert ke treeview
                            self.tree.insert("", tk.END, values=(
                                found_count,
                                year,
                                bulan_name,
                                nomor_surat,
                                file,
                                file_path
                            ))
                except Exception as e:
                    print(f"Error scanning {bulan_folder}: {e}")
        
        # Update status
        if found_count > 0:
            self.status_var.set(f"✅ Ditemukan {found_count} file PENGAJUAN_DANA.xlsm")
            self.export_btn.config(state=tk.NORMAL)
            self.analisa_btn.config(state=tk.NORMAL)
            messagebox.showinfo(
                "Scan Selesai",
                f"Berhasil menemukan {found_count} file PENGAJUAN_DANA.xlsm\n\n"
                f"Klik 'Export ke Excel' untuk menyimpan hasil.\n"
                f"Klik 'Analisa Data' untuk ekstrak data dari dalam file."
            )
        else:
            self.status_var.set("⚠️ Tidak ada file PENGAJUAN_DANA.xlsm ditemukan")
            self.export_btn.config(state=tk.DISABLED)
            self.analisa_btn.config(state=tk.DISABLED)
            messagebox.showinfo(
                "Scan Selesai",
                "Tidak ditemukan file PENGAJUAN_DANA.xlsm\n\n"
                f"Path yang di-scan: {base_path}"
            )
    
    def analisa_data(self):
        """Analisa data dari dalam file PENGAJUAN_DANA.xlsm"""
        if not self.scan_results:
            messagebox.showwarning("Tidak Ada Data", "Tidak ada data untuk dianalisa!")
            return
        
        # Konfirmasi dengan user
        result = messagebox.askyesno(
            "Konfirmasi Analisa",
            f"Akan menganalisa {len(self.scan_results)} file PENGAJUAN_DANA.xlsm\n\n"
            f"Proses ini akan:\n"
            f"• Membaca data dari dalam setiap file\n"
            f"• Mengekstrak nomor surat dari cell F8\n"
            f"• Menambahkan kolom data hasil analisa\n\n"
            f"Lanjutkan?"
        )
        
        if not result:
            return
        
        # Progress dialog
        progress_window = tk.Toplevel(self.root)
        progress_window.title("Analisa Data...")
        progress_window.geometry("400x150")
        progress_window.resizable(False, False)
        
        # Center window
        progress_window.update_idletasks()
        x = (progress_window.winfo_screenwidth() // 2) - 200
        y = (progress_window.winfo_screenheight() // 2) - 75
        progress_window.geometry(f'400x150+{x}+{y}')
        
        progress_window.transient(self.root)
        progress_window.grab_set()
        
        # Progress content
        progress_frame = ttk.Frame(progress_window, padding="20")
        progress_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        progress_label = ttk.Label(
            progress_frame, 
            text="Memproses file...",
            font=("Arial", 10)
        )
        progress_label.grid(row=0, column=0, pady=(0, 10))
        
        progress_bar = ttk.Progressbar(
            progress_frame, 
            mode='determinate',
            maximum=len(self.scan_results)
        )
        progress_bar.grid(row=1, column=0, sticky=(tk.W, tk.E))
        
        status_label = ttk.Label(
            progress_frame,
            text="",
            font=("Arial", 9),
            foreground="gray"
        )
        status_label.grid(row=2, column=0, pady=(10, 0))
        
        progress_window.update()
        
        # Analisa setiap file
        success_count = 0
        error_count = 0
        
        for idx, result in enumerate(self.scan_results):
            file_path = result['path']
            file_name = result['nama_file']
            
            # Update progress
            progress_bar['value'] = idx + 1
            status_label.config(text=f"File {idx+1}/{len(self.scan_results)}: {file_name}")
            progress_window.update()
            
            try:
                # === SHEET SURAT ===
                df_surat = pd.read_excel(file_path, sheet_name='Surat', header=None)
                
                # Ambil Nomor Surat dari cell F8 (row 7, col 5 - zero-indexed)
                nomor_surat_file = df_surat.iloc[7, 5] if len(df_surat) > 7 and len(df_surat.columns) > 5 else None
                
                # Ambil Nominal Input Kebutuhan dari cell I8 (row 7, col 8 - zero-indexed)
                nominal_input = df_surat.iloc[7, 8] if len(df_surat) > 7 and len(df_surat.columns) > 8 else None
                
                # === SHEET LAPORAN ===
                status_balance = None
                nominal_kebutuhan = None
                nama_bm = None
                
                try:
                    df_laporan = pd.read_excel(file_path, sheet_name='Laporan', header=None)
                    
                    # Ambil Status Balance dari cell A4 (row 3, col 0 - zero-indexed)
                    if len(df_laporan) > 3 and len(df_laporan.columns) > 0:
                        cell_a4 = str(df_laporan.iloc[3, 0])
                        if 'Ket.' in cell_a4 and ':' in cell_a4:
                            parts = cell_a4.split(':', 1)
                            if len(parts) > 1:
                                status_balance = parts[1].strip()
                    
                    # Ambil Nominal Kebutuhan dari cell F68 (row 67, col 5 - zero-indexed)
                    if len(df_laporan) > 67 and len(df_laporan.columns) > 5:
                        nominal_kebutuhan = df_laporan.iloc[67, 5]
                    
                    # Ambil Nama BM dari cell A83 (row 82, col 0 - zero-indexed)
                    if len(df_laporan) > 82 and len(df_laporan.columns) > 0:
                        nama_bm = df_laporan.iloc[82, 0]
                
                except Exception as e_laporan:
                    pass  # Jika gagal baca sheet Laporan, set None
                
                # === SHEET LAMPIRAN ===
                tanggal_disburse_awal = None
                tanggal_disburse_akhir = None
                
                try:
                    df_lampiran = pd.read_excel(file_path, sheet_name='Lampiran', header=None)
                    
                    # Ambil Tanggal Disburse Awal dari cell C3 (row 2, col 2 - zero-indexed)
                    if len(df_lampiran) > 2 and len(df_lampiran.columns) > 2:
                        tanggal_disburse_awal = df_lampiran.iloc[2, 2]
                    
                    # Ambil Tanggal Disburse Akhir dari cell E3 (row 2, col 4 - zero-indexed)
                    if len(df_lampiran) > 2 and len(df_lampiran.columns) > 4:
                        tanggal_disburse_akhir = df_lampiran.iloc[2, 4]
                
                except Exception as e_lampiran:
                    pass  # Jika gagal baca sheet Lampiran, set None
                
                # Simpan hasil analisa
                result['nomor_surat_file'] = nomor_surat_file
                result['nominal_input'] = nominal_input
                result['status_balance'] = status_balance
                result['nominal_kebutuhan'] = nominal_kebutuhan
                result['tanggal_disburse_awal'] = tanggal_disburse_awal
                result['tanggal_disburse_akhir'] = tanggal_disburse_akhir
                result['nama_bm'] = nama_bm
                result['status_analisa'] = 'SUCCESS'
                success_count += 1
                
            except Exception as e:
                result['nomor_surat_file'] = None
                result['nominal_input'] = None
                result['status_balance'] = None
                result['nominal_kebutuhan'] = None
                result['tanggal_disburse_awal'] = None
                result['tanggal_disburse_akhir'] = None
                result['nama_bm'] = None
                result['status_analisa'] = f'ERROR: {str(e)}'
                error_count += 1
        
        progress_window.destroy()
        
        # Update treeview dengan menambah kolom
        # Clear dan rebuild treeview
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        # Update columns untuk include data analisa (14 kolom total)
        self.tree['columns'] = ("No", "Tahun", "Bulan", "Nomor Surat", "Nomor di File", 
                                "Nominal Input", "Nominal Kebutuhan", "Status Balance", 
                                "Tgl Disburse Awal", "Tgl Disburse Akhir", "Nama BM", 
                                "Status", "Nama File", "Path")
        
        # Redefine headings
        self.tree.heading("No", text="No")
        self.tree.heading("Tahun", text="Tahun")
        self.tree.heading("Bulan", text="Bulan")
        self.tree.heading("Nomor Surat", text="No. Surat (Nama)")
        self.tree.heading("Nomor di File", text="No. Surat (F8)")
        self.tree.heading("Nominal Input", text="Nominal Input")
        self.tree.heading("Nominal Kebutuhan", text="Nominal Kebutuhan")
        self.tree.heading("Status Balance", text="Status Balance")
        self.tree.heading("Tgl Disburse Awal", text="Tgl Disburse Awal")
        self.tree.heading("Tgl Disburse Akhir", text="Tgl Disburse Akhir")
        self.tree.heading("Nama BM", text="Nama BM")
        self.tree.heading("Status", text="Status")
        self.tree.heading("Nama File", text="Nama File")
        self.tree.heading("Path", text="Path Lengkap")
        
        # Set column widths
        self.tree.column("No", width=50, anchor=tk.CENTER)
        self.tree.column("Tahun", width=70, anchor=tk.CENTER)
        self.tree.column("Bulan", width=90, anchor=tk.CENTER)
        self.tree.column("Nomor Surat", width=100, anchor=tk.CENTER)
        self.tree.column("Nomor di File", width=100, anchor=tk.CENTER)
        self.tree.column("Nominal Input", width=120, anchor=tk.E)
        self.tree.column("Nominal Kebutuhan", width=120, anchor=tk.E)
        self.tree.column("Status Balance", width=100, anchor=tk.CENTER)
        self.tree.column("Tgl Disburse Awal", width=120, anchor=tk.CENTER)
        self.tree.column("Tgl Disburse Akhir", width=120, anchor=tk.CENTER)
        self.tree.column("Nama BM", width=150, anchor=tk.W)
        self.tree.column("Status", width=80, anchor=tk.CENTER)
        self.tree.column("Nama File", width=200, anchor=tk.W)
        self.tree.column("Path", width=250, anchor=tk.W)
        
        # Repopulate treeview
        for idx, result in enumerate(self.scan_results, 1):
            nomor_file = result.get('nomor_surat_file', '')
            nominal_input = result.get('nominal_input', '')
            nominal_kebutuhan = result.get('nominal_kebutuhan', '')
            status_balance = result.get('status_balance', '')
            tgl_disburse_awal = result.get('tanggal_disburse_awal', '')
            tgl_disburse_akhir = result.get('tanggal_disburse_akhir', '')
            nama_bm = result.get('nama_bm', '')
            status = result.get('status_analisa', '')
            
            # Tentukan status display
            if status == 'SUCCESS':
                status_display = "✅"
            else:
                status_display = "❌"
            
            # Format tanggal jika ada
            tgl_awal_str = str(tgl_disburse_awal) if tgl_disburse_awal else '-'
            tgl_akhir_str = str(tgl_disburse_akhir) if tgl_disburse_akhir else '-'
            
            self.tree.insert("", tk.END, values=(
                idx,
                result['tahun'],
                result['bulan'],
                result['nomor_surat'],
                nomor_file if nomor_file else '-',
                nominal_input if nominal_input else '-',
                nominal_kebutuhan if nominal_kebutuhan else '-',
                status_balance if status_balance else '-',
                tgl_awal_str,
                tgl_akhir_str,
                nama_bm if nama_bm else '-',
                status_display,
                result['nama_file'],
                result['path']
            ))
        
        # Show result
        messagebox.showinfo(
            "Analisa Selesai",
            f"Analisa data selesai!\n\n"
            f"✅ Berhasil: {success_count} file\n"
            f"❌ Error: {error_count} file\n\n"
            f"Data yang diekstrak:\n"
            f"• Nomor Surat (F8) - Sheet Surat\n"
            f"• Nominal Input (I8) - Sheet Surat\n"
            f"• Nominal Kebutuhan (F68) - Sheet Laporan\n"
            f"• Status Balance (A4) - Sheet Laporan\n"
            f"• Tanggal Disburse Awal (C3) - Sheet Lampiran\n"
            f"• Tanggal Disburse Akhir (E3) - Sheet Lampiran\n"
            f"• Nama BM (A83) - Sheet Laporan"
        )
        
        self.status_var.set(f"✅ Analisa selesai: {success_count} sukses, {error_count} error")
    
    def export_to_excel(self):
        """Export hasil scan ke Excel"""
        if not self.scan_results:
            messagebox.showwarning("Tidak Ada Data", "Tidak ada data untuk di-export!")
            return
        
        # Ask for save location
        file_path = filedialog.asksaveasfilename(
            title="Export Hasil Scan Pengajuan Dana",
            defaultextension=".xlsx",
            initialfile=f"pengajuan_dana_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            filetypes=[
                ("Excel Files", "*.xlsx"),
                ("All Files", "*.*")
            ]
        )
        
        if not file_path:
            return
        
        try:
            # Create DataFrame
            df = pd.DataFrame(self.scan_results)
            
            # Check apakah sudah ada data analisa
            has_analisa = 'nomor_surat_file' in df.columns
            
            if has_analisa:
                # Include semua kolom analisa
                df = df[["tahun", "bulan", "bulan_code", "nomor_surat", "nomor_surat_file",
                        "nominal_input", "nominal_kebutuhan", "status_balance", 
                        "tanggal_disburse_awal", "tanggal_disburse_akhir", "nama_bm",
                        "status_analisa", "nama_file", "path"]]
                
                # Rename columns untuk export
                df.columns = ["Tahun", "Bulan", "Kode Bulan", "Nomor Surat (Nama File)", 
                             "Nomor Surat (F8)", "Nominal Input Kebutuhan (I8)", 
                             "Nominal Kebutuhan (F68)", "Status Balance (A4)", 
                             "Tanggal Disburse Awal (C3)", "Tanggal Disburse Akhir (E3)", 
                             "Nama BM (A83)", "Status Analisa", "Nama File", "Path Lengkap"]
            else:
                # Tanpa kolom analisa (export biasa)
                df = df[["tahun", "bulan", "bulan_code", "nomor_surat", "nama_file", "path"]]
                
                # Rename columns untuk export
                df.columns = ["Tahun", "Bulan", "Kode Bulan", "Nomor Surat", "Nama File", "Path Lengkap"]
            
            # Export ke Excel
            df.to_excel(file_path, index=False, sheet_name="Pengajuan Dana")
            
            info_msg = f"Data berhasil di-export!\n\n"
            info_msg += f"File: {os.path.basename(file_path)}\n"
            info_msg += f"Total: {len(self.scan_results)} rows"
            
            if has_analisa:
                info_msg += f"\n\nIncludes data analisa lengkap:\n"
                info_msg += f"• Nomor Surat (F8) - Sheet Surat\n"
                info_msg += f"• Nominal Input (I8) - Sheet Surat\n"
                info_msg += f"• Nominal Kebutuhan (F68) - Sheet Laporan\n"
                info_msg += f"• Status Balance (A4) - Sheet Laporan\n"
                info_msg += f"• Tanggal Disburse Awal (C3) - Sheet Lampiran\n"
                info_msg += f"• Tanggal Disburse Akhir (E3) - Sheet Lampiran\n"
                info_msg += f"• Nama BM (A83) - Sheet Laporan\n"
                info_msg += f"• Status analisa"
            
            messagebox.showinfo("Export Berhasil", info_msg)
            
        except Exception as e:
            messagebox.showerror("Export Gagal", f"Gagal export ke Excel:\n{str(e)}")
    
    def open_selected_file(self, event):
        """Buka file Excel yang dipilih dengan double-click"""
        # Ambil item yang dipilih
        selected_item = self.tree.selection()
        
        if not selected_item:
            return
        
        # Ambil values dari item yang dipilih
        item_values = self.tree.item(selected_item[0], "values")
        
        if not item_values:
            return
        
        # Path ada di kolom terakhir
        # Sebelum analisa: 6 kolom, path di index 5
        # Setelah analisa: 9 kolom, path di index 8
        file_path = item_values[-1]  # Ambil kolom terakhir (Path)
        
        # Validasi file exists
        if not os.path.exists(file_path):
            messagebox.showerror(
                "File Tidak Ditemukan",
                f"File tidak ditemukan:\n{file_path}"
            )
            return
        
        # Buka file dengan aplikasi default
        try:
            os.startfile(file_path)
        except Exception as e:
            messagebox.showerror(
                "Gagal Membuka File",
                f"Gagal membuka file:\n{str(e)}\n\n"
                f"Path: {file_path}"
            )
    
    def back_to_menu(self):
        """Kembali ke menu utama"""
        if self.parent_window:
            self.root.destroy()
            self.parent_window.deiconify()


class PDFToolApp:
    """Simple PDF Tool window with common utilities: merge, split, images<>pdf, compress.

    Note: This class uses PyPDF2 for PDF manipulation and Pillow for image handling.
    If libraries are missing, the UI will show instructions.
    """

    def __init__(self, root, parent_window=None):
        self.root = root
        self.parent_window = parent_window

        # Try to import optional dependencies lazily
        try:
            from PyPDF2 import PdfReader, PdfWriter
            self.PdfReader = PdfReader
            self.PdfWriter = PdfWriter
        except Exception as e:
            self.PdfReader = None
            self.PdfWriter = None
            print(f"Warning: PyPDF2 import failed - {e}")

        try:
            from PIL import Image
            self.PIL_Image = Image
        except Exception as e:
            self.PIL_Image = None
            print(f"Warning: Pillow import failed - {e}")

        self.setup_window()
        self.create_widgets()

    def setup_window(self):
        self.root.title("PDF Tool - Merge / Split / Convert")
        w, h = 600, 600
        self.root.geometry(f"{w}x{h}")
        self.root.minsize(500, 550)
        self.center_window()

    def center_window(self):
        self.root.update_idletasks()
        width = self.root.winfo_width()
        height = self.root.winfo_height()
        x = (self.root.winfo_screenwidth() // 2) - (width // 2)
        y = (self.root.winfo_screenheight() // 2) - (height // 2)
        self.root.geometry(f'{width}x{height}+{x}+{y}')

    def create_widgets(self):
        main = ttk.Frame(self.root, padding="20")
        main.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)

        title = ttk.Label(main, text="📄 PDF TOOL", font=("Arial", 20, "bold"))
        title.grid(row=0, column=0, sticky=tk.W, pady=(0, 5))

        subtitle = ttk.Label(main, text="Pilihan: Merge, Split, Convert Images ↔ PDF, Compress",
                             font=("Arial", 11), foreground="gray")
        subtitle.grid(row=1, column=0, sticky=tk.W, pady=(0, 20))

        # Frame untuk buttons dengan grid 2 kolom
        btn_frame = ttk.Frame(main)
        btn_frame.grid(row=2, column=0, sticky=(tk.W, tk.E))
        btn_frame.columnconfigure(0, weight=1)
        btn_frame.columnconfigure(1, weight=1)

        # Style untuk button yang lebih besar
        button_style = {
            'width': 25,
            'padding': 15
        }

        # Row 0 - Images → PDF (PRIORITAS UTAMA)
        img2pdf_btn = ttk.Button(btn_frame, text="�️ Images → PDF", command=self.images_to_pdf, **button_style)
        img2pdf_btn.grid(row=0, column=0, padx=5, pady=8, sticky=(tk.W, tk.E))

        # Row 0 - Merge PDFs
        merge_btn = ttk.Button(btn_frame, text="🔗 Merge PDFs", command=self.merge_pdfs, **button_style)
        merge_btn.grid(row=0, column=1, padx=5, pady=8, sticky=(tk.W, tk.E))

        # Row 1 - Split PDF
        split_btn = ttk.Button(btn_frame, text="✂️ Split PDF", command=self.split_pdf, **button_style)
        split_btn.grid(row=1, column=0, padx=5, pady=8, sticky=(tk.W, tk.E))

        # Row 1 - Compress PDF
        compress_btn = ttk.Button(btn_frame, text="�️ Compress PDF", command=self.compress_pdf, **button_style)
        compress_btn.grid(row=1, column=1, padx=5, pady=8, sticky=(tk.W, tk.E))

        # Row 2 - PDF → Images (PALING AKHIR)
        pdf2img_btn = ttk.Button(btn_frame, text="�️ PDF → Images", command=self.pdf_to_images, **button_style)
        pdf2img_btn.grid(row=2, column=0, padx=5, pady=8, sticky=(tk.W, tk.E))

        # Status and info
        self.status_var = tk.StringVar(value="✅ Ready - Pilih operasi PDF yang ingin dilakukan")
        status_label = ttk.Label(main, textvariable=self.status_var, font=("Arial", 10), foreground="blue")
        status_label.grid(row=3, column=0, sticky=tk.W, pady=(20, 0))

        footer = ttk.Frame(main)
        footer.grid(row=10, column=0, sticky=(tk.W, tk.E), pady=(20, 0))

        back_btn = ttk.Button(footer, text="⬅️ Kembali ke Menu", command=self.back_to_menu)
        back_btn.grid(row=0, column=0)

    def _ensure_pypdf(self):
        if not self.PdfReader or not self.PdfWriter:
            messagebox.showerror(
                "Dependency Missing",
                "PyPDF2 is required for PDF operations.\nInstall with: pip install PyPDF2"
            )
            return False
        return True

    def _ensure_pillow(self):
        if not self.PIL_Image:
            messagebox.showerror(
                "Dependency Missing",
                "Pillow is required for image operations.\nInstall with: pip install Pillow"
            )
            return False
        return True

    def merge_pdfs(self):
        if not self._ensure_pypdf():
            return

        paths = filedialog.askopenfilenames(
            title="Pilih file PDF untuk di-merge (bisa pilih multiple)", 
            filetypes=[("PDF Files", "*.pdf")]
        )
        if not paths or len(paths) == 0:
            return

        out_path = filedialog.asksaveasfilename(
            title="Simpan hasil merge sebagai", 
            defaultextension=".pdf", 
            filetypes=[("PDF Files", "*.pdf")]
        )
        if not out_path:
            return

        try:
            self.status_var.set(f"🔄 Menggabungkan {len(paths)} file PDF...")
            self.root.update()
            
            merger = self.PdfWriter()
            
            for idx, pdf_path in enumerate(paths, 1):
                self.status_var.set(f"🔄 Memproses file {idx}/{len(paths)}: {os.path.basename(pdf_path)}")
                self.root.update()
                
                with open(pdf_path, 'rb') as pdf_file:
                    reader = self.PdfReader(pdf_file)
                    for page in reader.pages:
                        merger.add_page(page)

            with open(out_path, 'wb') as output_file:
                merger.write(output_file)

            self.status_var.set(f"✅ Merge selesai: {os.path.basename(out_path)}")
            messagebox.showinfo(
                "Selesai", 
                f"✅ Berhasil menggabungkan {len(paths)} file PDF!\n\n"
                f"Output: {out_path}"
            )
        except Exception as e:
            self.status_var.set("❌ Error saat merge PDF")
            messagebox.showerror("Error", f"Gagal merge PDFs:\n\n{str(e)}")

    def split_pdf(self):
        if not self._ensure_pypdf():
            return

        path = filedialog.askopenfilename(
            title="Pilih file PDF untuk di-split", 
            filetypes=[("PDF Files", "*.pdf")]
        )
        if not path:
            return

        out_dir = filedialog.askdirectory(title="Pilih folder tujuan untuk menyimpan halaman")
        if not out_dir:
            return

        try:
            self.status_var.set("🔄 Membaca PDF...")
            self.root.update()
            
            with open(path, 'rb') as pdf_file:
                reader = self.PdfReader(pdf_file)
                total_pages = len(reader.pages)
                
                base_name = os.path.splitext(os.path.basename(path))[0]
                
                for i, page in enumerate(reader.pages, start=1):
                    self.status_var.set(f"🔄 Memproses halaman {i}/{total_pages}...")
                    self.root.update()
                    
                    writer = self.PdfWriter()
                    writer.add_page(page)
                    
                    out_file = os.path.join(out_dir, f"{base_name}_halaman_{i}.pdf")
                    with open(out_file, 'wb') as output_file:
                        writer.write(output_file)

            self.status_var.set(f"✅ Split selesai: {total_pages} halaman")
            messagebox.showinfo(
                "Selesai", 
                f"✅ PDF berhasil di-split menjadi {total_pages} file!\n\n"
                f"Lokasi: {out_dir}"
            )
        except Exception as e:
            self.status_var.set("❌ Error saat split PDF")
            messagebox.showerror("Error", f"Gagal split PDF:\n\n{str(e)}")

    def pdf_to_images(self):
        """Convert PDF pages to images with portable Poppler support"""
        # Check if pdf2image is available
        try:
            from pdf2image import convert_from_path
        except ImportError:
            messagebox.showerror(
                "Library Tidak Ditemukan",
                "Fitur PDF → Images membutuhkan library 'pdf2image'.\n\n"
                "Install dengan perintah:\n"
                "pip install pdf2image"
            )
            return
        
        # Select PDF file
        pdf_path = filedialog.askopenfilename(
            title="Pilih file PDF untuk dikonversi ke gambar",
            filetypes=[("PDF Files", "*.pdf")]
        )
        if not pdf_path:
            return
        
        # Select output folder
        out_dir = filedialog.askdirectory(title="Pilih folder tujuan untuk menyimpan gambar")
        if not out_dir:
            return
        
        # Ask for image format
        format_choice = messagebox.askquestion(
            "Format Gambar",
            "Simpan sebagai PNG?\n\n"
            "Yes = PNG (kualitas tinggi)\n"
            "No = JPG (ukuran lebih kecil)"
        )
        img_format = "PNG" if format_choice == "yes" else "JPEG"
        
        try:
            # Show progress
            self.status_var.set("🔄 Mengkonversi PDF ke gambar...")
            self.root.update()
            
            # Get Poppler path from config
            poppler_path = config_manager.config.get("poppler_path", "")
            
            # Try to convert with different methods
            images = None
            error_detail = ""
            
            # Method 1: Try with saved poppler path
            if poppler_path and os.path.exists(poppler_path):
                try:
                    images = convert_from_path(pdf_path, dpi=200, poppler_path=poppler_path)
                except Exception as e:
                    error_detail = f"Gagal dengan saved path: {e}"
            
            # Method 2: Try without poppler_path (auto-detect from PATH)
            if images is None:
                try:
                    images = convert_from_path(pdf_path, dpi=200)
                except Exception as e:
                    error_detail += f"\nGagal auto-detect: {e}"
            
            # Method 3: Check common portable locations
            if images is None:
                common_paths = [
                    # Struktur poppler-xx.xx.x (dari release GitHub)
                    os.path.join(os.getcwd(), "poppler-25.07.0", "Library", "bin"),
                    os.path.join(os.getcwd(), "poppler-25.07.0", "bin"),
                    os.path.join(os.path.dirname(os.path.abspath(__file__)), "poppler-25.07.0", "Library", "bin"),
                    os.path.join(os.path.dirname(os.path.abspath(__file__)), "poppler-25.07.0", "bin"),
                    # Struktur folder poppler generik
                    os.path.join(os.getcwd(), "poppler", "Library", "bin"),
                    os.path.join(os.getcwd(), "poppler", "bin"),
                    os.path.join(os.path.dirname(os.path.abspath(__file__)), "poppler", "Library", "bin"),
                    os.path.join(os.path.dirname(os.path.abspath(__file__)), "poppler", "bin")
                ]
                
                for path in common_paths:
                    if os.path.exists(path):
                        try:
                            images = convert_from_path(pdf_path, dpi=200, poppler_path=path)
                            # Save successful path
                            config_manager.config["poppler_path"] = path
                            config_manager.save_config()
                            break
                        except Exception:
                            continue
            
            # If still failed, ask user to locate Poppler
            if images is None:
                result = messagebox.askyesno(
                    "Poppler Tidak Ditemukan",
                    "Poppler tidak ditemukan!\n\n"
                    "Poppler diperlukan untuk konversi PDF ke gambar.\n"
                    "File poppler bisa diletakkan di folder 'poppler' dalam project ini (portable).\n\n"
                    "Download Poppler dari:\n"
                    "https://github.com/oschwartz10612/poppler-windows/releases/\n\n"
                    "Apakah Anda ingin memilih lokasi folder Poppler sekarang?\n"
                    "(Pilih folder 'Library\\bin' atau 'bin' dari hasil extract Poppler)"
                )
                
                if result:
                    selected_path = filedialog.askdirectory(
                        title="Pilih folder bin Poppler (contoh: poppler/Library/bin)"
                    )
                    
                    if selected_path and os.path.exists(selected_path):
                        try:
                            images = convert_from_path(pdf_path, dpi=200, poppler_path=selected_path)
                            # Save path to config
                            config_manager.config["poppler_path"] = selected_path
                            config_manager.save_config()
                            messagebox.showinfo(
                                "Berhasil",
                                f"Path Poppler berhasil disimpan!\n\n{selected_path}\n\n"
                                "Selanjutnya tidak perlu pilih lagi."
                            )
                        except Exception as e:
                            messagebox.showerror(
                                "Error",
                                f"Folder Poppler tidak valid!\n\n{e}\n\n"
                                "Pastikan memilih folder 'bin' atau 'Library/bin' dari Poppler."
                            )
                            return
                    else:
                        return
                else:
                    return
            
            # Save images
            if images:
                base_name = os.path.splitext(os.path.basename(pdf_path))[0]
                for i, image in enumerate(images, start=1):
                    self.status_var.set(f"🔄 Menyimpan gambar {i}/{len(images)}...")
                    self.root.update()
                    
                    ext = "png" if img_format == "PNG" else "jpg"
                    out_file = os.path.join(out_dir, f"{base_name}_halaman_{i}.{ext}")
                    image.save(out_file, img_format)
            
            self.status_var.set(f"✅ PDF → Images selesai: {len(images)} halaman")
            messagebox.showinfo(
                "Selesai",
                f"PDF berhasil dikonversi menjadi {len(images)} gambar di:\n{out_dir}"
            )
        except Exception as e:
            messagebox.showerror("Error", f"Gagal konversi PDF ke gambar:\n{e}")

    def images_to_pdf(self):
        if not self._ensure_pillow():
            return

        paths = filedialog.askopenfilenames(
            title="Pilih gambar untuk digabung jadi PDF (urutan sesuai pilihan)", 
            filetypes=[
                ("Image Files", "*.png *.jpg *.jpeg *.bmp *.tiff *.gif"),
                ("PNG Files", "*.png"),
                ("JPEG Files", "*.jpg *.jpeg"),
                ("All Files", "*.*")
            ]
        )
        if not paths or len(paths) == 0:
            return

        out_path = filedialog.asksaveasfilename(
            title="Simpan hasil PDF sebagai", 
            defaultextension=".pdf", 
            filetypes=[("PDF Files", "*.pdf")]
        )
        if not out_path:
            return

        try:
            self.status_var.set(f"🔄 Memproses {len(paths)} gambar...")
            self.root.update()
            
            images = []
            for idx, img_path in enumerate(paths, 1):
                self.status_var.set(f"🔄 Memproses gambar {idx}/{len(paths)}: {os.path.basename(img_path)}")
                self.root.update()
                
                img = self.PIL_Image.open(img_path)
                
                # Convert RGBA to RGB (PDF doesn't support transparency)
                if img.mode == 'RGBA':
                    # Create white background
                    background = self.PIL_Image.new('RGB', img.size, (255, 255, 255))
                    background.paste(img, mask=img.split()[3])  # Use alpha channel as mask
                    img = background
                elif img.mode != 'RGB':
                    img = img.convert('RGB')
                
                images.append(img)

            # Save as PDF
            if images:
                self.status_var.set("🔄 Menyimpan PDF...")
                self.root.update()
                
                images[0].save(
                    out_path, 
                    "PDF", 
                    resolution=100.0, 
                    save_all=True, 
                    append_images=images[1:] if len(images) > 1 else []
                )

            self.status_var.set(f"✅ Images → PDF selesai: {len(paths)} gambar")
            messagebox.showinfo(
                "Selesai", 
                f"✅ PDF berhasil dibuat dari {len(paths)} gambar!\n\n"
                f"Output: {out_path}"
            )
        except Exception as e:
            self.status_var.set("❌ Error saat convert images ke PDF")
            messagebox.showerror("Error", f"Gagal menggabungkan images menjadi PDF:\n\n{str(e)}")

    def compress_pdf(self):
        if not self._ensure_pypdf():
            return

        path = filedialog.askopenfilename(
            title="Pilih PDF untuk di-compress", 
            filetypes=[("PDF Files", "*.pdf")]
        )
        if not path:
            return

        out_path = filedialog.asksaveasfilename(
            title="Simpan hasil compress sebagai", 
            defaultextension=".pdf", 
            filetypes=[("PDF Files", "*.pdf")]
        )
        if not out_path:
            return

        try:
            self.status_var.set("🔄 Membaca dan menganalisis PDF...")
            self.root.update()
            
            with open(path, 'rb') as pdf_file:
                reader = self.PdfReader(pdf_file)
                writer = self.PdfWriter()
                
                total_pages = len(reader.pages)
                
                # Copy pages and compress content streams
                for i, page in enumerate(reader.pages, 1):
                    self.status_var.set(f"🔄 Memproses halaman {i}/{total_pages}...")
                    self.root.update()
                    
                    # Compress content streams BEFORE adding to writer
                    try:
                        page.compress_content_streams()
                    except Exception:
                        pass
                    
                    writer.add_page(page)

                # Transfer minimal metadata
                if reader.metadata:
                    try:
                        # Only copy essential metadata
                        essential_meta = {}
                        for key in ['/Title', '/Author', '/Subject']:
                            if key in reader.metadata:
                                essential_meta[key] = reader.metadata[key]
                        if essential_meta:
                            writer.add_metadata(essential_meta)
                    except Exception:
                        pass
                
                # Remove duplicate objects and compress
                self.status_var.set("🔄 Mengoptimalkan dan menghapus duplikasi...")
                self.root.update()
                
                # Write with compression
                with open(out_path, 'wb') as output_file:
                    writer.write(output_file)

            # Get file sizes
            original_size = os.path.getsize(path)
            compressed_size = os.path.getsize(out_path)
            reduction = ((original_size - compressed_size) / original_size * 100)

            self.status_var.set(f"✅ Optimasi selesai: {os.path.basename(out_path)}")
            
            # Format sizes
            def format_size(size_bytes):
                if size_bytes >= 1024*1024:
                    return f"{size_bytes / (1024*1024):.2f} MB"
                else:
                    return f"{size_bytes / 1024:.1f} KB"
            
            if reduction > 0:
                size_info = (
                    f"✅ PDF berhasil dioptimasi!\n\n"
                    f"Ukuran awal: {format_size(original_size)}\n"
                    f"Ukuran akhir: {format_size(compressed_size)}\n"
                    f"Pengurangan: {reduction:.1f}%\n\n"
                    f"💾 File disimpan di:\n{out_path}"
                )
            else:
                # File malah bertambah atau sama
                increase = abs(reduction)
                size_info = (
                    f"⚠️ PDF telah dioptimasi\n\n"
                    f"Ukuran awal: {format_size(original_size)}\n"
                    f"Ukuran akhir: {format_size(compressed_size)}\n"
                    f"Perubahan: +{increase:.1f}%\n\n"
                    f"ℹ️ Catatan: PDF ini sudah teroptimasi atau\n"
                    f"mengandung konten yang tidak bisa dikompresi lebih lanjut.\n\n"
                    f"💾 File disimpan di:\n{out_path}"
                )
            
            messagebox.showinfo("Selesai", size_info)
        except Exception as e:
            self.status_var.set("❌ Error saat compress PDF")
            messagebox.showerror("Error", f"Gagal melakukan compress PDF:\n\n{str(e)}")

    def back_to_menu(self):
        if self.parent_window:
            self.root.destroy()
            self.parent_window.deiconify()

class ArsipDigitalApp:
    def __init__(self, root, parent_window=None):
        self.root = root
        self.parent_window = parent_window
        
        # Initialize business logic processors
        self.arsip_processor = ArsipProcessor()
        self.file_manager = FileManager()
        
        self.setup_window()
        self.create_widgets()
        
        # Variables untuk menyimpan path yang dipilih
        self.selected_folder = ""
        self.selected_file = ""
    
    def setup_window(self):
        """Setup window utama aplikasi"""
        self.root.title("Aplikasi Arsip Digital - Halaman Awal")
        
        # Get screen dimensions for responsive design
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        
        # Calculate appropriate window size (responsive)
        if screen_width >= 1920:  # Large screens
            window_width, window_height = 600, 600
        elif screen_width >= 1366:  # Medium screens (laptop)
            window_width, window_height = 580, 580
        elif screen_width >= 1024:  # Small laptop
            window_width, window_height = 520, 550
        else:  # Very small screens
            window_width, window_height = 480, 500
        
        # Ensure window doesn't exceed 85% of screen size
        max_width = int(screen_width * 0.85)
        max_height = int(screen_height * 0.85)
        window_width = min(window_width, max_width)
        window_height = min(window_height, max_height)
        
        self.root.geometry(f"{window_width}x{window_height}")
        self.root.resizable(True, True)
        
        # Set minimum size to prevent too small windows
        self.root.minsize(450, 400)
        
        # Center window
        self.center_window()
        
        # Set window icon (optional)
        try:
            self.root.iconbitmap("icon.ico")  # Jika ada file icon
        except:
            pass
    
    def center_window(self):
        """Center window di layar"""
        self.root.update_idletasks()
        width = self.root.winfo_width()
        height = self.root.winfo_height()
        x = (self.root.winfo_screenwidth() // 2) - (width // 2)
        y = (self.root.winfo_screenheight() // 2) - (height // 2)
        self.root.geometry(f'{width}x{height}+{x}+{y}')
    
    def create_widgets(self):
        """Membuat semua widget GUI"""
        # Main frame dengan padding responsif
        screen_width = self.root.winfo_screenwidth()
        padding_size = 20 if screen_width >= 1366 else 15 if screen_width >= 1024 else 12
        
        main_frame = ttk.Frame(self.root, padding=str(padding_size))
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Configure grid weights
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(0, weight=1)
        
        # Title dengan ukuran font responsif
        title_size = 16 if screen_width >= 1366 else 14 if screen_width >= 1024 else 12
        title_label = ttk.Label(
            main_frame, 
            text="CEK ARSIP DIGITAL", 
            font=("Arial", title_size, "bold")
        )
        title_label.grid(row=0, column=0, pady=(0, 30))
        
        # Subtitle
        subtitle_size = 10 if screen_width >= 1366 else 9 if screen_width >= 1024 else 8
        subtitle_label = ttk.Label(
            main_frame, 
            text="Cocokkan data folder dengan database Excel anggota",
            font=("Arial", subtitle_size)
        )
        subtitle_label.grid(row=1, column=0, pady=(0, 20))
        
        # Frame untuk folder selection dengan padding responsif
        frame_padding = 15 if screen_width >= 1366 else 12 if screen_width >= 1024 else 10
        folder_frame = ttk.LabelFrame(main_frame, text="Pilih Folder Data Anggota", padding=str(frame_padding))
        folder_frame.grid(row=2, column=0, sticky=(tk.W, tk.E), pady=(0, 15))
        folder_frame.columnconfigure(0, weight=1)
        
        # Folder path display dengan wraplength responsif
        wrap_length = 500 if screen_width >= 1366 else 400 if screen_width >= 1024 else 350
        self.folder_var = tk.StringVar(value="Belum ada folder yang dipilih...")
        folder_path_label = ttk.Label(
            folder_frame, 
            textvariable=self.folder_var,
            foreground="gray",
            wraplength=wrap_length
        )
        folder_path_label.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        
        # Folder browse button
        folder_btn = ttk.Button(
            folder_frame, 
            text="Browse Folder", 
            command=self.browse_folder
        )
        folder_btn.grid(row=1, column=0)
        
        # Frame untuk file selection
        file_frame = ttk.LabelFrame(main_frame, text="Pilih File Excel Database Anggota", padding="15")
        file_frame.grid(row=3, column=0, sticky=(tk.W, tk.E), pady=(0, 15))
        file_frame.columnconfigure(0, weight=1)
        
        # File path display
        self.file_var = tk.StringVar(value="Belum ada file yang dipilih...")
        file_path_label = ttk.Label(
            file_frame, 
            textvariable=self.file_var,
            foreground="gray",
            wraplength=600
        )
        file_path_label.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        
        # File browse button
        file_btn = ttk.Button(
            file_frame, 
            text="Browse File", 
            command=self.browse_file
        )
        file_btn.grid(row=1, column=0)
        
        # Frame untuk informasi yang dipilih
        info_frame = ttk.LabelFrame(main_frame, text="Informasi Pilihan", padding="15")
        info_frame.grid(row=4, column=0, sticky=(tk.W, tk.E), pady=(0, 20))
        info_frame.columnconfigure(1, weight=1)
        
        # Info labels
        ttk.Label(info_frame, text="Folder:").grid(row=0, column=0, sticky=tk.W, pady=2)
        self.info_folder_var = tk.StringVar(value="-")
        ttk.Label(info_frame, textvariable=self.info_folder_var, foreground="blue").grid(
            row=0, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=2
        )
        
        ttk.Label(info_frame, text="File:").grid(row=1, column=0, sticky=tk.W, pady=2)
        self.info_file_var = tk.StringVar(value="-")
        ttk.Label(info_frame, textvariable=self.info_file_var, foreground="blue").grid(
            row=1, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=2
        )
        
        # Action buttons frame
        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=5, column=0, pady=(10, 0))
        
        # Process button
        process_btn = ttk.Button(
            button_frame, 
            text="Proses Arsip", 
            command=self.process_archive,
            style="Accent.TButton"
        )
        process_btn.grid(row=0, column=0, padx=(0, 10))
        
        # Reset button
        reset_btn = ttk.Button(
            button_frame, 
            text="Reset", 
            command=self.reset_selections
        )
        reset_btn.grid(row=0, column=1, padx=(10, 0))
        
        # Back to menu button (if parent window exists)
        if self.parent_window:
            back_btn = ttk.Button(
                button_frame, 
                text="Kembali ke Menu", 
                command=self.back_to_menu
            )
            back_btn.grid(row=0, column=2, padx=(10, 0))
        
    
    def browse_folder(self):
        """Fungsi untuk memilih folder"""
        # Gunakan default folder jika ada
        default_folder = config_manager.get_default_folder()
        initial_dir = default_folder if default_folder and os.path.exists(default_folder) else os.getcwd()
        
        folder_path = filedialog.askdirectory(
            title="Pilih Folder Arsip Digital",
            initialdir=initial_dir
        )
        
        if folder_path:
            # Validasi folder menggunakan business logic
            if self.file_manager.validate_folder_path(folder_path):
                self.selected_folder = folder_path
                # Update display dengan path yang dipilih
                self.folder_var.set(folder_path)
                self.info_folder_var.set(os.path.basename(folder_path))
                
                # Reset file selection jika folder berubah
                if self.selected_file and not self.file_manager.is_file_in_folder(self.selected_file, folder_path):
                    self.reset_file_selection()
            else:
                messagebox.showerror("Error", "Folder yang dipilih tidak valid atau tidak dapat diakses!")
                self.selected_folder = ""
    
    def browse_file(self):
        """Fungsi untuk memilih file Excel database anggota"""
        # Tentukan initial directory
        initial_dir = self.selected_folder if self.selected_folder else os.getcwd()
        
        file_path = filedialog.askopenfilename(
            title="Pilih File Excel Database Anggota (Header: B3-Y3)",
            initialdir=initial_dir,
            filetypes=[
                ("Excel Files", "*.xlsx;*.xls"),
                ("All Files", "*.*")
            ]
        )
        
        if file_path:
            # Validasi file menggunakan business logic
            if self.file_manager.validate_file_path(file_path):
                self.selected_file = file_path
                # Update display dengan path yang dipilih
                self.file_var.set(file_path)
                self.info_file_var.set(os.path.basename(file_path))
            else:
                messagebox.showerror("Error", "File yang dipilih tidak valid atau tidak dapat diakses!")
                self.selected_file = ""
    
    def reset_file_selection(self):
        """Reset pilihan file"""
        self.selected_file = ""
        self.file_var.set("Belum ada file yang dipilih...")
        self.info_file_var.set("-")
    
    def reset_selections(self):
        """Reset semua pilihan"""
        self.selected_folder = ""
        self.selected_file = ""
        
        self.folder_var.set("Belum ada folder yang dipilih...")
        self.file_var.set("Belum ada file yang dipilih...")
        self.info_folder_var.set("-")
        self.info_file_var.set("-")
        
        messagebox.showinfo("Reset", "Semua pilihan telah direset!")
    
    def process_archive(self):
        """Fungsi untuk memproses arsip - mencocokkan folder scan dengan database Excel"""
        if not self.selected_folder:
            messagebox.showwarning("Peringatan", "Silakan pilih folder terlebih dahulu!")
            return
        
        if not self.selected_file:
            messagebox.showwarning("Peringatan", "Silakan pilih file Excel database terlebih dahulu!")
            return
        
        # Validasi folder
        if not self.file_manager.validate_folder_path(self.selected_folder):
            messagebox.showerror("Error", "Folder yang dipilih tidak valid atau tidak dapat diakses!")
            return
        
        # Validasi file Excel
        if not self.file_manager.validate_file_path(self.selected_file):
            messagebox.showerror("Error", "File Excel yang dipilih tidak valid atau tidak dapat diakses!")
            return
        
        try:
            # Initialize AnggotaFolderReader
            anggota_reader = AnggotaFolderReader()
            
            # Tentukan jenis scan berdasarkan struktur folder
            folder_name = os.path.basename(self.selected_folder)
            
            # Progress dialog
            progress_window = self.show_progress_dialog("Memproses arsip folder dan database Excel...")
            
            result = None
            scan_type = None
            
            # Cek apakah ini folder anggota (6digit_nama)
            if anggota_reader.validate_anggota_folder(folder_name):
                result = anggota_reader.scan_anggota_folder(self.selected_folder)
                scan_type = "anggota"
            # Cek apakah ini folder center (4digit)
            elif anggota_reader.validate_center_folder(folder_name):
                result = anggota_reader.scan_center_folder(self.selected_folder)
                scan_type = "center"
            # Jika tidak sesuai pattern, coba scan sebagai root
            else:
                result = anggota_reader.scan_data_anggota_root(self.selected_folder)
                scan_type = "root"
            
            if not result or not result.get("success", False):
                progress_window.destroy()
                error_msg = result.get("error", "Unknown error") if result else "Gagal melakukan scan folder"
                messagebox.showerror("Error Scan", f"Gagal memproses folder:\n{error_msg}")
                return
            
            # Baca file Excel database
            try:
                df_database = pd.read_excel(self.selected_file, header=2, usecols="B:Y", skiprows=[3, 4])
                
                # Clean data: Hapus baris kosong jika masih ada
                df_database = df_database.dropna(how='all')  # Hapus baris yang semua kolomnya kosong
                df_database = df_database.reset_index(drop=True)  # Reset index setelah menghapus baris
                
                # Validasi bahwa ini file Excel dengan format yang benar
                if df_database.empty:
                    progress_window.destroy()
                    messagebox.showerror("Error", "File Excel kosong atau format tidak sesuai!")
                    return
            except Exception as e:
                progress_window.destroy()
                messagebox.showerror("Error", f"Gagal membaca file Excel:\n{str(e)}")
                return
            
            progress_window.destroy()
            
            # Tampilkan hasil matching
            self.match_and_export(result, scan_type, df_database, anggota_reader)
                
        except Exception as e:
            if 'progress_window' in locals():
                progress_window.destroy()
            messagebox.showerror("Error", f"Terjadi kesalahan saat memproses arsip:\n{str(e)}")
    
    def match_and_export(self, scan_result, scan_type, df_database, anggota_reader):
        """Menampilkan preview matching dan pilihan export"""
        # Window preview
        preview_window = tk.Toplevel(self.root)
        preview_window.title("Preview Pencocokan Data")
        preview_window.geometry("900x700")
        preview_window.resizable(True, True)
        
        # Center window
        preview_window.update_idletasks()
        width = preview_window.winfo_width()
        height = preview_window.winfo_height()
        x = (preview_window.winfo_screenwidth() // 2) - (width // 2)
        y = (preview_window.winfo_screenheight() // 2) - (height // 2)
        preview_window.geometry(f'{width}x{height}+{x}+{y}')
        
        # Main frame
        main_frame = ttk.Frame(preview_window, padding="15")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        preview_window.columnconfigure(0, weight=1)
        preview_window.rowconfigure(0, weight=1)
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(1, weight=1)
        
        # Title
        title_label = ttk.Label(main_frame, text="🔍 PREVIEW PENCOCOKAN DATA", font=("Arial", 14, "bold"))
        title_label.grid(row=0, column=0, pady=(0, 15))
        
        # Info text area dengan scrollbar
        text_frame = ttk.Frame(main_frame)
        text_frame.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        text_frame.columnconfigure(0, weight=1)
        text_frame.rowconfigure(0, weight=1)
        
        text_widget = tk.Text(text_frame, wrap=tk.WORD, font=("Consolas", 10))
        scrollbar = ttk.Scrollbar(text_frame, orient="vertical", command=text_widget.yview)
        text_widget.configure(yscrollcommand=scrollbar.set)
        
        text_widget.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))
        
        # Generate preview info
        preview_info = []
        preview_info.append("=" * 80)
        preview_info.append("INFORMASI DATABASE EXCEL")
        preview_info.append("=" * 80)
        preview_info.append(f"File: {os.path.basename(self.selected_file)}")
        preview_info.append(f"Total Records: {len(df_database)}")
        preview_info.append(f"Columns: {', '.join(df_database.columns[:5])}...")
        preview_info.append("")
        
        preview_info.append("=" * 80)
        preview_info.append("INFORMASI SCAN FOLDER")
        preview_info.append("=" * 80)
        
        total_anggota_scanned = 0
        if scan_type == "anggota":
            total_anggota_scanned = 1
            preview_info.append(f"Tipe Scan: Single Anggota")
            preview_info.append(f"ID Anggota: {scan_result['anggota_info']['id']}")
        elif scan_type == "center":
            total_anggota_scanned = len(scan_result['anggota_folders'])
            preview_info.append(f"Tipe Scan: Center")
            preview_info.append(f"Kode Center: {scan_result['center_info']['code']}")
            preview_info.append(f"Total Anggota di-scan: {total_anggota_scanned}")
        else:  # root
            total_anggota_scanned = scan_result['root_info']['total_anggota']
            preview_info.append(f"Tipe Scan: Root (Multi-Center)")
            preview_info.append(f"Total Center: {scan_result['root_info']['total_centers']}")
            preview_info.append(f"Total Anggota di-scan: {total_anggota_scanned}")
        
        preview_info.append("")
        preview_info.append("=" * 80)
        preview_info.append("PILIHAN EXPORT")
        preview_info.append("=" * 80)
        export_path = get_export_path()
        preview_info.append(f"1. Export ke {export_path}")
        preview_info.append("   - Sheet 'databaseanggota': Data dari Excel database")
        preview_info.append("   - Sheet 'hasilscan': Data hasil scan folder (dengan ADA/TIDAK ADA)")
        preview_info.append(f"   - Lokasi: {export_path}")
        preview_info.append("")
        preview_info.append("2. Export Hanya Hasil Scan (Custom nama file)")
        preview_info.append("   - Hanya data dari hasil scan folder")
        preview_info.append("   - Format sederhana dengan kolom ADA/TIDAK ADA")
        preview_info.append("")
        
        text_widget.insert(tk.END, "\n".join(preview_info))
        text_widget.config(state=tk.DISABLED)
        
        # Button frame
        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=2, column=0, pady=(15, 0))
        
        # Export combined button
        export_combined_btn = ttk.Button(
            button_frame, 
            text="Lanjutkan Proses", 
            command=lambda: self.export_combined_data(scan_result, scan_type, df_database, preview_window)
        )
        export_combined_btn.grid(row=0, column=0, padx=(0, 10))
        
        # Export scan only button
        export_scan_btn = ttk.Button(
            button_frame, 
            text="📁 Export Hanya Scan", 
            command=lambda: self.export_scan_only(scan_result, scan_type, preview_window)
        )
        export_scan_btn.grid(row=0, column=1, padx=(10, 10))
        
        # Close button
        close_btn = ttk.Button(button_frame, text="Tutup", command=preview_window.destroy)
        close_btn.grid(row=0, column=2, padx=(10, 0))
    
    def export_combined_data(self, scan_result, scan_type, df_database, parent_window):
        """Export data database dan hasil scan ke file_export.xlsx dengan 2 sheet, lalu analisa dan buat sheet matching"""
        try:
            # Gunakan path AppData untuk export
            file_path = get_export_path()
            
            # Generate data scan
            scan_data = self.generate_scan_data(scan_result, scan_type)
            df_scan = pd.DataFrame(scan_data)
            
            # Export kedua data ke Excel dengan 2 sheet
            with pd.ExcelWriter(file_path, engine='openpyxl', mode='w') as writer:
                df_database.to_excel(writer, sheet_name='databaseanggota', index=False)
                df_scan.to_excel(writer, sheet_name='hasilscan', index=False)
            
            parent_window.destroy()
            
            # Analisa dan matching data berdasarkan ID
            self.analyze_and_match_data(df_database, df_scan)
            
        except Exception as e:
            messagebox.showerror("Export Gagal", f"Gagal export data:\n{str(e)}")
    
    def analyze_and_match_data(self, df_database, df_scan):
        """Analisa dan matching data berdasarkan kombinasi Center + ID"""
        try:
            # Buat salinan dataframe untuk normalisasi
            df_db_normalized = df_database.copy()
            df_scan_normalized = df_scan.copy()
            
            # Fungsi normalisasi ID yang lebih robust
            def normalize_id(id_value):
                """Normalisasi ID ke format 6 digit string"""
                try:
                    if pd.isna(id_value):
                        return ''
                    # Convert ke string dan hapus whitespace
                    id_str = str(id_value).strip()
                    # Hapus .0 jika ada (dari float)
                    if '.' in id_str:
                        id_str = id_str.split('.')[0]
                    # Convert ke int lalu ke 6 digit string
                    return str(int(float(id_str))).zfill(6)
                except:
                    return ''
            
            # Fungsi normalisasi center code ke format 4 digit
            def normalize_center(center_value):
                """Normalisasi center ke format 4 digit string"""
                try:
                    if pd.isna(center_value):
                        return ''
                    # Convert ke string dan hapus whitespace
                    center_str = str(center_value).strip()
                    # Hapus .0 jika ada (dari float)
                    if '.' in center_str:
                        center_str = center_str.split('.')[0]
                    # Convert ke int lalu ke 4 digit string
                    return str(int(float(center_str))).zfill(4)
                except:
                    return ''
            
            # Matching berdasarkan kombinasi Center + Sort_ID/ID
            print("=== Matching berdasarkan Center + Sort_ID (database) dengan center + id_anggota (scan) ===")
            
            # Normalisasi Center di kedua dataframe
            # Database: cari kolom "Center"
            if 'Center' in df_database.columns:
                df_db_normalized['center_normalized'] = df_db_normalized['Center'].apply(normalize_center)
            else:
                print("⚠️ Kolom 'Center' tidak ditemukan di database!")
                df_db_normalized['center_normalized'] = ''
            
            # Scan: kolom "center" sudah ada
            df_scan_normalized['center_normalized'] = df_scan_normalized['center'].apply(normalize_center)
            
            # Normalisasi ID
            if 'Sort_ID' in df_database.columns:
                df_db_normalized['id_normalized'] = df_db_normalized['Sort_ID'].apply(normalize_id)
            elif 'No' in df_database.columns:
                df_db_normalized['id_normalized'] = df_db_normalized['No'].apply(normalize_id)
            else:
                db_id_column = df_database.columns[0]
                df_db_normalized['id_normalized'] = df_db_normalized[db_id_column].apply(normalize_id)
            
            df_scan_normalized['id_normalized'] = df_scan_normalized['id_anggota'].apply(normalize_id)
            
            # Buat composite key: Center + ID
            df_db_normalized['composite_key'] = df_db_normalized['center_normalized'] + '_' + df_db_normalized['id_normalized']
            df_scan_normalized['composite_key'] = df_scan_normalized['center_normalized'] + '_' + df_scan_normalized['id_normalized']
            
            # Hapus baris dengan key kosong atau tidak valid
            df_db_normalized = df_db_normalized[
                (df_db_normalized['center_normalized'] != '') & 
                (df_db_normalized['id_normalized'] != '')
            ]
            df_scan_normalized = df_scan_normalized[
                (df_scan_normalized['center_normalized'] != '') & 
                (df_scan_normalized['id_normalized'] != '')
            ]
            
            # Debug: Lihat beberapa composite key
            print("=== DEBUG COMPOSITE KEY ===")
            print(f"Database - Sample Keys: {df_db_normalized['composite_key'].head(10).tolist()}")
            print(f"Scan - Sample Keys: {df_scan_normalized['composite_key'].head(10).tolist()}")
            print(f"Database unique keys: {len(df_db_normalized['composite_key'].unique())}")
            print(f"Scan unique keys: {len(df_scan_normalized['composite_key'].unique())}")
            
            # Cari key yang ada di kedua dataset
            db_keys = set(df_db_normalized['composite_key'].unique())
            scan_keys = set(df_scan_normalized['composite_key'].unique())
            matching_keys = db_keys.intersection(scan_keys)
            
            print(f"Matching keys found: {len(matching_keys)}")
            print(f"Sample matching keys: {list(matching_keys)[:10]}")
            
            # Merge berdasarkan composite key
            df_matched = pd.merge(
                df_scan_normalized,
                df_db_normalized,
                left_on='composite_key',
                right_on='composite_key',
                how='inner',
                suffixes=('_scan', '_db')
            )
            
            print(f"Final matched rows: {len(df_matched)}")
            print("=== END DEBUG ===")
            
            if len(df_matched) > 0:
                # Hitung statistik matching
                total_db = len(df_database)
                total_scan = len(df_scan)
                total_matched = len(df_matched)
                only_db = total_db - total_matched
                only_scan = total_scan - total_matched
                
                # Cari data yang ada di database tapi TIDAK ada di hasil scan (belum diarsip)
                # Gunakan composite key yang sudah dibuat
                matched_keys = set(df_matched['composite_key']) if 'composite_key' in df_matched.columns else set()
                
                # Filter database: ambil yang composite_key-nya TIDAK ada di matched_keys
                if 'composite_key' in df_db_normalized.columns:
                    df_belum_diarsip = df_db_normalized[~df_db_normalized['composite_key'].isin(matched_keys)].copy()
                    # Hapus kolom helper dari df_belum_diarsip
                    cols_to_drop = ['composite_key', 'center_normalized', 'id_normalized']
                    for col in cols_to_drop:
                        if col in df_belum_diarsip.columns:
                            df_belum_diarsip = df_belum_diarsip.drop(columns=[col])
                else:
                    # Fallback jika tidak ada composite_key
                    df_belum_diarsip = df_database.head(0)  # Empty dataframe
                
                # Hapus kolom helper dari df_matched (setelah digunakan untuk filter)
                columns_to_drop = ['composite_key', 'center_normalized_scan', 'center_normalized_db', 
                                 'id_normalized_scan', 'id_normalized_db']
                for col in columns_to_drop:
                    if col in df_matched.columns:
                        df_matched = df_matched.drop(columns=[col])
                
                total_belum_diarsip = len(df_belum_diarsip)
                
                # Ada data yang match, tanyakan user untuk save as
                result = messagebox.askyesno(
                    "Data Matching Ditemukan",
                    f"📊 STATISTIK MATCHING:\n\n"
                    f"✅ Data yang MATCH: {total_matched} rows\n"
                    f"   (ID ada di database DAN hasil scan)\n\n"
                    f"⚠️  Belum Diarsip: {total_belum_diarsip} rows\n"
                    f"   (ID ada di database tapi TIDAK ada di hasil scan)\n\n"
                    f"📄 Total di database: {total_db} rows\n"
                    f"📁 Total di hasil scan: {total_scan} rows\n\n"
                    f"Apakah Anda ingin menyimpan file baru dengan 4 sheet?\n\n"
                    f"Sheet yang akan dibuat:\n"
                    f"1. databaseanggota (semua data database)\n"
                    f"2. hasilscan (semua data scan)\n"
                    f"3. datamatching ({total_matched} data yang match)\n"
                    f"4. belumdiarsip ({total_belum_diarsip} belum ada arsip)"
                )
                
                if result:
                    # Dialog Save As
                    new_file_path = filedialog.asksaveasfilename(
                        title="Simpan File dengan Data Matching",
                        defaultextension=".xlsx",
                        initialfile=f"data_matching_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                        filetypes=[
                            ("Excel Files", "*.xlsx"),
                            ("All Files", "*.*")
                        ]
                    )
                    
                    if new_file_path:
                        # Export ke file baru dengan 4 sheet
                        with pd.ExcelWriter(new_file_path, engine='openpyxl', mode='w') as writer:
                            df_database.to_excel(writer, sheet_name='databaseanggota', index=False)
                            df_scan.to_excel(writer, sheet_name='hasilscan', index=False)
                            df_matched.to_excel(writer, sheet_name='datamatching', index=False)
                            df_belum_diarsip.to_excel(writer, sheet_name='belumdiarsip', index=False)
                        
                        messagebox.showinfo(
                            "Export Berhasil",
                            f"File berhasil disimpan dengan data lengkap!\n\n"
                            f"File: {new_file_path}\n\n"
                            f"Sheet 1: databaseanggota ({len(df_database)} rows)\n"
                            f"Sheet 2: hasilscan ({len(df_scan)} rows)\n"
                            f"Sheet 3: datamatching ({len(df_matched)} rows)\n"
                            f"Sheet 4: belumdiarsip ({len(df_belum_diarsip)} rows)\n\n"
                            f"✅ datamatching = Data yang sudah diarsip\n"
                            f"⚠️  belumdiarsip = Data yang belum ada arsipnya"
                        )
                else:
                    # User tidak ingin save, tampilkan info saja
                    export_path = get_export_path()
                    messagebox.showinfo(
                        "Export Selesai",
                        f"Data berhasil di-export ke:\n{export_path}\n\n"
                        f"Sheet 1: databaseanggota ({len(df_database)} rows)\n"
                        f"Sheet 2: hasilscan ({len(df_scan)} rows)\n\n"
                        f"Ditemukan {len(df_matched)} data yang match (tidak disimpan)."
                    )
            else:
                # Tidak ada data yang match
                export_path = get_export_path()
                messagebox.showinfo(
                    "Export Selesai",
                    f"Data berhasil di-export ke:\n{export_path}\n\n"
                    f"Sheet 1: databaseanggota ({len(df_database)} rows)\n"
                    f"Sheet 2: hasilscan ({len(df_scan)} rows)\n\n"
                    f"⚠️ Tidak ada data yang match berdasarkan ID."
                )
                
        except Exception as e:
            messagebox.showerror("Error Matching", f"Gagal melakukan matching data:\n{str(e)}")
    
    def export_scan_only(self, scan_result, scan_type, parent_window):
        """Export hanya hasil scan ke Excel format sederhana"""
        try:
            file_path = filedialog.asksaveasfilename(
                title="Export Hasil Scan",
                defaultextension=".xlsx",
                initialfile=f"hasil_scan_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                filetypes=[
                    ("Excel Files", "*.xlsx"),
                    ("All Files", "*.*")
                ]
            )
            
            if not file_path:
                return
            
            export_result = self.export_simple_excel_with_ya_tidak(scan_result, scan_type, file_path)
            
            if export_result.get("success", False):
                parent_window.destroy()
                messagebox.showinfo(
                    "Export Berhasil",
                    f"Data scan berhasil di-export!\n\n"
                    f"File: {export_result['file_path']}\n"
                    f"Rows: {export_result['rows_exported']}\n"
                    f"Waktu: {export_result['timestamp']}"
                )
            else:
                messagebox.showerror("Export Gagal", f"Gagal export:\n{export_result.get('error', 'Unknown error')}")
                
        except Exception as e:
            messagebox.showerror("Export Gagal", f"Gagal export hasil scan:\n{str(e)}")
    
    def generate_scan_data(self, scan_result, scan_type):
        """Generate list of dictionaries dari hasil scan untuk DataFrame"""
        data_rows = []
        
        if scan_type == "anggota":
            anggota_info = scan_result["anggota_info"]
            file_categories = scan_result["file_categories"]
            
            row = {
                "center": anggota_info.get("center_code", ""),
                "anggota_folder": anggota_info["folder_name"],
                "id_anggota": anggota_info["id"],
                "nama": self.extract_nama_from_folder(anggota_info["folder_name"]),
                "file_ditemukan": scan_result["file_summary"]["total_files"]
            }
            
            for i in range(1, 13):
                code = f"{i:02d}"
                files = file_categories.get(code, [])
                row[f"file_{code}"] = "ADA" if len(files) > 0 else "TIDAK ADA"
            
            data_rows.append(row)
            
        elif scan_type == "center":
            center_code = scan_result["center_info"]["code"]
            
            for anggota in scan_result["anggota_folders"]:
                anggota_info = anggota["anggota_info"]
                file_categories = anggota["file_categories"]
                
                row = {
                    "center": center_code,
                    "anggota_folder": anggota_info["folder_name"],
                    "id_anggota": anggota_info["id"],
                    "nama": self.extract_nama_from_folder(anggota_info["folder_name"]),
                    "file_ditemukan": anggota["file_summary"]["total_files"]
                }
                
                for i in range(1, 13):
                    code = f"{i:02d}"
                    files = file_categories.get(code, [])
                    row[f"file_{code}"] = "ADA" if len(files) > 0 else "TIDAK ADA"
                
                data_rows.append(row)
                
        elif scan_type == "root":
            for center in scan_result["center_folders"]:
                center_code = center["center_info"]["code"]
                
                for anggota in center["anggota_folders"]:
                    anggota_info = anggota["anggota_info"]
                    file_categories = anggota["file_categories"]
                    
                    row = {
                        "center": center_code,
                        "anggota_folder": anggota_info["folder_name"],
                        "id_anggota": anggota_info["id"],
                        "nama": self.extract_nama_from_folder(anggota_info["folder_name"]),
                        "file_ditemukan": anggota["file_summary"]["total_files"]
                    }
                    
                    for i in range(1, 13):
                        code = f"{i:02d}"
                        files = file_categories.get(code, [])
                        row[f"file_{code}"] = "ADA" if len(files) > 0 else "TIDAK ADA"
                    
                    data_rows.append(row)
        
        return data_rows
    
    def show_progress_dialog(self, message):
        """Menampilkan dialog progress"""
        progress_window = tk.Toplevel(self.root)
        progress_window.title("Processing...")
        progress_window.geometry("300x100")
        progress_window.resizable(False, False)
        
        # Center window
        progress_window.update_idletasks()
        x = (progress_window.winfo_screenwidth() // 2) - (150)
        y = (progress_window.winfo_screenheight() // 2) - (50)
        progress_window.geometry(f'300x100+{x}+{y}')
        
        # Make it modal
        progress_window.transient(self.root)
        progress_window.grab_set()
        
        # Progress content
        frame = ttk.Frame(progress_window, padding="20")
        frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        ttk.Label(frame, text=message, font=("Arial", 10)).grid(row=0, column=0, pady=(0, 10))
        
        progress_bar = ttk.Progressbar(frame, mode='indeterminate')
        progress_bar.grid(row=1, column=0, sticky=(tk.W, tk.E))
        progress_bar.start()
        
        progress_window.update()
        return progress_window
    
    def show_scan_result(self, result, scan_type, anggota_reader):
        """Menampilkan hasil scan dalam window terpisah"""
        result_window = tk.Toplevel(self.root)
        result_window.title("Hasil Proses Arsip Digital")
        result_window.geometry("800x600")
        result_window.resizable(True, True)
        
        # Center window
        result_window.update_idletasks()
        width = result_window.winfo_width()
        height = result_window.winfo_height()
        x = (result_window.winfo_screenwidth() // 2) - (width // 2)
        y = (result_window.winfo_screenheight() // 2) - (height // 2)
        result_window.geometry(f'{width}x{height}+{x}+{y}')
        
        # Main frame
        main_frame = ttk.Frame(result_window, padding="15")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        result_window.columnconfigure(0, weight=1)
        result_window.rowconfigure(0, weight=1)
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(1, weight=1)
        
        # Title
        title_label = ttk.Label(main_frame, text="📊 HASIL PROSES ARSIP DIGITAL", font=("Arial", 14, "bold"))
        title_label.grid(row=0, column=0, pady=(0, 15))
        
        # Results text area dengan scrollbar
        text_frame = ttk.Frame(main_frame)
        text_frame.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        text_frame.columnconfigure(0, weight=1)
        text_frame.rowconfigure(0, weight=1)
        
        text_widget = tk.Text(text_frame, wrap=tk.WORD, font=("Consolas", 10))
        scrollbar = ttk.Scrollbar(text_frame, orient="vertical", command=text_widget.yview)
        text_widget.configure(yscrollcommand=scrollbar.set)
        
        text_widget.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))
        
        # Generate report berdasarkan scan type
        if scan_type == "anggota":
            report = anggota_reader.generate_anggota_report(result)
        elif scan_type == "center":
            report = self.generate_center_report(result)
        else:  # root
            report = self.generate_root_report(result)
        
        # Insert report
        text_widget.insert(tk.END, report)
        text_widget.config(state=tk.DISABLED)
        
        # Button frame
        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=2, column=0, pady=(15, 0))
        
        # Export Excel button
        export_excel_btn = ttk.Button(
            button_frame, 
            text="📊 Export ke Excel", 
            command=lambda: self.export_scan_to_excel(result, scan_type, anggota_reader)
        )
        export_excel_btn.grid(row=0, column=0, padx=(0, 10))
        
        # Export Text button
        export_text_btn = ttk.Button(
            button_frame, 
            text="💾 Export ke Text", 
            command=lambda: self.export_scan_to_text(report)
        )
        export_text_btn.grid(row=0, column=1, padx=(10, 10))
        
        # Close button
        close_btn = ttk.Button(button_frame, text="Tutup", command=result_window.destroy)
        close_btn.grid(row=0, column=2, padx=(10, 0))
    
    def generate_center_report(self, result):
        """Generate laporan untuk scan center"""
        report = []
        report.append("=" * 70)
        report.append("LAPORAN SCAN FOLDER CENTER")
        report.append("=" * 70)
        
        center_info = result["center_info"]
        report.append(f"Kode Center: {center_info['code']}")
        report.append(f"Path: {center_info['path']}")
        report.append(f"Total Anggota: {center_info['total_anggota']}")
        report.append(f"Anggota Lengkap: {center_info['complete_anggota']}")
        report.append(f"Total File: {center_info['total_files']}")
        report.append(f"Tingkat Kelengkapan: {result['summary']['completion_rate']:.1f}%")
        report.append("")
        
        # List anggota
        report.append("DAFTAR ANGGOTA:")
        for anggota in result["anggota_folders"]:
            info = anggota["anggota_info"]
            completeness = anggota["completeness"]
            status = "✓ LENGKAP" if completeness["complete"] else f"✗ {completeness['percentage']:.1f}%"
            report.append(f"  {info['id']} - {info['nama']}: {status} ({anggota['file_summary']['total_files']} files)")
        
        # Invalid folders
        if result["invalid_folders"]:
            report.append("")
            report.append("FOLDER TIDAK VALID:")
            for invalid in result["invalid_folders"]:
                report.append(f"  ❌ {invalid['name']}: {invalid['error']}")
        
        report.append("")
        report.append("=" * 70)
        
        return "\n".join(report)
    
    def generate_root_report(self, result):
        """Generate laporan untuk scan root"""
        report = []
        report.append("=" * 80)
        report.append("LAPORAN SCAN ROOT DATA_ANGGOTA")
        report.append("=" * 80)
        
        root_info = result["root_info"]
        report.append(f"Path Root: {root_info['path']}")
        report.append(f"Total Center: {root_info['total_centers']}")
        report.append(f"Total Anggota: {root_info['total_anggota']}")
        report.append(f"Total File: {root_info['total_files']}")
        report.append(f"Anggota Lengkap: {root_info['complete_anggota']}")
        report.append(f"Tingkat Kelengkapan Keseluruhan: {result['summary']['overall_completion_rate']:.1f}%")
        report.append("")
        
        # List center
        report.append("DAFTAR CENTER:")
        for center in result["center_folders"]:
            center_info = center["center_info"]
            completion = center["summary"]["completion_rate"]
            report.append(f"  🏢 {center_info['code']}: {center_info['total_anggota']} anggota, {completion:.1f}% lengkap")
        
        # Invalid centers
        if result["invalid_centers"]:
            report.append("")
            report.append("CENTER TIDAK VALID:")
            for invalid in result["invalid_centers"]:
                report.append(f"  ❌ {invalid['name']}: {invalid['error']}")
        
        report.append("")
        report.append("=" * 80)
        
        return "\n".join(report)
    
    def export_scan_to_excel(self, result, scan_type, anggota_reader):
        """Export hasil scan ke Excel"""
        try:
            # Tentukan nama file default
            if scan_type == "anggota":
                anggota_info = result["anggota_info"]
                default_name = f"data_anggota_{anggota_info['id']}_{anggota_info['nama']}"
            elif scan_type == "center":
                center_info = result["center_info"]
                default_name = f"data_center_{center_info['code']}"
            else:  # root
                default_name = "data_root_all_anggota"
            
            file_path = filedialog.asksaveasfilename(
                title="Export ke Excel",
                defaultextension=".xlsx",
                initialfile=f"{default_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                filetypes=[
                    ("Excel Files", "*.xlsx"),
                    ("All Files", "*.*")
                ]
            )
            
            if file_path:
                # Tanya user format export yang diinginkan
                export_choice = messagebox.askyesnocancel(
                    "Pilih Format Export",
                    "Pilih format export Excel:\n\n"
                    "YES = Format Sederhana (kolom: center, anggota_folder, id_anggota, file_01-12)\n"
                    "NO = Format Lengkap (3 sheet dengan detail dokumen)\n"
                    "CANCEL = Batal export"
                )
                
                if export_choice is None:  # Cancel
                    return
                elif export_choice:  # YES - Format sederhana
                    export_result = self.export_simple_excel(result, scan_type, file_path)
                else:  # NO - Format lengkap
                    export_result = anggota_reader.export_to_excel(result, scan_type, file_path)
                
                if export_result.get("success", False):
                    messagebox.showinfo(
                        "Export Berhasil", 
                        f"Data berhasil di-export ke Excel!\n\n"
                        f"File: {export_result['file_path']}\n"
                        f"Rows: {export_result['rows_exported']}\n"
                        f"Format: {'Sederhana' if export_choice else 'Lengkap'}"
                    )
                else:
                    messagebox.showerror("Export Gagal", f"Gagal export ke Excel:\n{export_result.get('error', 'Unknown error')}")
                    print(f"Error during export to Excel: {export_result.get('error', 'Unknown error')}")
        except Exception as e:
            print(f"Exception during export to Excel: {str(e)}")
            messagebox.showerror("Export Gagal", f"Gagal export ke Excel:\n{str(e)}")
            
    
    def export_simple_excel(self, scan_result, scan_type, output_file):
        """Export hasil scan ke Excel dengan format sederhana"""
        try:
            import pandas as pd
            
            data_rows = []
            
            if scan_type == "anggota":
                # Single anggota
                anggota_info = scan_result["anggota_info"]
                file_categories = scan_result["file_categories"]
                
                row = {
                    "center": anggota_info.get("center_code", ""),
                    "anggota_folder": anggota_info["folder_name"],
                    "id_anggota": anggota_info["id"],
                    "file_ditemukan": scan_result["file_summary"]["total_files"]
                }
                
                # Add file columns (01-12)
                for i in range(1, 13):
                    code = f"{i:02d}"
                    files = file_categories.get(code, [])
                    row[f"file_{code}"] = len(files)
                
                data_rows.append(row)
                
            elif scan_type == "center":
                # Multiple anggota dalam center
                center_code = scan_result["center_info"]["code"]
                
                for anggota in scan_result["anggota_folders"]:
                    anggota_info = anggota["anggota_info"]
                    file_categories = anggota["file_categories"]
                    
                    row = {
                        "center": center_code,
                        "anggota_folder": anggota_info["folder_name"], 
                        "id_anggota": anggota_info["id"],
                        "file_ditemukan": anggota["file_summary"]["total_files"]
                    }
                    
                    # Add file columns (01-12)
                    for i in range(1, 13):
                        code = f"{i:02d}"
                        files = file_categories.get(code, [])
                        row[f"file_{code}"] = len(files)
                    
                    data_rows.append(row)
                    
            elif scan_type == "root":
                # Multiple center dengan multiple anggota
                for center in scan_result["center_folders"]:
                    center_code = center["center_info"]["code"]
                    
                    for anggota in center["anggota_folders"]:
                        anggota_info = anggota["anggota_info"]
                        file_categories = anggota["file_categories"]
                        
                        row = {
                            "center": center_code,
                            "anggota_folder": anggota_info["folder_name"],
                            "id_anggota": anggota_info["id"], 
                            "file_ditemukan": anggota["file_summary"]["total_files"]
                        }
                        
                        # Add file columns (01-12)
                        for i in range(1, 13):
                            code = f"{i:02d}"
                            files = file_categories.get(code, [])
                            row[f"file_{code}"] = len(files)
                        
                        data_rows.append(row)
            
            # Create DataFrame
            df = pd.DataFrame(data_rows)
            
            # Ensure column order
            column_order = ["center", "anggota_folder", "id_anggota", "file_ditemukan"]
            column_order.extend([f"file_{i:02d}" for i in range(1, 13)])
            
            df = df[column_order]
            
            # Export to Excel
            df.to_excel(output_file, index=False, sheet_name="Data_Scan")
            
            return {
                "success": True,
                "file_path": output_file,
                "rows_exported": len(df),
                "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            }
            
        except Exception as e:
            return {
                "success": False,
                "error": str(e)
            }
    
    def extract_nama_from_folder(self, folder_name):
        """Extract nama dari folder_name dengan format 6digit_nama"""
        try:
            if "_" in folder_name:
                parts = folder_name.split("_", 1)  # Split maksimal jadi 2 bagian
                if len(parts) > 1:
                    nama_part = parts[1]  # Ambil bagian setelah underscore pertama
                    # Ganti underscore dengan spasi dan title case
                    return nama_part.replace("_", " ").title()
            return ""
        except:
            return ""
    
    def export_simple_excel_with_ya_tidak(self, scan_result, scan_type, output_file):
        """Export hasil scan ke Excel dengan format sederhana menggunakan ADA/TIDAK ADA"""
        try:
            import pandas as pd
            
            data_rows = []
            
            if scan_type == "anggota":
                # Single anggota
                anggota_info = scan_result["anggota_info"]
                file_categories = scan_result["file_categories"]
                
                row = {
                    "center": anggota_info.get("center_code", ""),
                    "anggota_folder": anggota_info["folder_name"],
                    "id_anggota": anggota_info["id"],
                    "nama": self.extract_nama_from_folder(anggota_info["folder_name"]),
                    "file_ditemukan": scan_result["file_summary"]["total_files"]
                }
                
                # Add file columns (01-12) dengan ADA/TIDAK ADA
                for i in range(1, 13):
                    code = f"{i:02d}"
                    files = file_categories.get(code, [])
                    row[f"file_{code}"] = "ADA" if len(files) > 0 else "TIDAK ADA"
                
                # Add nama column
                row["nama"] = self.extract_nama_from_folder(row["anggota_folder"])
                
                data_rows.append(row)
                
            elif scan_type == "center":
                # Multiple anggota dalam center
                center_code = scan_result["center_info"]["code"]
                
                for anggota in scan_result["anggota_folders"]:
                    anggota_info = anggota["anggota_info"]
                    file_categories = anggota["file_categories"]
                    
                    row = {
                        "center": center_code,
                        "anggota_folder": anggota_info["folder_name"], 
                        "id_anggota": anggota_info["id"],
                        "file_ditemukan": anggota["file_summary"]["total_files"]
                    }
                    
                    # Add file columns (01-12) dengan ADA/TIDAK ADA
                    for i in range(1, 13):
                        code = f"{i:02d}"
                        files = file_categories.get(code, [])
                        row[f"file_{code}"] = "ADA" if len(files) > 0 else "TIDAK ADA"
                    
                    # Add nama column
                    row["nama"] = self.extract_nama_from_folder(row["anggota_folder"])
                    
                    data_rows.append(row)
                    
            elif scan_type == "root":
                # Multiple center dengan multiple anggota
                for center in scan_result["center_folders"]:
                    center_code = center["center_info"]["code"]
                    
                    for anggota in center["anggota_folders"]:
                        anggota_info = anggota["anggota_info"]
                        file_categories = anggota["file_categories"]
                        
                        row = {
                            "center": center_code,
                            "anggota_folder": anggota_info["folder_name"],
                            "id_anggota": anggota_info["id"], 
                            "file_ditemukan": anggota["file_summary"]["total_files"]
                        }
                        
                        # Add file columns (01-12) dengan ADA/TIDAK ADA
                        for i in range(1, 13):
                            code = f"{i:02d}"
                            files = file_categories.get(code, [])
                            row[f"file_{code}"] = "ADA" if len(files) > 0 else "TIDAK ADA"
                        
                        # Add nama column
                        row["nama"] = self.extract_nama_from_folder(row["anggota_folder"])
                        
                        data_rows.append(row)
            
            # Create DataFrame
            df = pd.DataFrame(data_rows)
            
            # Ensure column order
            column_order = ["center", "anggota_folder", "nama", "id_anggota", "file_ditemukan"]
            column_order.extend([f"file_{i:02d}" for i in range(1, 13)])
            
            df = df[column_order]
            
            # Export to Excel
            df.to_excel(output_file, index=False, sheet_name="Data_Scan")
            
            return {
                "success": True,
                "file_path": output_file,
                "rows_exported": len(df),
                "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            }
            
        except Exception as e:
            return {
                "success": False,
                "error": str(e)
            }
    
    def export_scan_to_text(self, report):
        """Export hasil scan ke file text"""
        try:
            file_path = filedialog.asksaveasfilename(
                title="Export ke Text",
                defaultextension=".txt",
                initialfile=f"hasil_scan_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt",
                filetypes=[
                    ("Text Files", "*.txt"),
                    ("All Files", "*.*")
                ]
            )
            
            if file_path:
                with open(file_path, 'w', encoding='utf-8') as f:
                    f.write(report)
                messagebox.showinfo("Export Berhasil", f"Hasil scan berhasil disimpan ke:\n{file_path}")
        except Exception as e:
            messagebox.showerror("Export Gagal", f"Gagal menyimpan file:\n{str(e)}")
    
    def show_process_result(self, result, summary):
        """Menampilkan hasil proses dalam window terpisah"""
        result_window = tk.Toplevel(self.root)
        result_window.title("Hasil Proses Arsip Digital")
        result_window.geometry("700x500")
        result_window.resizable(True, True)
        
        # Center window
        result_window.update_idletasks()
        width = result_window.winfo_width()
        height = result_window.winfo_height()
        x = (result_window.winfo_screenwidth() // 2) - (width // 2)
        y = (result_window.winfo_screenheight() // 2) - (height // 2)
        result_window.geometry(f'{width}x{height}+{x}+{y}')
        
        # Main frame
        main_frame = ttk.Frame(result_window, padding="15")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        result_window.columnconfigure(0, weight=1)
        result_window.rowconfigure(0, weight=1)
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(1, weight=1)
        
        # Title
        title_label = ttk.Label(main_frame, text="📊 HASIL PROSES ARSIP DIGITAL", font=("Arial", 14, "bold"))
        title_label.grid(row=0, column=0, pady=(0, 15))
        
        # Summary text area dengan scrollbar
        text_frame = ttk.Frame(main_frame)
        text_frame.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        text_frame.columnconfigure(0, weight=1)
        text_frame.rowconfigure(0, weight=1)
        
        text_widget = tk.Text(text_frame, wrap=tk.WORD, font=("Consolas", 10))
        scrollbar = ttk.Scrollbar(text_frame, orient="vertical", command=text_widget.yview)
        text_widget.configure(yscrollcommand=scrollbar.set)
        
        text_widget.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))
        
        # Insert summary
        text_widget.insert(tk.END, summary)
        text_widget.config(state=tk.DISABLED)
        
        # Button frame
        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=2, column=0, pady=(15, 0))
        
        # Close button
        close_btn = ttk.Button(button_frame, text="Tutup", command=result_window.destroy)
        close_btn.grid(row=0, column=0, padx=(0, 10))
        
        # Export button (placeholder)
        export_btn = ttk.Button(button_frame, text="Export ke File", 
                               command=lambda: self.export_to_file(summary))
        export_btn.grid(row=0, column=1)
    
    def export_to_file(self, summary):
        """Export summary ke file text"""
        try:
            file_path = filedialog.asksaveasfilename(
                title="Export Hasil Proses",
                defaultextension=".txt",
                filetypes=[
                    ("Text Files", "*.txt"),
                    ("All Files", "*.*")
                ]
            )
            
            if file_path:
                with open(file_path, 'w', encoding='utf-8') as f:
                    f.write(summary)
                messagebox.showinfo("Export Berhasil", f"Hasil proses berhasil disimpan ke:\n{file_path}")
        except Exception as e:
            messagebox.showerror("Export Gagal", f"Gagal menyimpan file:\n{str(e)}")
    
    def back_to_menu(self):
        """Kembali ke menu utama"""
        if self.parent_window:
            self.root.destroy()
            self.parent_window.deiconify()
    
    def exit_app(self):
        """Keluar dari aplikasi"""
        if messagebox.askokcancel("Keluar", "Apakah Anda yakin ingin keluar dari aplikasi?"):
            if self.parent_window:
                self.parent_window.destroy()
            else:
                self.root.destroy()


class ScanFolderApp:
    """Form untuk Scan Folder Arsip Digital (Owncloud)"""
    
    def __init__(self, root, parent_window=None):
        self.root = root
        self.parent_window = parent_window
        
        # Initialize business logic
        self.anggota_reader = AnggotaFolderReader()
        
        self.setup_window()
        self.create_widgets()
        
        # Variable untuk menyimpan path yang dipilih
        self.selected_folder = ""
        self.current_scan_result = None
        self.scan_type = None
    
    def setup_window(self):
        """Setup window utama"""
        self.root.title("Scan Folder Arsip Digital - Owncloud")
        
        # Get screen dimensions
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        
        # Get responsive dimensions
        width, height, self.padding, self.fonts = get_responsive_dimensions(
            700, 650, screen_width, screen_height
        )
        
        self.root.geometry(f"{width}x{height}")
        self.root.resizable(True, True)
        
        # Set minimum size
        self.root.minsize(550, 500)
        
        # Center window
        self.center_window()
    
    def center_window(self):
        """Center window di layar"""
        self.root.update_idletasks()
        width = self.root.winfo_width()
        height = self.root.winfo_height()
        x = (self.root.winfo_screenwidth() // 2) - (width // 2)
        y = (self.root.winfo_screenheight() // 2) - (height // 2)
        self.root.geometry(f'{width}x{height}+{x}+{y}')
    
    def create_widgets(self):
        """Membuat semua widget GUI"""
        # Main frame dengan padding responsif
        main_frame = ttk.Frame(self.root, padding=str(self.padding))
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Configure grid weights
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(2, weight=1)
        
        # Title dengan font responsif
        title_label = ttk.Label(
            main_frame, 
            text="📂 SCAN FOLDER ARSIP DIGITAL", 
            font=("Arial", self.fonts['title'], "bold")
        )
        title_label.grid(row=0, column=0, pady=(0, 10))
        
        # Subtitle
        subtitle_label = ttk.Label(
            main_frame, 
            text="Scan folder arsip digital dari Owncloud",
            font=("Arial", self.fonts['subtitle']),
            foreground="gray"
        )
        subtitle_label.grid(row=1, column=0, pady=(0, 20))
        
        # Control frame dengan padding responsif
        frame_padding = max(10, self.padding - 5)
        control_frame = ttk.LabelFrame(main_frame, text="Pilih Folder Arsip", padding=str(frame_padding))
        control_frame.grid(row=2, column=0, sticky=(tk.W, tk.E), pady=(0, 15))
        control_frame.columnconfigure(0, weight=1)
        
        # Path display dengan wraplength responsif
        wrap_length = max(400, min(800, int(self.root.winfo_screenwidth() * 0.6)))
        self.path_var = tk.StringVar(value="Belum ada folder yang dipilih...")
        path_label = ttk.Label(
            control_frame, 
            textvariable=self.path_var,
            foreground="gray",
            wraplength=wrap_length
        )
        path_label.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 15))
        
        # Browse button
        browse_btn = ttk.Button(
            control_frame, 
            text="📁 Browse Folder Arsip Digital", 
            command=self.browse_folder
        )
        browse_btn.grid(row=1, column=0, pady=(0, 10))
        
        # Scan button
        self.scan_btn = ttk.Button(
            control_frame, 
            text="🔍 Mulai Scan", 
            command=self.scan_folder,
            state="disabled"
        )
        self.scan_btn.grid(row=2, column=0)
        
        # Results frame
        result_frame = ttk.LabelFrame(main_frame, text="Hasil Scan", padding="15")
        result_frame.grid(row=3, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(0, 15))
        result_frame.columnconfigure(0, weight=1)
        result_frame.rowconfigure(0, weight=1)
        
        # Results text area dengan scrollbar
        text_frame = ttk.Frame(result_frame)
        text_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        text_frame.columnconfigure(0, weight=1)
        text_frame.rowconfigure(0, weight=1)
        
        self.result_text = tk.Text(text_frame, wrap=tk.WORD, font=("Consolas", 9), height=15)
        scrollbar = ttk.Scrollbar(text_frame, orient="vertical", command=self.result_text.yview)
        self.result_text.configure(yscrollcommand=scrollbar.set)
        
        self.result_text.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))
        
        self.result_text.config(state=tk.DISABLED)
        
        # Action buttons frame
        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=4, column=0, pady=(10, 0))
        
        # Export Excel button
        self.export_btn = ttk.Button(
            button_frame, 
            text="📊 Export ke Excel", 
            command=self.export_to_excel,
            state="disabled"
        )
        self.export_btn.grid(row=0, column=0, padx=(0, 10))
        
        # Count Files button
        self.count_files_btn = ttk.Button(
            button_frame, 
            text="📈 Hitung File & Export CSV", 
            command=self.count_files_and_export,
            state="disabled"
        )
        self.count_files_btn.grid(row=0, column=1, padx=(10, 10))
        
        # Export Struktur Lengkap button
        self.export_struktur_btn = ttk.Button(
            button_frame, 
            text="📋 Simpan Dan Singkron", 
            command=self.export_struktur_lengkap,
            state="disabled"
        )
        self.export_struktur_btn.grid(row=0, column=2, padx=(10, 10))
        
        # Back button
        if self.parent_window:
            back_btn = ttk.Button(
                button_frame, 
                text="🔙 Kembali ke Menu", 
                command=self.back_to_menu
            )
            back_btn.grid(row=0, column=3, padx=(10, 0))
    
    def browse_folder(self):
        """Fungsi untuk memilih folder arsip digital"""
        # Gunakan default folder jika ada
        default_folder = config_manager.get_default_folder()
        initial_dir = default_folder if default_folder and os.path.exists(default_folder) else os.getcwd()
        
        folder_path = filedialog.askdirectory(
            title="Pilih Folder Arsip Digital (Owncloud)",
            initialdir=initial_dir
        )
        
        if folder_path:
            self.selected_folder = folder_path
            self.path_var.set(folder_path)
            self.scan_btn.config(state="normal")
            
            # Reset hasil sebelumnya
            self.current_scan_result = None
            self.scan_type = None
            self.result_text.config(state=tk.NORMAL)
            self.result_text.delete(1.0, tk.END)
            self.result_text.config(state=tk.DISABLED)
            self.export_btn.config(state="disabled")
    
    def scan_folder(self):
        """Fungsi untuk melakukan scan folder Owncloud"""
        if not self.selected_folder:
            messagebox.showwarning("Peringatan", "Silakan pilih folder terlebih dahulu!")
            return
        
        self.result_text.config(state=tk.NORMAL)
        self.result_text.delete(1.0, tk.END)
        self.result_text.insert(tk.END, f"🔍 Memulai scan folder: {self.selected_folder}\n\n")
        self.result_text.update()
        
        try:
            # Daftar folder standar yang harus dicek
            standard_folders = [
                "01.SURAT_MENYURAT",
                "02.DATA_ANGGOTA",
                "03.DATA_ANGGOTA_KELUAR",
                "04.DATA_DANA_RESIKO",
                "05.DATA_HARI_RAYA_ANGGOTA",
                "06.LAPORAN_BULANAN",
                "07.BUKU_BANK",
                "08.DATA_LWK"
            ]
            
            # Scan folder dan cek keberadaan
            result = self.scan_owncloud_folder(self.selected_folder, standard_folders)
            
            # Generate report
            report = self.generate_owncloud_report(result)
            self.result_text.insert(tk.END, report)
            
            # Simpan hasil untuk export
            self.current_scan_result = result
            self.scan_type = "owncloud"
            self.export_btn.config(state="normal")
            self.count_files_btn.config(state="normal")
            self.export_struktur_btn.config(state="normal")
            
            self.result_text.config(state=tk.DISABLED)
            
        except Exception as e:
            self.result_text.insert(tk.END, f"\n❌ Terjadi kesalahan:\n{str(e)}")
            self.result_text.config(state=tk.DISABLED)
            messagebox.showerror("Error", f"Terjadi kesalahan saat scan:\n{str(e)}")
    
    def scan_owncloud_folder(self, folder_path, standard_folders):
        """Scan folder Owncloud dan cek keberadaan folder standar"""
        result = {
            "success": True,
            "folder_path": folder_path,
            "folder_name": os.path.basename(folder_path),
            "scan_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "standard_folders": [],
            "other_folders": [],
            "summary": {
                "total_standard": len(standard_folders),
                "found_standard": 0,
                "missing_standard": 0,
                "other_count": 0
            }
        }
        
        try:
            # Cek apakah folder path valid
            if not os.path.exists(folder_path):
                result["success"] = False
                result["error"] = "Folder tidak ditemukan"
                return result
            
            if not os.path.isdir(folder_path):
                result["success"] = False
                result["error"] = "Path bukan folder"
                return result
            
            # Scan semua folder di dalam path
            existing_folders = []
            try:
                existing_folders = [f for f in os.listdir(folder_path) 
                                  if os.path.isdir(os.path.join(folder_path, f))]
            except Exception as e:
                result["success"] = False
                result["error"] = f"Gagal membaca isi folder: {str(e)}"
                return result
            
            # Cek setiap folder standar
            for std_folder in standard_folders:
                folder_info = {
                    "name": std_folder,
                    "status": "TIDAK ADA",
                    "exists": False,
                    "path": ""
                }
                
                if std_folder in existing_folders:
                    folder_info["status"] = "ADA"
                    folder_info["exists"] = True
                    folder_info["path"] = os.path.join(folder_path, std_folder)
                    result["summary"]["found_standard"] += 1
                else:
                    result["summary"]["missing_standard"] += 1
                
                result["standard_folders"].append(folder_info)
            
            # Cek folder lain yang tidak ada di daftar standar
            for folder in existing_folders:
                if folder not in standard_folders:
                    result["other_folders"].append({
                        "name": folder,
                        "path": os.path.join(folder_path, folder)
                    })
                    result["summary"]["other_count"] += 1
            
            return result
            
        except Exception as e:
            result["success"] = False
            result["error"] = str(e)
            return result
    
    def generate_owncloud_report(self, result):
        """Generate laporan untuk scan folder Owncloud"""
        report = []
        report.append("=" * 80)
        report.append("LAPORAN SCAN FOLDER OWNCLOUD")
        report.append("=" * 80)
        
        report.append(f"Folder: {result['folder_name']}")
        report.append(f"Path: {result['folder_path']}")
        report.append(f"Waktu Scan: {result['scan_time']}")
        report.append("")
        
        # Summary
        summary = result["summary"]
        report.append("RINGKASAN:")
        report.append(f"  Total Folder Standar: {summary['total_standard']}")
        report.append(f"  ✅ Folder yang Ada: {summary['found_standard']}")
        report.append(f"  ❌ Folder yang Tidak Ada: {summary['missing_standard']}")
        report.append(f"  📁 Folder Lain: {summary['other_count']}")
        report.append("")
        
        # Daftar folder standar
        report.append("=" * 80)
        report.append("DAFTAR FOLDER STANDAR:")
        report.append("=" * 80)
        
        for folder_info in result["standard_folders"]:
            if folder_info["exists"]:
                icon = "✅"
                status = "ADA"
            else:
                icon = "❌"
                status = "TIDAK ADA"
            
            report.append(f"{icon} {folder_info['name']:<40} [{status}]")
        
        # Folder lain (jika ada)
        if result["other_folders"]:
            report.append("")
            report.append("=" * 80)
            report.append("FOLDER LAIN (TIDAK STANDAR):")
            report.append("=" * 80)
            
            for folder in result["other_folders"]:
                report.append(f"📁 {folder['name']}")
        
        report.append("")
        report.append("=" * 80)
        
        return "\n".join(report)
    
    def export_to_excel(self):
        """Export hasil scan ke Excel"""
        if not self.current_scan_result or not self.scan_type:
            messagebox.showwarning("Peringatan", "Belum ada hasil scan untuk di-export!")
            return
        
        try:
            # Tentukan nama file default
            folder_name = self.current_scan_result.get("folder_name", "folder")
            default_name = f"scan_owncloud_{folder_name}"
            
            file_path = filedialog.asksaveasfilename(
                title="Export ke Excel",
                defaultextension=".xlsx",
                initialfile=f"{default_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                filetypes=[
                    ("Excel Files", "*.xlsx"),
                    ("All Files", "*.*")
                ]
            )
            
            if file_path:
                # Export hasil scan Owncloud
                export_result = self.export_owncloud_to_excel(self.current_scan_result, file_path)
                
                if export_result.get("success", False):
                    messagebox.showinfo(
                        "Export Berhasil", 
                        f"Data berhasil di-export ke Excel!\n\n"
                        f"File: {export_result['file_path']}\n"
                        f"Rows: {export_result['rows_exported']}"
                    )
                else:
                    messagebox.showerror(
                        "Export Gagal", 
                        f"Gagal export ke Excel:\n{export_result.get('error', 'Unknown error')}"
                    )
                    
        except Exception as e:
            messagebox.showerror("Export Gagal", f"Gagal export ke Excel:\n{str(e)}")
    
    def export_owncloud_to_excel(self, result, output_file):
        """Export hasil scan Owncloud ke Excel"""
        try:
            import pandas as pd
            
            # Buat data untuk DataFrame
            data_rows = []
            
            # Tambahkan folder standar
            for folder_info in result["standard_folders"]:
                data_rows.append({
                    "No": len(data_rows) + 1,
                    "Nama Folder": folder_info["name"],
                    "Status": folder_info["status"],
                    "Tipe": "Standar",
                    "Path": folder_info.get("path", "-")
                })
            
            # Tambahkan folder lain
            for folder in result["other_folders"]:
                data_rows.append({
                    "No": len(data_rows) + 1,
                    "Nama Folder": folder["name"],
                    "Status": "ADA",
                    "Tipe": "Lain-lain",
                    "Path": folder["path"]
                })
            
            # Create DataFrame
            df = pd.DataFrame(data_rows)
            
            # Export to Excel
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Scan_Folder', index=False)
                
                # Tambahkan sheet summary
                summary_data = [{
                    "Informasi": "Folder yang di-scan",
                    "Value": result["folder_name"]
                }, {
                    "Informasi": "Path",
                    "Value": result["folder_path"]
                }, {
                    "Informasi": "Waktu Scan",
                    "Value": result["scan_time"]
                }, {
                    "Informasi": "Total Folder Standar",
                    "Value": result["summary"]["total_standard"]
                }, {
                    "Informasi": "Folder yang Ada",
                    "Value": result["summary"]["found_standard"]
                }, {
                    "Informasi": "Folder yang Tidak Ada",
                    "Value": result["summary"]["missing_standard"]
                }, {
                    "Informasi": "Folder Lain",
                    "Value": result["summary"]["other_count"]
                }]
                
                df_summary = pd.DataFrame(summary_data)
                df_summary.to_excel(writer, sheet_name='Summary', index=False)
            
            return {
                "success": True,
                "file_path": output_file,
                "rows_exported": len(df),
                "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            }
            
        except Exception as e:
            return {
                "success": False,
                "error": str(e)
            }
    
    def back_to_menu(self):
        """Kembali ke menu utama"""
        if self.parent_window:
            self.root.destroy()
            self.parent_window.deiconify()
    
    def count_files_and_export(self):
        """Menghitung jumlah file di setiap folder dan export ke CSV"""
        if not self.selected_folder:
            messagebox.showwarning("Peringatan", "Silakan pilih folder terlebih dahulu!")
            return
        
        try:
            # Progress dialog
            progress_window = tk.Toplevel(self.root)
            progress_window.title("Menghitung File...")
            progress_window.geometry("400x100")
            progress_window.resizable(False, False)
            
            # Center window
            progress_window.update_idletasks()
            x = (progress_window.winfo_screenwidth() // 2) - 200
            y = (progress_window.winfo_screenheight() // 2) - 50
            progress_window.geometry(f'400x100+{x}+{y}')
            
            progress_window.transient(self.root)
            progress_window.grab_set()
            
            frame = ttk.Frame(progress_window, padding="20")
            frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
            
            ttk.Label(frame, text="Sedang menghitung file...", font=("Arial", 10)).grid(row=0, column=0, pady=(0, 10))
            
            progress_bar = ttk.Progressbar(frame, mode='indeterminate')
            progress_bar.grid(row=1, column=0, sticky=(tk.W, tk.E))
            progress_bar.start()
            
            progress_window.update()
            
            # Hitung file di semua folder
            result = self.count_files_in_directory(self.selected_folder)
            
            progress_window.destroy()
            
            if result["success"]:
                # Tampilkan dialog save file
                default_filename = f"rekap_jumlah_file_{os.path.basename(self.selected_folder)}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
                
                file_path = filedialog.asksaveasfilename(
                    title="Simpan Rekap Jumlah File",
                    defaultextension=".csv",
                    initialfile=default_filename,
                    initialdir=self.selected_folder,
                    filetypes=[
                        ("CSV Files", "*.csv"),
                        ("Excel Files", "*.xlsx"),
                        ("All Files", "*.*")
                    ]
                )
                
                if file_path:
                    # Export ke CSV atau Excel
                    if file_path.endswith('.xlsx'):
                        export_result = self.export_file_count_to_excel(result, file_path)
                    else:
                        export_result = self.export_file_count_to_csv(result, file_path)
                    
                    if export_result["success"]:
                        messagebox.showinfo(
                            "Export Berhasil",
                            f"Rekap jumlah file berhasil disimpan!\n\n"
                            f"File: {export_result['file_path']}\n"
                            f"Total Folder: {result['summary']['total_folders']}\n"
                            f"Total File: {result['summary']['total_files']}"
                        )
                    else:
                        messagebox.showerror("Export Gagal", f"Gagal menyimpan file:\n{export_result.get('error', 'Unknown error')}")
            else:
                messagebox.showerror("Error", f"Gagal menghitung file:\n{result.get('error', 'Unknown error')}")
                
        except Exception as e:
            if 'progress_window' in locals():
                progress_window.destroy()
            messagebox.showerror("Error", f"Terjadi kesalahan:\n{str(e)}")
    
    def count_files_in_directory(self, root_path):
        """Menghitung jumlah file di setiap folder dan subfolder"""
        result = {
            "success": True,
            "root_path": root_path,
            "root_name": os.path.basename(root_path),
            "scan_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "folders": [],
            "summary": {
                "total_folders": 0,
                "total_files": 0,
                "total_size_bytes": 0,
                "max_depth": 0
            }
        }
        
        try:
            # Walk through all directories
            for dirpath, dirnames, filenames in os.walk(root_path):
                # Hitung file di folder ini
                file_count = len(filenames)
                
                # Hitung total size
                folder_size = 0
                for filename in filenames:
                    try:
                        file_path = os.path.join(dirpath, filename)
                        folder_size += os.path.getsize(file_path)
                    except:
                        pass
                
                # Dapatkan relative path dari root
                rel_path = os.path.relpath(dirpath, root_path)
                if rel_path == ".":
                    rel_path = os.path.basename(root_path)
                    path_parts = [os.path.basename(root_path)]
                else:
                    # Split path menjadi parts
                    path_parts = rel_path.split(os.sep)
                
                # Tentukan level/depth
                level = len(path_parts) - 1
                folder_name = os.path.basename(dirpath)
                
                # Update max depth
                if level > result["summary"]["max_depth"]:
                    result["summary"]["max_depth"] = level
                
                # Validasi kelengkapan berdasarkan struktur standar
                status = self.validate_folder_structure(rel_path, file_count)
                
                folder_info = {
                    "path": dirpath,
                    "relative_path": rel_path,
                    "path_parts": path_parts,
                    "folder_name": folder_name,
                    "level": level,
                    "file_count": file_count,
                    "folder_size": folder_size,
                    "status": status,
                    "subfolder_count": len(dirnames)
                }
                
                result["folders"].append(folder_info)
                result["summary"]["total_folders"] += 1
                result["summary"]["total_files"] += file_count
                result["summary"]["total_size_bytes"] += folder_size
            
            return result
            
        except Exception as e:
            result["success"] = False
            result["error"] = str(e)
            return result
    
    def validate_folder_structure(self, rel_path, file_count):
        """Validasi kelengkapan folder berdasarkan struktur standar"""
        # Daftar folder standar yang harus ada file
        standard_folders_with_files = [
            "01.SURAT_MENYURAT", "02.DATA_ANGGOTA", "03.DATA_ANGGOTA_KELUAR",
            "04.DATA_DANA_RESIKO", "05.DATA_HARI_RAYA_ANGGOTA", "06.LAPORAN_BULANAN",
            "07.BUKU_BANK", "08.DATA_LWK"
        ]
        
        # Cek apakah ini folder standar atau subfolder standar
        for std_folder in standard_folders_with_files:
            if std_folder in rel_path:
                if file_count > 0:
                    return "TERISI"
                else:
                    return "KOSONG"
        
        # Untuk folder lain
        if file_count > 0:
            return "TERISI"
        elif file_count == 0:
            return "KOSONG"
        
        return "OK"
    
    def validate_data_anggota_structure(self, path_parts):
        """Validasi khusus untuk struktur folder 02.DATA_ANGGOTA"""
        keterangan = ""
        
        # Cek apakah ini path untuk 02.DATA_ANGGOTA
        if len(path_parts) > 0 and "02.DATA_ANGGOTA" in path_parts[0]:
            # Cek SUB FOLDER 1 (seharusnya 4 digit nomor center)
            if len(path_parts) >= 2:
                sub_folder_1 = path_parts[1]
                
                # Cek apakah 4 digit angka
                if sub_folder_1.isdigit() and len(sub_folder_1) == 4:
                    keterangan = f"Center: {sub_folder_1}"
                else:
                    # Bukan 4 digit angka
                    keterangan = f"⚠️ Bukan format center (bukan 4 digit): {sub_folder_1}"
            
            # Cek SUB FOLDER 2 (seharusnya format IDANGGOTA_NAMAANGGOTA)
            if len(path_parts) >= 3:
                sub_folder_2 = path_parts[2]
                
                # Cek apakah mengandung underscore dan angka di awal
                if "_" in sub_folder_2:
                    parts = sub_folder_2.split("_", 1)
                    if len(parts) == 2 and parts[0].isdigit():
                        if keterangan:
                            keterangan += f" | Anggota: {parts[0]}"
                        else:
                            keterangan = f"Anggota: {parts[0]}"
                    else:
                        if keterangan:
                            keterangan += f" | ⚠️ Format anggota tidak sesuai"
                        else:
                            keterangan = "⚠️ Format anggota tidak sesuai"
                else:
                    if keterangan:
                        keterangan += f" | ⚠️ Folder anggota tanpa underscore"
                    else:
                        keterangan = "⚠️ Folder anggota tanpa underscore"
        
        return keterangan
    
    def format_path_parts(self, path_parts):
        """Format path parts dengan aturan khusus untuk 02.DATA_ANGGOTA"""
        formatted_parts = []
        
        for i, part in enumerate(path_parts):
            # Cek apakah ini SUB FOLDER 1 dari 02.DATA_ANGGOTA
            if i == 1 and len(path_parts) > 0 and "02.DATA_ANGGOTA" in path_parts[0]:
                # Cek apakah ini angka
                if part.isdigit():
                    # Format menjadi 4 digit dengan leading zero
                    formatted_parts.append(part.zfill(4))
                else:
                    # Bukan angka, biarkan seperti aslinya
                    formatted_parts.append(part)
            else:
                # Folder lain, biarkan seperti aslinya
                formatted_parts.append(part)
        
        return formatted_parts
    
    def export_struktur_lengkap(self):
        if not self.selected_folder:
            messagebox.showwarning("Peringatan", "Silakan pilih folder terlebih dahulu!")
            return

        try:
            # === Progress dialog ===
            progress_window = tk.Toplevel(self.root)
            progress_window.title("Membuat Struktur Lengkap...")
            progress_window.geometry("400x100")
            progress_window.resizable(False, False)
            progress_window.update_idletasks()
            x = (progress_window.winfo_screenwidth() // 2) - 200
            y = (progress_window.winfo_screenheight() // 2) - 50
            progress_window.geometry(f'400x100+{x}+{y}')
            progress_window.transient(self.root)
            progress_window.grab_set()

            frame = ttk.Frame(progress_window, padding="20")
            frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
            ttk.Label(frame, text="Sedang membuat struktur lengkap...", font=("Arial", 10)).grid(row=0, column=0, pady=(0, 10))
            progress_bar = ttk.Progressbar(frame, mode='indeterminate')
            progress_bar.grid(row=1, column=0, sticky=(tk.W, tk.E))
            progress_bar.start()
            progress_window.update()

            # === Scan struktur lengkap dengan file ===
            result = self.scan_struktur_lengkap(self.selected_folder)

            progress_window.destroy()

            if result["success"]:
                # === Tahap 1: Buat database.xlsx terlebih dahulu di AppData ===
                temp_path = get_database_path()
                export_result = self.export_struktur_to_excel(result, temp_path)

                if export_result["success"]:
                    # === Tahap 2: Setelah selesai, baru dialog Save As ===
                    default_filename = f"struktur_lengkap_{os.path.basename(self.selected_folder)}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                    file_path = filedialog.asksaveasfilename(
                        title="Simpan Struktur Lengkap",
                        defaultextension=".xlsx",
                        initialfile=default_filename,
                        initialdir=self.selected_folder,
                        filetypes=[("Excel Files", "*.xlsx")]
                    )

                    if file_path:
                        # Salin hasil ke lokasi pilihan user
                        import shutil
                        shutil.copy2(temp_path, file_path)

                        messagebox.showinfo(
                            "Export Berhasil",
                            f"Struktur lengkap berhasil disimpan!\n\n"
                            f"File: {file_path}\n"
                            f"Total Sheet: {export_result['total_sheets']}"
                        )


                else:
                    messagebox.showerror("Export Gagal", f"Gagal menyimpan file sementara:\n{export_result.get('error', 'Unknown error')}")

            else:
                messagebox.showerror("Error", f"Gagal scan struktur:\n{result.get('error', 'Unknown error')}")

        except Exception as e:
            if 'progress_window' in locals():
                progress_window.destroy()
            messagebox.showerror("Error", f"Terjadi kesalahan:\n{str(e)}")

    def scan_struktur_lengkap(self, root_path):
        """Scan struktur lengkap termasuk file"""
        result = {
            "success": True,
            "root_path": root_path,
            "root_name": os.path.basename(root_path),
            "scan_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "folders": {}
        }
        
        try:
            # Daftar folder standar yang akan dijadikan sheet
            standard_folders = [
                "01.SURAT_MENYURAT",
                "02.DATA_ANGGOTA",
                "03.DATA_ANGGOTA_KELUAR",
                "04.DATA_DANA_RESIKO",
                "05.DATA_HARI_RAYA_ANGGOTA",
                "06.LAPORAN_BULANAN",
                "07.BUKU_BANK",
                "08.DATA_LWK"
            ]
            
            # Scan setiap folder standar
            for std_folder in standard_folders:
                folder_path = os.path.join(root_path, std_folder)
                
                if os.path.exists(folder_path) and os.path.isdir(folder_path):
                    result["folders"][std_folder] = self.scan_folder_recursive(folder_path, std_folder)
                else:
                    result["folders"][std_folder] = {
                        "exists": False,
                        "items": []
                    }
            
            return result
            
        except Exception as e:
            result["success"] = False
            result["error"] = str(e)
            return result
    
    def scan_folder_recursive(self, folder_path, folder_name):
        """Scan folder secara rekursif hingga file"""
        folder_data = {
            "exists": True,
            "path": folder_path,
            "items": []
        }
        
        try:
            for item in os.listdir(folder_path):
                item_path = os.path.join(folder_path, item)
                rel_path = os.path.relpath(item_path, folder_path)
                
                if os.path.isdir(item_path):
                    # Ini folder - scan rekursif
                    folder_info = {
                        "type": "folder",
                        "name": item,
                        "path": item_path,
                        "relative_path": rel_path,
                        "status": "FOLDER",
                        "size": 0,
                        "children": []
                    }
                    
                    # Scan semua children secara rekursif
                    folder_info["children"] = self.get_folder_children(item_path, folder_path)
                    
                    folder_data["items"].append(folder_info)
                else:
                    # Ini file di root folder standar
                    try:
                        file_size = os.path.getsize(item_path)
                        folder_data["items"].append({
                            "type": "file",
                            "name": item,
                            "path": item_path,
                            "relative_path": rel_path,
                            "status": "FILE",
                            "size": file_size
                        })
                    except:
                        pass
            
            return folder_data
            
        except Exception as e:
            folder_data["error"] = str(e)
            return folder_data
    
    def get_folder_children(self, folder_path, root_path):
        """Helper untuk mendapatkan children folder secara rekursif"""
        children = []
        
        try:
            for item in os.listdir(folder_path):
                item_path = os.path.join(folder_path, item)
                rel_path = os.path.relpath(item_path, root_path)
                
                if os.path.isdir(item_path):
                    # Subfolder - scan rekursif lagi
                    folder_info = {
                        "type": "folder",
                        "name": item,
                        "path": item_path,
                        "relative_path": rel_path,
                        "status": "FOLDER",
                        "size": 0,
                        "children": []
                    }
                    
                    # Rekursif untuk subfolder ini
                    folder_info["children"] = self.get_folder_children(item_path, root_path)
                    children.append(folder_info)
                else:
                    # File
                    try:
                        file_size = os.path.getsize(item_path)
                        children.append({
                            "type": "file",
                            "name": item,
                            "path": item_path,
                            "relative_path": rel_path,
                            "status": "FILE",
                            "size": file_size
                        })
                    except:
                        pass
        except:
            pass
        
        return children
    
    def export_struktur_to_excel(self, result, output_path):
        """Export struktur ke Excel dengan multiple sheets"""
        try:
            import pandas as pd
            from openpyxl import load_workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            # Create Excel writer
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                sheet_count = 0
                
                # Buat sheet untuk setiap folder standar
                for folder_name, folder_data in result["folders"].items():
                    if not folder_data["exists"]:
                        # Folder tidak ada
                        df = pd.DataFrame([{
                            "Status": "FOLDER TIDAK ADA",
                            "Nama": folder_name,
                            "Keterangan": "Folder ini tidak ditemukan"
                        }])
                        sheet_name = folder_name[:31]  # Excel sheet name max 31 chars
                        df.to_excel(writer, sheet_name=sheet_name, index=False)
                        sheet_count += 1
                        continue
                    
                    # Panggil handler spesifik per folder
                    if folder_name == "01.SURAT_MENYURAT":
                        rows = self.handle_surat_menyurat(folder_data)
                    elif folder_name == "02.DATA_ANGGOTA":
                        rows = self.handle_data_anggota(folder_data)
                    elif folder_name == "03.DATA_ANGGOTA_KELUAR":
                        rows = self.handle_data_anggota_keluar(folder_data)
                    elif folder_name == "04.DATA_DANA_RESIKO":
                        rows = self.handle_data_dana_resiko(folder_data)
                    elif folder_name == "05.DATA_HARI_RAYA_ANGGOTA":
                        rows = self.handle_hari_raya_anggota(folder_data)
                    elif folder_name == "06.LAPORAN_BULANAN":
                        rows = self.handle_laporan_bulanan(folder_data)
                    elif folder_name == "07.BUKU_BANK":
                        rows = self.handle_buku_bank(folder_data)
                    elif folder_name == "08.DATA_LWK":
                        rows = self.handle_data_lwk(folder_data)
                    else:
                        rows = []
                    
                    if rows:
                        df = pd.DataFrame(rows)
                        sheet_name = folder_name[:31]
                        df.to_excel(writer, sheet_name=sheet_name, index=False)
                        sheet_count += 1
                    else:
                        # Folder kosong
                        df = pd.DataFrame([{
                            "Status": "FOLDER KOSONG",
                            "Nama": folder_name,
                            "Keterangan": "Tidak ada file atau subfolder"
                        }])
                        sheet_name = folder_name[:31]
                        df.to_excel(writer, sheet_name=sheet_name, index=False)
                        sheet_count += 1
            
            # Buat hyperlink untuk kolom PATH
            self.add_hyperlinks_to_excel(output_path)
            
            return {
                "success": True,
                "file_path": output_path,
                "total_sheets": sheet_count
            }
            
        except Exception as e:
            return {
                "success": False,
                "error": str(e)
            }
    
    def add_hyperlinks_to_excel(self, excel_path):
        """Tambahkan hyperlink ke kolom PATH di semua sheet"""
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Font
            
            wb = load_workbook(excel_path)
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Cari kolom PATH
                path_col = None
                for idx, cell in enumerate(ws[1], 1):
                    if cell.value == "PATH":
                        path_col = idx
                        break
                
                if path_col:
                    # Tambahkan hyperlink untuk setiap cell di kolom PATH
                    for row in range(2, ws.max_row + 1):
                        cell = ws.cell(row=row, column=path_col)
                        if cell.value and isinstance(cell.value, str) and len(cell.value) > 0:
                            # Buat hyperlink
                            cell.hyperlink = cell.value
                            cell.font = Font(color="0563C1", underline="single")
            
            wb.save(excel_path)
            
        except Exception as e:
            print(f"Error adding hyperlinks: {str(e)}")
            pass
    
    def handle_surat_menyurat(self, folder_data):
        """
        Format: JENIS_SURAT | TAHUN | BULAN | NAMA_FILE | TYPE | UKURAN | PATH
        Struktur: 01.SURAT_MASUK/02.SURAT_KELUAR -> Tahun -> Bulan
        """
        rows = []
        
        for item in folder_data["items"]:
            jenis_surat = item["name"]  # 01.SURAT_MASUK atau 02.SURAT_KELUAR
            
            if "children" in item:
                for tahun_item in item["children"]:
                    tahun = tahun_item["name"]
                    
                    if tahun_item["type"] == "folder" and "children" in tahun_item:
                        for bulan_item in tahun_item["children"]:
                            bulan = bulan_item["name"]
                            
                            if bulan_item["type"] == "folder":
                                # Tambahkan row untuk folder bulan
                                rows.append({
                                    "JENIS_SURAT": jenis_surat,
                                    "TAHUN": tahun,
                                    "BULAN": bulan,
                                    "NAMA_FILE": "",
                                    "TYPE": "FOLDER",
                                    "UKURAN_KB": "",
                                    "PATH": bulan_item["path"]
                                })
                                
                                # Tambahkan file-file di dalam bulan
                                if "children" in bulan_item:
                                    for file_item in bulan_item["children"]:
                                        rows.append({
                                            "JENIS_SURAT": jenis_surat,
                                            "TAHUN": tahun,
                                            "BULAN": bulan,
                                            "NAMA_FILE": file_item["name"],
                                            "TYPE": file_item["status"],
                                            "UKURAN_KB": round(file_item["size"] / 1024, 2) if file_item["type"] == "file" else "",
                                            "PATH": file_item["path"]
                                        })
                            else:
                                # File langsung di tahun (bukan di bulan)
                                rows.append({
                                    "JENIS_SURAT": jenis_surat,
                                    "TAHUN": tahun,
                                    "BULAN": "",
                                    "NAMA_FILE": bulan_item["name"],
                                    "TYPE": bulan_item["status"],
                                    "UKURAN_KB": round(bulan_item["size"] / 1024, 2) if bulan_item["type"] == "file" else "",
                                    "PATH": bulan_item["path"]
                                })
                    else:
                        # File langsung di jenis surat
                        rows.append({
                            "JENIS_SURAT": jenis_surat,
                            "TAHUN": tahun if tahun_item["type"] == "folder" else "",
                            "BULAN": "",
                            "NAMA_FILE": tahun_item["name"] if tahun_item["type"] == "file" else "",
                            "TYPE": tahun_item["status"],
                            "UKURAN_KB": round(tahun_item["size"] / 1024, 2) if tahun_item["type"] == "file" else "",
                            "PATH": tahun_item["path"]
                        })
        
        return rows
    
    def handle_data_anggota(self, folder_data):
        """
        Format: NOMOR_CENTER | ID_NAMA_ANGGOTA | TYPE | UKURAN | PATH
        Struktur: Nomor Center (4 digit) -> IDANGGOTA_NAMAANGGOTA
        """
        rows = []
        
        for item in folder_data["items"]:
            nomor_center = item["name"]
            
            # Format nomor center ke 4 digit
            if nomor_center.isdigit():
                nomor_center = nomor_center.zfill(4)
            
            if "children" in item:
                for anggota_item in item["children"]:
                    if anggota_item["type"] == "folder":
                        # Tambahkan row untuk folder anggota
                        rows.append({
                            "NOMOR_CENTER": nomor_center,
                            "ID_NAMA_ANGGOTA": anggota_item["name"],
                            "NAMA_FILE": "",
                            "TYPE": "FOLDER",
                            "UKURAN_KB": "",
                            "PATH": anggota_item["path"]
                        })
                        
                        # Tambahkan file-file di dalam folder anggota
                        if "children" in anggota_item:
                            for file_item in anggota_item["children"]:
                                rows.append({
                                    "NOMOR_CENTER": nomor_center,
                                    "ID_NAMA_ANGGOTA": anggota_item["name"],
                                    "NAMA_FILE": file_item["name"],
                                    "TYPE": file_item["status"],
                                    "UKURAN_KB": round(file_item["size"] / 1024, 2) if file_item["type"] == "file" else "",
                                    "PATH": file_item["path"]
                                })
                    else:
                        # File langsung di nomor center
                        rows.append({
                            "NOMOR_CENTER": nomor_center,
                            "ID_NAMA_ANGGOTA": "",
                            "NAMA_FILE": anggota_item["name"],
                            "TYPE": anggota_item["status"],
                            "UKURAN_KB": round(anggota_item["size"] / 1024, 2) if anggota_item["type"] == "file" else "",
                            "PATH": anggota_item["path"]
                        })
            else:
                rows.append({
                    "NOMOR_CENTER": nomor_center,
                    "ID_NAMA_ANGGOTA": "",
                    "NAMA_FILE": "",
                    "TYPE": item["status"],
                    "UKURAN_KB": "",
                    "PATH": item["path"]
                })
        
        return rows
    
    def handle_data_anggota_keluar(self, folder_data):
        """
        Format: TAHUN | BULAN | ID_NAMA_ANGGOTA | TYPE | UKURAN | PATH
        Struktur: Tahun -> Bulan -> IDANGGOTA_NAMAANGGOTA
        """
        rows = []
        
        for item in folder_data["items"]:
            tahun = item["name"]
            
            if "children" in item:
                for bulan_item in item["children"]:
                    bulan = bulan_item["name"]
                    
                    if bulan_item["type"] == "folder":
                        # Tambahkan row untuk folder bulan
                        rows.append({
                            "TAHUN": tahun,
                            "BULAN": bulan,
                            "ID_NAMA_ANGGOTA": "",
                            "NAMA_FILE": "",
                            "TYPE": "FOLDER",
                            "UKURAN_KB": "",
                            "PATH": bulan_item["path"]
                        })
                        
                        # Tambahkan folder anggota dan file di dalamnya
                        if "children" in bulan_item:
                            for anggota_item in bulan_item["children"]:
                                if anggota_item["type"] == "folder":
                                    # Folder anggota
                                    rows.append({
                                        "TAHUN": tahun,
                                        "BULAN": bulan,
                                        "ID_NAMA_ANGGOTA": anggota_item["name"],
                                        "NAMA_FILE": "",
                                        "TYPE": "FOLDER",
                                        "UKURAN_KB": "",
                                        "PATH": anggota_item["path"]
                                    })
                                    
                                    # File di dalam folder anggota
                                    if "children" in anggota_item:
                                        for file_item in anggota_item["children"]:
                                            rows.append({
                                                "TAHUN": tahun,
                                                "BULAN": bulan,
                                                "ID_NAMA_ANGGOTA": anggota_item["name"],
                                                "NAMA_FILE": file_item["name"],
                                                "TYPE": file_item["status"],
                                                "UKURAN_KB": round(file_item["size"] / 1024, 2) if file_item["type"] == "file" else "",
                                                "PATH": file_item["path"]
                                            })
                                else:
                                    # File langsung di bulan
                                    rows.append({
                                        "TAHUN": tahun,
                                        "BULAN": bulan,
                                        "ID_NAMA_ANGGOTA": "",
                                        "NAMA_FILE": anggota_item["name"],
                                        "TYPE": anggota_item["status"],
                                        "UKURAN_KB": round(anggota_item["size"] / 1024, 2) if anggota_item["type"] == "file" else "",
                                        "PATH": anggota_item["path"]
                                    })
                    else:
                        # File langsung di tahun
                        rows.append({
                            "TAHUN": tahun,
                            "BULAN": "",
                            "ID_NAMA_ANGGOTA": "",
                            "NAMA_FILE": bulan_item["name"],
                            "TYPE": bulan_item["status"],
                            "UKURAN_KB": round(bulan_item["size"] / 1024, 2) if bulan_item["type"] == "file" else "",
                            "PATH": bulan_item["path"]
                        })
        
        return rows
    
    def handle_data_dana_resiko(self, folder_data):
        """
        Format: TAHUN | BULAN | ID_NAMA_ANGGOTA | NAMA_FILE | TYPE | UKURAN | PATH
        Struktur: Tahun -> Bulan -> Folder ID_NAMA -> File
        """
        rows = []
        
        for item in folder_data["items"]:
            tahun = item["name"]
            
            if "children" in item:
                for bulan_item in item["children"]:
                    bulan = bulan_item["name"]
                    
                    if bulan_item["type"] == "folder":
                        # Tambahkan row untuk folder bulan
                        rows.append({
                            "TAHUN": tahun,
                            "BULAN": bulan,
                            "ID_NAMA_ANGGOTA": "",
                            "NAMA_FILE": "",
                            "TYPE": "FOLDER",
                            "UKURAN_KB": "",
                            "PATH": bulan_item["path"]
                        })
                        
                        # Tambahkan folder ID anggota dan file di dalamnya
                        if "children" in bulan_item:
                            for anggota_item in bulan_item["children"]:
                                if anggota_item["type"] == "folder":
                                    # Folder ID_NAMA anggota
                                    rows.append({
                                        "TAHUN": tahun,
                                        "BULAN": bulan,
                                        "ID_NAMA_ANGGOTA": anggota_item["name"],
                                        "NAMA_FILE": "",
                                        "TYPE": "FOLDER",
                                        "UKURAN_KB": "",
                                        "PATH": anggota_item["path"]
                                    })
                                    
                                    # File di dalam folder anggota
                                    if "children" in anggota_item:
                                        for file_item in anggota_item["children"]:
                                            rows.append({
                                                "TAHUN": tahun,
                                                "BULAN": bulan,
                                                "ID_NAMA_ANGGOTA": anggota_item["name"],
                                                "NAMA_FILE": file_item["name"],
                                                "TYPE": file_item["status"],
                                                "UKURAN_KB": round(file_item["size"] / 1024, 2) if file_item["type"] == "file" else "",
                                                "PATH": file_item["path"]
                                            })
                                else:
                                    # File langsung di bulan (tanpa folder anggota)
                                    rows.append({
                                        "TAHUN": tahun,
                                        "BULAN": bulan,
                                        "ID_NAMA_ANGGOTA": "",
                                        "NAMA_FILE": anggota_item["name"],
                                        "TYPE": anggota_item["status"],
                                        "UKURAN_KB": round(anggota_item["size"] / 1024, 2) if anggota_item["type"] == "file" else "",
                                        "PATH": anggota_item["path"]
                                    })
                    else:
                        # File langsung di tahun
                        rows.append({
                            "TAHUN": tahun,
                            "BULAN": "",
                            "ID_NAMA_ANGGOTA": "",
                            "NAMA_FILE": bulan_item["name"],
                            "TYPE": bulan_item["status"],
                            "UKURAN_KB": round(bulan_item["size"] / 1024, 2) if bulan_item["type"] == "file" else "",
                            "PATH": bulan_item["path"]
                        })
        
        return rows
    
    def handle_hari_raya_anggota(self, folder_data):
        """
        Format: TAHUN | NAMA_FILE | TYPE | UKURAN | PATH
        Struktur: Tahun -> File bulan (01.JANUARI.xlsx, 02.FEBRUARI.xlsx, ...)
        """
        rows = []
        
        for item in folder_data["items"]:
            tahun = item["name"]
            
            if "children" in item:
                for file_item in item["children"]:
                    rows.append({
                        "TAHUN": tahun,
                        "NAMA_FILE": file_item["name"],
                        "TYPE": file_item["status"],
                        "UKURAN_KB": round(file_item["size"] / 1024, 2) if file_item["type"] == "file" else "",
                        "PATH": file_item["path"]
                    })
            else:
                rows.append({
                    "TAHUN": tahun,
                    "NAMA_FILE": "",
                    "TYPE": item["status"],
                    "UKURAN_KB": "",
                    "PATH": item["path"]
                })
        
        return rows
    
    def handle_laporan_bulanan(self, folder_data):
        """
        Format: TAHUN | BULAN | JENIS_DOKUMEN | NAMA_FILE | TYPE | UKURAN | PATH
        Struktur: Tahun -> Bulan -> File laporan (01.NERACA.pdf, dst)
        """
        rows = []
        
        # Mapping kode dokumen
        doc_types = {
            "01.NERACA": "01.NERACA.pdf",
            "02.PERHITUNGAN_HASIL_USAHA": "02.PERHITUNGAN_HASIL_USAHA.pdf",
            "03.TRIAL_BALANCE": "03.TRIAL_BALANCE.pdf",
            "04.FIXED_ASSET": "04.FIXED_ASSET.pdf",
            "05.JOURNAL_VOUCHER": "05.JOURNAL_VOUCHER.pdf",
            "06.INFORMASI_PORTOFOLIO": "06.INFORMASI_PORTOFOLIO.pdf",
            "07.DELIQUENCY": "07.DELIQUENCY.pdf",
            "08.MONTHLY_PROJECT_STATEMENT": "08.MONTHLY_PROJECT_STATEMENT.pdf",
            "09.STATISTIK": "09.STATISTIK.pdf",
            "10.STATISTIK_PETUGAS_LAPANG": "10.STATISTIK_PETUGAS_LAPANG.pdf",
            "11.STATISTIK_WILAYAH": "11.STATISTIK_WILAYAH.pdf",
            "12.LOAN_PURPOSE": "12.LOAN_PURPOSE.pdf"
        }
        
        for item in folder_data["items"]:
            tahun = item["name"]
            
            if "children" in item:
                for bulan_item in item["children"]:
                    bulan = bulan_item["name"]
                    
                    if bulan_item["type"] == "folder":
                        # Tambahkan row untuk folder bulan
                        rows.append({
                            "TAHUN": tahun,
                            "BULAN": bulan,
                            "JENIS_DOKUMEN": "",
                            "NAMA_FILE": "",
                            "TYPE": "FOLDER",
                            "UKURAN_KB": "",
                            "PATH": bulan_item["path"]
                        })
                        
                        # Tambahkan file di dalam bulan
                        if "children" in bulan_item:
                            for file_item in bulan_item["children"]:
                                # Deteksi jenis dokumen dari nama file
                                jenis_dok = ""
                                filename = file_item["name"]
                                for key in doc_types.keys():
                                    if key in filename.upper():
                                        jenis_dok = key
                                        break
                                
                                rows.append({
                                    "TAHUN": tahun,
                                    "BULAN": bulan,
                                    "JENIS_DOKUMEN": jenis_dok,
                                    "NAMA_FILE": filename,
                                    "TYPE": file_item["status"],
                                    "UKURAN_KB": round(file_item["size"] / 1024, 2) if file_item["type"] == "file" else "",
                                    "PATH": file_item["path"]
                                })
                    else:
                        rows.append({
                            "TAHUN": tahun,
                            "BULAN": bulan if bulan_item["type"] == "folder" else "",
                            "JENIS_DOKUMEN": "",
                            "NAMA_FILE": bulan if bulan_item["type"] == "file" else "",
                            "TYPE": bulan_item["status"],
                            "UKURAN_KB": round(bulan_item["size"] / 1024, 2) if bulan_item["type"] == "file" else "",
                            "PATH": bulan_item["path"]
                        })
        
        return rows
    
    def handle_buku_bank_old(self, folder_data):
        """
        Format: TAHUN | BULAN | TANGGAL | NAMA_FILE | TYPE | UKURAN | PATH
        Struktur: Tahun -> Bulan -> File (XX_BUKUBANK.XLSX, 2 digit tanggal)
        """
        rows = []
        
        for item in folder_data["items"]:
            tahun = item["name"]
            
            if "children" in item:
                for bulan_item in item["children"]:
                    bulan = bulan_item["name"]
                    
                    if bulan_item["type"] == "folder":
                        # Tambahkan row untuk folder bulan
                        rows.append({
                            "TAHUN": tahun,
                            "BULAN": bulan,
                            "TANGGAL": "",
                            "NAMA_FILE": "",
                            "TYPE": "FOLDER",
                            "UKURAN_KB": "",
                            "PATH": bulan_item["path"]
                        })
                        
                        # Tambahkan file di dalam bulan
                        if "children" in bulan_item:
                            for file_item in bulan_item["children"]:
                                # Extract tanggal dari nama file (2 digit pertama)
                                tanggal = ""
                                filename = file_item["name"]
                                if len(filename) >= 2 and filename[:2].isdigit():
                                    tanggal = filename[:2]
                                
                                rows.append({
                                    "TAHUN": tahun,
                                    "BULAN": bulan,
                                    "TANGGAL": tanggal,
                                    "NAMA_FILE": filename,
                                    "TYPE": file_item["status"],
                                    "UKURAN_KB": round(file_item["size"] / 1024, 2) if file_item["type"] == "file" else "",
                                    "PATH": file_item["path"]
                                })
                    else:
                        rows.append({
                            "TAHUN": tahun,
                            "BULAN": bulan if bulan_item["type"] == "folder" else "",
                            "TANGGAL": "",
                            "NAMA_FILE": bulan if bulan_item["type"] == "file" else "",
                            "TYPE": bulan_item["status"],
                            "UKURAN_KB": round(bulan_item["size"] / 1024, 2) if bulan_item["type"] == "file" else "",
                            "PATH": bulan_item["path"]
                        })
        
        return rows
    
    def handle_buku_bank(self, folder_data):
        """
        Format Output:
        TAHUN | BULAN | TANGGAL | NAMA_FILE | TYPE | UKURAN_KB | PATH
        Struktur:
        Tahun -> File langsung (contoh: 01.JANUARI_BUKUBANK.xlsx)
        """
        rows = []

        for item in folder_data["items"]:
            tahun = item["name"]

            if "children" in item:
                for file_item in item["children"]:
                    filename = file_item["name"]
                    bulan = ""
                    tanggal = ""

                    # Ekstrak bulan dari nama file (sebelum '_')
                    if "_BUKUBANK" in filename:
                        bulan_part = filename.split("_BUKUBANK")[0]
                        bulan = bulan_part  # contoh: 01.JANUARI

                    # Ekstrak tanggal dari 2 digit pertama jika angka
                    if len(filename) >= 2 and filename[:2].isdigit():
                        tanggal = filename[:2]

                    rows.append({
                        "TAHUN": tahun,
                        "BULAN": bulan,
                        "NAMA_FILE": filename,
                        "TYPE": file_item.get("status", "FILE"),
                        "UKURAN_KB": round(file_item["size"] / 1024, 2) if file_item["type"] == "file" else "",
                        "PATH": file_item["path"]
                    })

            else:
                # Jika langsung file tanpa folder tahun (jarang terjadi, tapi tetap di-handle)
                filename = item["name"]
                rows.append({
                    "TAHUN": "",
                    "BULAN": "",
                    "NAMA_FILE": filename,
                    "TYPE": item.get("status", "FILE"),
                    "UKURAN_KB": round(item["size"] / 1024, 2) if item["type"] == "file" else "",
                    "PATH": item["path"]
                })

        return rows

    def handle_data_lwk(self, folder_data):
        """
        Format: TAHUN | BULAN | TANGGAL | NOMOR_CENTER | NAMA_FILE | TYPE | UKURAN | PATH
        Struktur: Tahun -> Bulan -> File (XX_CCCC.PDF, 2 digit tanggal + 4 digit center)
        """
        rows = []
        
        for item in folder_data["items"]:
            tahun = item["name"]
            
            if "children" in item:
                for bulan_item in item["children"]:
                    bulan = bulan_item["name"]
                    
                    if bulan_item["type"] == "folder":
                        # Tambahkan row untuk folder bulan
                        rows.append({
                            "TAHUN": tahun,
                            "BULAN": bulan,
                            "TANGGAL": "",
                            "NOMOR_CENTER": "",
                            "NAMA_FILE": "",
                            "TYPE": "FOLDER",
                            "UKURAN_KB": "",
                            "PATH": bulan_item["path"]
                        })
                        
                        # Tambahkan file di dalam bulan
                        if "children" in bulan_item:
                            for file_item in bulan_item["children"]:
                                # Extract tanggal dan center dari nama file (XX_CCCC)
                                tanggal = ""
                                nomor_center = ""
                                filename = file_item["name"]
                                
                                # Format: XX_CCCC.PDF
                                parts = filename.split("_")
                                if len(parts) >= 2:
                                    if len(parts[0]) == 2 and parts[0].isdigit():
                                        tanggal = parts[0]
                                    center_part = parts[1].split(".")[0]  # Ambil sebelum ekstensi
                                    if len(center_part) == 4 and center_part.isdigit():
                                        nomor_center = center_part
                                
                                rows.append({
                                    "TAHUN": tahun,
                                    "BULAN": bulan,
                                    "TANGGAL": tanggal,
                                    "NOMOR_CENTER": nomor_center,
                                    "NAMA_FILE": filename,
                                    "TYPE": file_item["status"],
                                    "UKURAN_KB": round(file_item["size"] / 1024, 2) if file_item["type"] == "file" else "",
                                    "PATH": file_item["path"]
                                })
                    else:
                        rows.append({
                            "TAHUN": tahun,
                            "BULAN": bulan if bulan_item["type"] == "folder" else "",
                            "TANGGAL": "",
                            "NOMOR_CENTER": "",
                            "NAMA_FILE": bulan if bulan_item["type"] == "file" else "",
                            "TYPE": bulan_item["status"],
                            "UKURAN_KB": round(bulan_item["size"] / 1024, 2) if bulan_item["type"] == "file" else "",
                            "PATH": bulan_item["path"]
                        })
        
        return rows
    
    def export_file_count_to_csv(self, result, output_path):
        """Export hasil perhitungan file ke CSV dengan kolom hierarki terpisah"""
        try:
            import csv
            
            # Tentukan jumlah kolom level maksimal
            max_depth = result["summary"]["max_depth"]
            
            # Buat header dinamis
            fieldnames = ['No']
            for i in range(max_depth + 1):
                if i == 0:
                    fieldnames.append('FOLDER')
                else:
                    fieldnames.append(f'SUB FOLDER {i}')
            fieldnames.extend(['Level', 'Jumlah_File', 'Jumlah_Subfolder', 'Ukuran_MB', 'Status', 'Keterangan'])
            
            with open(output_path, 'w', newline='', encoding='utf-8-sig') as csvfile:
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                
                writer.writeheader()
                
                for idx, folder in enumerate(result["folders"], start=1):
                    size_mb = folder["folder_size"] / (1024 * 1024)  # Convert to MB
                    
                    # Buat row data
                    row_data = {'No': idx}
                    
                    # Isi kolom hierarki dengan format khusus untuk 02.DATA_ANGGOTA
                    path_parts = folder["path_parts"]
                    formatted_parts = self.format_path_parts(path_parts)
                    
                    for i in range(max_depth + 1):
                        if i == 0:
                            col_name = 'FOLDER'
                        else:
                            col_name = f'SUB FOLDER {i}'
                        
                        if i < len(formatted_parts):
                            row_data[col_name] = formatted_parts[i]
                        else:
                            row_data[col_name] = ''
                    
                    # Validasi khusus untuk 02.DATA_ANGGOTA
                    keterangan = self.validate_data_anggota_structure(path_parts)
                    
                    # Isi kolom lainnya
                    row_data['Level'] = folder["level"]
                    row_data['Jumlah_File'] = folder["file_count"]
                    row_data['Jumlah_Subfolder'] = folder["subfolder_count"]
                    row_data['Ukuran_MB'] = f"{size_mb:.2f}"
                    row_data['Status'] = folder["status"]
                    row_data['Keterangan'] = keterangan
                    
                    writer.writerow(row_data)
                
                # Tambahkan summary di akhir
                writer.writerow({})
                summary_row = {'No': 'RINGKASAN'}
                writer.writerow(summary_row)
                
                summary_row = {'No': 'Total Folder', 'FOLDER': result["summary"]["total_folders"]}
                writer.writerow(summary_row)
                
                summary_row = {'No': 'Total File', 'FOLDER': result["summary"]["total_files"]}
                writer.writerow(summary_row)
                
                summary_row = {'No': 'Total Ukuran', 'FOLDER': f"{result['summary']['total_size_bytes'] / (1024 * 1024):.2f} MB"}
                writer.writerow(summary_row)
            
            return {
                "success": True,
                "file_path": output_path
            }
            
        except Exception as e:
            return {
                "success": False,
                "error": str(e)
            }
    
    def export_file_count_to_excel(self, result, output_path):
        """Export hasil perhitungan file ke Excel dengan kolom hierarki terpisah"""
        try:
            import pandas as pd
            
            # Tentukan jumlah kolom level maksimal
            max_depth = result["summary"]["max_depth"]
            
            # Buat data untuk DataFrame
            data_rows = []
            
            for idx, folder in enumerate(result["folders"], start=1):
                size_mb = folder["folder_size"] / (1024 * 1024)
                
                # Buat row data
                row_data = {'No': idx}
                
                # Isi kolom hierarki dengan format khusus untuk 02.DATA_ANGGOTA
                path_parts = folder["path_parts"]
                formatted_parts = self.format_path_parts(path_parts)
                
                for i in range(max_depth + 1):
                    if i == 0:
                        col_name = 'FOLDER'
                    else:
                        col_name = f'SUB FOLDER {i}'
                    
                    if i < len(formatted_parts):
                        row_data[col_name] = formatted_parts[i]
                    else:
                        row_data[col_name] = ''
                
                # Validasi khusus untuk 02.DATA_ANGGOTA
                keterangan = self.validate_data_anggota_structure(path_parts)
                
                # Isi kolom lainnya
                row_data['Level'] = folder["level"]
                row_data['Jumlah_File'] = folder["file_count"]
                row_data['Jumlah_Subfolder'] = folder["subfolder_count"]
                row_data['Ukuran_MB'] = round(size_mb, 2)
                row_data['Status'] = folder["status"]
                row_data['Keterangan'] = keterangan
                
                data_rows.append(row_data)
            
            df = pd.DataFrame(data_rows)
            
            # Buat summary data
            summary_data = [{
                'Informasi': 'Folder Root',
                'Value': result["root_name"]
            }, {
                'Informasi': 'Path',
                'Value': result["root_path"]
            }, {
                'Informasi': 'Waktu Scan',
                'Value': result["scan_time"]
            }, {
                'Informasi': 'Total Folder',
                'Value': result["summary"]["total_folders"]
            }, {
                'Informasi': 'Total File',
                'Value': result["summary"]["total_files"]
            }, {
                'Informasi': 'Total Ukuran (MB)',
                'Value': round(result["summary"]["total_size_bytes"] / (1024 * 1024), 2)
            }, {
                'Informasi': 'Kedalaman Maksimal',
                'Value': result["summary"]["max_depth"]
            }]
            
            df_summary = pd.DataFrame(summary_data)
            
            # Export to Excel dengan 2 sheet
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Rekap_File', index=False)
                df_summary.to_excel(writer, sheet_name='Summary', index=False)
            
            return {
                "success": True,
                "file_path": output_path
            }
            
        except Exception as e:
            return {
                "success": False,
                "error": str(e)
            }



class UniversalScanApp:
    """Form untuk Universal Scan - Menscan seluruh folder dan file yang dipilih"""
    
    def __init__(self, root, parent_window=None):
        self.root = root
        self.parent_window = parent_window
        
        # Initialize variables
        self.selected_folder = ""
        self.scan_results = []
        self.database_file = "universal_scan_database.xlsx"
        
        self.setup_window()
        self.create_widgets()
        self.load_existing_database()
    
    def setup_window(self):
        """Setup window utama untuk universal scan"""
        self.root.title("Universal Scan - Database Folder & File")
        
        # Get screen dimensions
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        
        # Get responsive dimensions
        width, height, self.padding, self.fonts = get_responsive_dimensions(
            1000, 800, screen_width, screen_height
        )
        
        self.root.geometry(f"{width}x{height}")
        self.root.resizable(True, True)
        
        # Set minimum size to prevent too small windows
        self.root.minsize(700, 600)
        
        # Center window
        self.center_window()
    
    def center_window(self):
        """Center window di layar"""
        self.root.update_idletasks()
        width = self.root.winfo_width()
        height = self.root.winfo_height()
        x = (self.root.winfo_screenwidth() // 2) - (width // 2)
        y = (self.root.winfo_screenheight() // 2) - (height // 2)
        self.root.geometry(f'{width}x{height}+{x}+{y}')
    
    def create_widgets(self):
        """Membuat widget untuk universal scan"""
        # Main frame dengan padding responsif
        main_frame = ttk.Frame(self.root, padding=str(self.padding))
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Configure grid weights
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(5, weight=1)  # Row untuk treeview
        
        # Title dengan ukuran font responsif
        title_label = ttk.Label(
            main_frame, 
            text="🗄️ UNIVERSAL SCAN DATABASE", 
            font=("Arial", self.fonts['title'], "bold")
        )
        title_label.grid(row=0, column=0, pady=(0, 10))
        
        # Subtitle
        subtitle_label = ttk.Label(
            main_frame, 
            text="Scan seluruh folder dan file untuk dijadikan database Excel dengan synchronize",
            font=("Arial", self.fonts['subtitle']),
            foreground="gray"
        )
        subtitle_label.grid(row=1, column=0, pady=(0, 20))
        
        # Database info frame
        db_frame = ttk.LabelFrame(main_frame, text="📊 Database Info", padding=str(self.padding-5))
        db_frame.grid(row=2, column=0, sticky=(tk.W, tk.E), pady=(0, 15))
        db_frame.columnconfigure(1, weight=1)
        
        # Database file path
        ttk.Label(db_frame, text="Database File:", font=("Arial", self.fonts['normal'], "bold")).grid(row=0, column=0, sticky=tk.W, padx=(0, 10))
        self.db_file_var = tk.StringVar(value=self.database_file)
        ttk.Label(db_frame, textvariable=self.db_file_var, foreground="blue", font=("Arial", self.fonts['normal'])).grid(row=0, column=1, sticky=tk.W)
        
        # Record count
        ttk.Label(db_frame, text="Total Records:", font=("Arial", self.fonts['normal'], "bold")).grid(row=1, column=0, sticky=tk.W, padx=(0, 10))
        self.record_count_var = tk.StringVar(value="0")
        ttk.Label(db_frame, textvariable=self.record_count_var, foreground="green", font=("Arial", self.fonts['normal'])).grid(row=1, column=1, sticky=tk.W)
        
        # Folder selection frame dengan padding responsif
        folder_frame = ttk.LabelFrame(main_frame, text="📂 Pilih Folder untuk Discan", padding=str(self.padding-5))
        folder_frame.grid(row=3, column=0, sticky=(tk.W, tk.E), pady=(0, 15))
        folder_frame.columnconfigure(0, weight=1)
        
        # Folder path display dengan wraplength responsif
        wrap_length = max(400, int(self.root.winfo_width() * 0.7)) if hasattr(self, 'root') else 500
        
        self.folder_var = tk.StringVar(value="Belum ada folder yang dipilih...")
        folder_path_label = ttk.Label(
            folder_frame, 
            textvariable=self.folder_var,
            foreground="gray",
            wraplength=wrap_length,
            font=("Arial", self.fonts['normal'])
        )
        folder_path_label.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        
        # Buttons frame
        btn_frame = ttk.Frame(folder_frame)
        btn_frame.grid(row=1, column=0)
        
        # Browse folder button
        browse_btn = ttk.Button(
            btn_frame, 
            text="📂 Pilih Folder", 
            command=self.browse_folder,
            style="Accent.TButton"
        )
        browse_btn.grid(row=0, column=0, padx=(0, 10))
        
        # Scan button
        self.scan_btn = ttk.Button(
            btn_frame, 
            text="🔍 Scan & Update Database", 
            command=self.scan_and_update_database,
            state=tk.DISABLED
        )
        self.scan_btn.grid(row=0, column=1, padx=(10, 10))
        
        # Synchronize button
        self.sync_btn = ttk.Button(
            btn_frame, 
            text="🔄 Synchronize Database", 
            command=self.synchronize_database,
            state=tk.DISABLED
        )
        self.sync_btn.grid(row=0, column=2, padx=(10, 10))
        
        # Export button
        self.export_btn = ttk.Button(
            btn_frame, 
            text="📊 Export Database", 
            command=self.export_database
        )
        self.export_btn.grid(row=0, column=3, padx=(10, 0))
        
        # Status info frame
        status_frame = ttk.Frame(main_frame)
        status_frame.grid(row=4, column=0, sticky=(tk.W, tk.E), pady=(10, 10))
        
        self.status_var = tk.StringVar(value="Database siap digunakan")
        status_label = ttk.Label(
            status_frame,
            textvariable=self.status_var,
            font=("Arial", self.fonts['normal']),
            foreground="blue"
        )
        status_label.grid(row=0, column=0, sticky=tk.W)
        
        # Results frame dengan treeview
        results_frame = ttk.LabelFrame(main_frame, text="📋 Database Records", padding=str(self.padding-5))
        results_frame.grid(row=5, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(0, 15))
        results_frame.columnconfigure(0, weight=1)
        results_frame.rowconfigure(0, weight=1)
        
        # Create scrollbars
        tree_scroll_y = ttk.Scrollbar(results_frame, orient=tk.VERTICAL)
        tree_scroll_y.grid(row=0, column=1, sticky=(tk.N, tk.S))
        
        tree_scroll_x = ttk.Scrollbar(results_frame, orient=tk.HORIZONTAL)
        tree_scroll_x.grid(row=1, column=0, sticky=(tk.W, tk.E))
        
        # Treeview
        self.tree = ttk.Treeview(
            results_frame,
            columns=("ID", "File_Path", "Status", "Ukuran_MB", "Last_Modified", "Scan_Time"),
            show="headings",
            yscrollcommand=tree_scroll_y.set,
            xscrollcommand=tree_scroll_x.set
        )
        tree_scroll_y.config(command=self.tree.yview)
        tree_scroll_x.config(command=self.tree.xview)
        
        # Define columns dengan ukuran responsif
        self.tree.heading("ID", text="ID")
        self.tree.heading("File_Path", text="File Path")
        self.tree.heading("Status", text="Status")
        self.tree.heading("Ukuran_MB", text="Ukuran (MB)")
        self.tree.heading("Last_Modified", text="Last Modified")
        self.tree.heading("Scan_Time", text="Scan Time")
        
        # Set column widths
        self.tree.column("ID", width=60, anchor=tk.CENTER)
        self.tree.column("File_Path", width=400, anchor=tk.W)
        self.tree.column("Status", width=80, anchor=tk.CENTER)
        self.tree.column("Ukuran_MB", width=100, anchor=tk.E)
        self.tree.column("Last_Modified", width=150, anchor=tk.CENTER)
        self.tree.column("Scan_Time", width=150, anchor=tk.CENTER)
        
        self.tree.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Bind double-click to open file/folder
        self.tree.bind("<Double-Button-1>", self.open_selected_item)
        
        # Footer buttons
        footer_frame = ttk.Frame(main_frame)
        footer_frame.grid(row=6, column=0, pady=(10, 0))
        
        # Clear database button
        clear_btn = ttk.Button(
            footer_frame, 
            text="🗑️ Clear Database", 
            command=self.clear_database
        )
        clear_btn.grid(row=0, column=0, padx=(0, 10))
        
        # Refresh view button
        refresh_btn = ttk.Button(
            footer_frame, 
            text="🔄 Refresh View", 
            command=self.refresh_treeview
        )
        refresh_btn.grid(row=0, column=1, padx=(10, 10))
        
        # Back button
        if self.parent_window:
            back_btn = ttk.Button(
                footer_frame, 
                text="⬅️ Kembali ke Menu", 
                command=self.back_to_menu
            )
            back_btn.grid(row=0, column=2, padx=(10, 0))
    
    def load_existing_database(self):
        """Load database yang sudah ada jika tersedia"""
        if os.path.exists(self.database_file):
            try:
                df = pd.read_excel(self.database_file)
                self.scan_results = df.to_dict('records')
                self.record_count_var.set(str(len(self.scan_results)))
                self.status_var.set(f"✅ Database loaded: {len(self.scan_results)} records")
                self.refresh_treeview()
                
                # Enable synchronize button if there are records
                if self.scan_results:
                    self.sync_btn.config(state=tk.NORMAL)
                    
            except Exception as e:
                self.status_var.set(f"⚠️ Error loading database: {str(e)}")
        else:
            self.status_var.set("📋 Database baru akan dibuat saat scan pertama")
    
    def browse_folder(self):
        """Fungsi untuk memilih folder"""
        # Gunakan default folder jika ada
        default_folder = config_manager.get_default_folder()
        initial_dir = default_folder if default_folder and os.path.exists(default_folder) else os.getcwd()
        
        folder_path = filedialog.askdirectory(
            title="Pilih Folder untuk Universal Scan",
            initialdir=initial_dir
        )
        
        if folder_path:
            self.selected_folder = folder_path
            self.folder_var.set(folder_path)
            self.scan_btn.config(state=tk.NORMAL)
            self.status_var.set(f"📂 Folder dipilih: {os.path.basename(folder_path)}")
    
    def scan_and_update_database(self):
        """Scan folder dan update database"""
        if not self.selected_folder:
            messagebox.showwarning("Peringatan", "Silakan pilih folder terlebih dahulu!")
            return
        
        # Konfirmasi
        result = messagebox.askyesno(
            "Konfirmasi Scan",
            f"Akan melakukan scan pada folder:\n{self.selected_folder}\n\n"
            f"Proses ini akan:\n"
            f"• Scan semua folder dan file\n"
            f"• Update database dengan data terbaru\n"
            f"• Menambahkan record baru\n"
            f"• Update status file yang sudah ada\n\n"
            f"Lanjutkan?"
        )
        
        if not result:
            return
        
        # Progress dialog
        progress_window = tk.Toplevel(self.root)
        progress_window.title("Scanning...")
        progress_window.geometry("500x200")
        progress_window.resizable(False, False)
        
        # Center progress window
        progress_window.update_idletasks()
        x = (progress_window.winfo_screenwidth() // 2) - 250
        y = (progress_window.winfo_screenheight() // 2) - 100
        progress_window.geometry(f'500x200+{x}+{y}')
        
        progress_window.transient(self.root)
        progress_window.grab_set()
        
        # Progress content
        progress_frame = ttk.Frame(progress_window, padding="20")
        progress_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        progress_label = ttk.Label(
            progress_frame, 
            text="Scanning folder dan file...",
            font=("Arial", 12, "bold")
        )
        progress_label.grid(row=0, column=0, pady=(0, 10))
        
        self.progress_var = tk.DoubleVar()
        progress_bar = ttk.Progressbar(
            progress_frame, 
            mode='indeterminate'
        )
        progress_bar.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        progress_bar.start(10)
        
        status_label = ttk.Label(
            progress_frame,
            text="Memulai scan...",
            font=("Arial", 9),
            foreground="gray"
        )
        status_label.grid(row=2, column=0, pady=(10, 0))
        
        progress_window.update()
        
        # Lakukan scan
        try:
            new_records = []
            updated_count = 0
            new_count = 0
            scan_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            # Create lookup untuk existing records
            existing_paths = {record['file_path']: record for record in self.scan_results}
            
            # Scan folder secara rekursif
            for root, dirs, files in os.walk(self.selected_folder):
                # Update progress status
                status_label.config(text=f"Scanning: {os.path.basename(root)}...")
                progress_window.update()
                
                # Scan folders
                for dir_name in dirs:
                    dir_path = os.path.join(root, dir_name)
                    relative_path = os.path.relpath(dir_path, self.selected_folder)
                    
                    try:
                        dir_size = self.get_folder_size(dir_path)
                        dir_size_mb = round(dir_size / (1024 * 1024), 2)
                        last_modified = datetime.fromtimestamp(os.path.getmtime(dir_path)).strftime("%Y-%m-%d %H:%M:%S")
                        
                        record = {
                            'file_path': dir_path,
                            'status': 'folder',
                            'ukuran_mb': dir_size_mb,
                            'last_modified': last_modified,
                            'scan_time': scan_time,
                            'relative_path': relative_path
                        }
                        
                        # Check if already exists
                        if dir_path in existing_paths:
                            # Update existing record
                            existing_record = existing_paths[dir_path]
                            if (existing_record['last_modified'] != last_modified or 
                                existing_record['ukuran_mb'] != dir_size_mb):
                                existing_record.update(record)
                                updated_count += 1
                        else:
                            # New record
                            new_records.append(record)
                            new_count += 1
                    
                    except Exception as e:
                        print(f"Error scanning folder {dir_path}: {e}")
                
                # Scan files
                for file_name in files:
                    file_path = os.path.join(root, file_name)
                    relative_path = os.path.relpath(file_path, self.selected_folder)
                    
                    try:
                        file_size = os.path.getsize(file_path)
                        file_size_mb = round(file_size / (1024 * 1024), 2)
                        last_modified = datetime.fromtimestamp(os.path.getmtime(file_path)).strftime("%Y-%m-%d %H:%M:%S")
                        
                        record = {
                            'file_path': file_path,
                            'status': 'file',
                            'ukuran_mb': file_size_mb,
                            'last_modified': last_modified,
                            'scan_time': scan_time,
                            'relative_path': relative_path
                        }
                        
                        # Check if already exists
                        if file_path in existing_paths:
                            # Update existing record
                            existing_record = existing_paths[file_path]
                            if (existing_record['last_modified'] != last_modified or 
                                existing_record['ukuran_mb'] != file_size_mb):
                                existing_record.update(record)
                                updated_count += 1
                        else:
                            # New record
                            new_records.append(record)
                            new_count += 1
                    
                    except Exception as e:
                        print(f"Error scanning file {file_path}: {e}")
            
            # Add new records to scan_results
            self.scan_results.extend(new_records)
            
            # Save to database
            self.save_database()
            
            progress_window.destroy()
            
            # Update UI
            self.record_count_var.set(str(len(self.scan_results)))
            self.status_var.set(f"✅ Scan selesai: {new_count} baru, {updated_count} updated")
            self.refresh_treeview()
            self.sync_btn.config(state=tk.NORMAL)
            
            # Show result
            messagebox.showinfo(
                "Scan Selesai",
                f"Scan berhasil diselesaikan!\n\n"
                f"📊 Statistik:\n"
                f"• Record baru: {new_count}\n"
                f"• Record di-update: {updated_count}\n"
                f"• Total record: {len(self.scan_results)}\n\n"
                f"💾 Database disimpan di: {self.database_file}"
            )
        
        except Exception as e:
            progress_window.destroy()
            messagebox.showerror("Error", f"Terjadi error saat scanning:\n{str(e)}")
            self.status_var.set(f"❌ Error: {str(e)}")
    
    def get_folder_size(self, folder_path):
        """Hitung ukuran total folder"""
        total_size = 0
        try:
            for root, dirs, files in os.walk(folder_path):
                for file in files:
                    try:
                        file_path = os.path.join(root, file)
                        total_size += os.path.getsize(file_path)
                    except (OSError, FileNotFoundError):
                        pass
        except:
            pass
        return total_size
    
    def synchronize_database(self):
        """Synchronize database dengan kondisi file saat ini"""
        if not self.scan_results:
            messagebox.showwarning("Peringatan", "Tidak ada data untuk disynchronize!")
            return
        
        result = messagebox.askyesno(
            "Konfirmasi Synchronize",
            f"Akan melakukan synchronize database dengan kondisi file saat ini.\n\n"
            f"Proses ini akan:\n"
            f"• Cek keberadaan semua file/folder dalam database\n"
            f"• Update status file yang sudah tidak ada\n"
            f"• Update ukuran dan tanggal modifikasi\n"
            f"• Tandai file yang hilang/moved\n\n"
            f"Total records: {len(self.scan_results)}\n\n"
            f"Lanjutkan?"
        )
        
        if not result:
            return
        
        # Progress dialog
        progress_window = tk.Toplevel(self.root)
        progress_window.title("Synchronizing...")
        progress_window.geometry("450x150")
        progress_window.resizable(False, False)
        
        # Center progress window
        progress_window.update_idletasks()
        x = (progress_window.winfo_screenwidth() // 2) - 225
        y = (progress_window.winfo_screenheight() // 2) - 75
        progress_window.geometry(f'450x150+{x}+{y}')
        
        progress_window.transient(self.root)
        progress_window.grab_set()
        
        # Progress content
        progress_frame = ttk.Frame(progress_window, padding="20")
        progress_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        progress_label = ttk.Label(
            progress_frame, 
            text="Synchronizing database...",
            font=("Arial", 10, "bold")
        )
        progress_label.grid(row=0, column=0, pady=(0, 10))
        
        progress_bar = ttk.Progressbar(
            progress_frame, 
            mode='determinate',
            maximum=len(self.scan_results)
        )
        progress_bar.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        
        status_label = ttk.Label(
            progress_frame,
            text="",
            font=("Arial", 8),
            foreground="gray"
        )
        status_label.grid(row=2, column=0, pady=(5, 0))
        
        progress_window.update()
        
        # Synchronize
        try:
            updated_count = 0
            missing_count = 0
            sync_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            for idx, record in enumerate(self.scan_results):
                progress_bar['value'] = idx + 1
                file_path = record['file_path']
                status_label.config(text=f"Checking: {os.path.basename(file_path)}")
                progress_window.update()
                
                if os.path.exists(file_path):
                    # File/folder exists - update info
                    try:
                        if record['status'] == 'folder':
                            new_size = self.get_folder_size(file_path)
                        else:
                            new_size = os.path.getsize(file_path)
                        
                        new_size_mb = round(new_size / (1024 * 1024), 2)
                        new_modified = datetime.fromtimestamp(os.path.getmtime(file_path)).strftime("%Y-%m-%d %H:%M:%S")
                        
                        # Update if changed
                        if (record['ukuran_mb'] != new_size_mb or 
                            record['last_modified'] != new_modified):
                            record['ukuran_mb'] = new_size_mb
                            record['last_modified'] = new_modified
                            record['scan_time'] = sync_time
                            updated_count += 1
                        
                        # Ensure status is not missing
                        if record.get('status_sync') == 'MISSING':
                            record['status_sync'] = 'EXISTS'
                            updated_count += 1
                    
                    except Exception as e:
                        print(f"Error updating {file_path}: {e}")
                else:
                    # File/folder missing
                    if record.get('status_sync') != 'MISSING':
                        record['status_sync'] = 'MISSING'
                        record['scan_time'] = sync_time
                        missing_count += 1
            
            # Save database
            self.save_database()
            
            progress_window.destroy()
            
            # Update UI
            self.status_var.set(f"🔄 Sync selesai: {updated_count} updated, {missing_count} missing")
            self.refresh_treeview()
            
            # Show result
            messagebox.showinfo(
                "Synchronize Selesai",
                f"Database berhasil di-synchronize!\n\n"
                f"📊 Statistik:\n"
                f"• Record di-update: {updated_count}\n"
                f"• File/folder hilang: {missing_count}\n"
                f"• Total record: {len(self.scan_results)}\n\n"
                f"💾 Database disimpan di: {self.database_file}"
            )
        
        except Exception as e:
            progress_window.destroy()
            messagebox.showerror("Error", f"Terjadi error saat synchronize:\n{str(e)}")
    
    def save_database(self):
        """Simpan database ke Excel"""
        try:
            if self.scan_results:
                # Prepare data untuk DataFrame
                df_data = []
                for idx, record in enumerate(self.scan_results, 1):
                    df_data.append({
                        'ID': idx,
                        'File_Path': record['file_path'],
                        'Status': record['status'],
                        'Ukuran_MB': record['ukuran_mb'],
                        'Last_Modified': record['last_modified'],
                        'Scan_Time': record['scan_time'],
                        'Relative_Path': record.get('relative_path', ''),
                        'Status_Sync': record.get('status_sync', 'EXISTS')
                    })
                
                df = pd.DataFrame(df_data)
                df.to_excel(self.database_file, index=False, sheet_name="Universal_Scan_DB")
        
        except Exception as e:
            messagebox.showerror("Error", f"Gagal menyimpan database:\n{str(e)}")
    
    def export_database(self):
        """Export database ke lokasi lain"""
        if not self.scan_results:
            messagebox.showwarning("Peringatan", "Tidak ada data untuk di-export!")
            return
        
        file_path = filedialog.asksaveasfilename(
            title="Export Universal Scan Database",
            defaultextension=".xlsx",
            initialfile=f"universal_scan_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            filetypes=[
                ("Excel Files", "*.xlsx"),
                ("All Files", "*.*")
            ]
        )
        
        if file_path:
            try:
                # Prepare data dengan format yang lebih detail
                df_data = []
                for idx, record in enumerate(self.scan_results, 1):
                    df_data.append({
                        'ID': idx,
                        'File_Path': record['file_path'],
                        'Status': record['status'],
                        'Ukuran_MB': record['ukuran_mb'],
                        'Last_Modified': record['last_modified'],
                        'Scan_Time': record['scan_time'],
                        'Relative_Path': record.get('relative_path', ''),
                        'Status_Sync': record.get('status_sync', 'EXISTS'),
                        'File_Name': os.path.basename(record['file_path']),
                        'Directory': os.path.dirname(record['file_path']),
                        'Extension': os.path.splitext(record['file_path'])[1] if record['status'] == 'file' else ''
                    })
                
                df = pd.DataFrame(df_data)
                
                # Create summary sheet
                summary_data = [{
                    'Informasi': 'Export Time',
                    'Value': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                }, {
                    'Informasi': 'Total Records',
                    'Value': len(self.scan_results)
                }, {
                    'Informasi': 'Total Files',
                    'Value': len([r for r in self.scan_results if r['status'] == 'file'])
                }, {
                    'Informasi': 'Total Folders',
                    'Value': len([r for r in self.scan_results if r['status'] == 'folder'])
                }, {
                    'Informasi': 'Missing Items',
                    'Value': len([r for r in self.scan_results if r.get('status_sync') == 'MISSING'])
                }, {
                    'Informasi': 'Total Size (MB)',
                    'Value': round(sum([r['ukuran_mb'] for r in self.scan_results]), 2)
                }]
                
                df_summary = pd.DataFrame(summary_data)
                
                # Export to Excel with multiple sheets
                with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name="Database", index=False)
                    df_summary.to_excel(writer, sheet_name="Summary", index=False)
                
                messagebox.showinfo(
                    "Export Berhasil",
                    f"Database berhasil di-export!\n\n"
                    f"File: {os.path.basename(file_path)}\n"
                    f"Total records: {len(self.scan_results)}\n"
                    f"Sheets: Database, Summary"
                )
            
            except Exception as e:
                messagebox.showerror("Error", f"Gagal export database:\n{str(e)}")
    
    def refresh_treeview(self):
        """Refresh tampilan treeview"""
        # Clear existing items
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        # Add items from scan_results
        for idx, record in enumerate(self.scan_results, 1):
            status_display = record['status']
            if record.get('status_sync') == 'MISSING':
                status_display = f"{record['status']} (MISSING)"
            
            self.tree.insert("", tk.END, values=(
                idx,
                record['file_path'],
                status_display,
                record['ukuran_mb'],
                record['last_modified'],
                record['scan_time']
            ))
    
    def open_selected_item(self, event):
        """Buka file/folder yang dipilih dengan double-click"""
        selected_item = self.tree.selection()
        
        if not selected_item:
            return
        
        item_values = self.tree.item(selected_item[0], "values")
        
        if not item_values:
            return
        
        file_path = item_values[1]  # File_Path column
        
        if not os.path.exists(file_path):
            messagebox.showerror(
                "File/Folder Tidak Ditemukan",
                f"File atau folder tidak ditemukan:\n{file_path}\n\n"
                f"Mungkin sudah dipindah atau dihapus.\n"
                f"Gunakan 'Synchronize Database' untuk update status."
            )
            return
        
        try:
            os.startfile(file_path)
        except Exception as e:
            messagebox.showerror(
                "Gagal Membuka",
                f"Gagal membuka file/folder:\n{str(e)}\n\n"
                f"Path: {file_path}"
            )
    
    def clear_database(self):
        """Clear semua data database"""
        if not self.scan_results:
            messagebox.showinfo("Info", "Database sudah kosong!")
            return
        
        result = messagebox.askyesno(
            "Konfirmasi Clear Database",
            f"Apakah Anda yakin ingin menghapus semua data database?\n\n"
            f"Total records yang akan dihapus: {len(self.scan_results)}\n\n"
            f"⚠️ Tindakan ini tidak dapat di-undo!"
        )
        
        if result:
            self.scan_results = []
            self.record_count_var.set("0")
            self.refresh_treeview()
            self.sync_btn.config(state=tk.DISABLED)
            
            # Delete database file
            if os.path.exists(self.database_file):
                try:
                    os.remove(self.database_file)
                    self.status_var.set("🗑️ Database berhasil di-clear")
                except Exception as e:
                    self.status_var.set(f"⚠️ Error deleting file: {str(e)}")
            
            messagebox.showinfo("Database Cleared", "Database berhasil di-clear!")
    
    def back_to_menu(self):
        """Kembali ke menu utama"""
        if self.parent_window:
            self.root.destroy()
            self.parent_window.deiconify()


class ScanLargeFilesApp:
    """Form untuk Scan File Besar (>10MB) dari Folder Arsip Digital Owncloud"""
    
    def __init__(self, root, parent_window=None):
        self.root = root
        self.parent_window = parent_window
        
        # File yang akan diabaikan (owncloud sync files)
        self.ignored_files = {
            '.owncloudsync.log',
            '.owncloudsync.log.1',
            '.sync_journal.db',
            '.sync_journal.db-wal'
        }
        
        # Format dokumen yang umum/diizinkan (untuk mode format)
        self.allowed_extensions = {
            # Office Documents
            '.doc', '.docx', '.xls', '.xlsx', '.xlsm', '.ppt', '.pptx',
            '.odt', '.ods', '.odp',  # OpenOffice/LibreOffice
            # PDF
            '.pdf',
            # Text
            '.txt', '.rtf', '.csv',
            # Images
            '.jpg', '.jpeg', '.png', '.gif', '.bmp', 
        }
        
        # Default minimum size in MB
        self.min_size_mb = 10
        
        self.setup_window()
        self.create_widgets()
        
        # Variables
        self.selected_folder = ""
        self.scan_results = []
    
    def setup_window(self):
        """Setup window utama aplikasi"""
        self.root.title("Scan File Besar (>10MB) - Arsip Digital Owncloud")
        
        # Get screen dimensions
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        
        # Get responsive dimensions
        width, height, self.padding, self.fonts = get_responsive_dimensions(
            900, 900, screen_width, screen_height
        )
        
        self.root.geometry(f"{width}x{height}")
        self.root.resizable(True, True)
        
        # Set minimum size
        self.root.minsize(650, 600)
        
        # Center window
        self.center_window()
    
    def center_window(self):
        """Center window di layar"""
        self.root.update_idletasks()
        width = self.root.winfo_width()
        height = self.root.winfo_height()
        x = (self.root.winfo_screenwidth() // 2) - (width // 2)
        y = (self.root.winfo_screenheight() // 2) - (height // 2)
        self.root.geometry(f'{width}x{height}+{x}+{y}')
    
    def create_widgets(self):
        """Membuat semua widget GUI dengan scrollable canvas"""
        # Main container frame
        container = ttk.Frame(self.root)
        container.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Configure grid weights
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        container.columnconfigure(0, weight=1)
        container.rowconfigure(0, weight=1)
        
        # Create canvas with scrollbar
        canvas = tk.Canvas(container, highlightthickness=0)
        scrollbar = ttk.Scrollbar(container, orient="vertical", command=canvas.yview)
        
        # Scrollable frame inside canvas
        scrollable_frame = ttk.Frame(canvas)
        
        # Configure canvas scroll region
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        # Create window in canvas
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Grid layout
        canvas.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))
        
        # Enable mousewheel scrolling
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        
        canvas.bind_all("<MouseWheel>", _on_mousewheel)
        
        # Main frame dengan padding responsif
        main_frame = ttk.Frame(scrollable_frame, padding=str(self.padding))
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Configure main_frame
        scrollable_frame.columnconfigure(0, weight=1)
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(5, weight=1)  # Result frame expandable
        
        # Title dengan font responsif
        title_label = ttk.Label(
            main_frame, 
            text="🔍 SCAN FILE BESAR & FORMAT NON-DOKUMEN", 
            font=("Arial", self.fonts['title'], "bold")
        )
        title_label.grid(row=0, column=0, pady=(0, 10))
        
        # Subtitle
        subtitle_label = ttk.Label(
            main_frame, 
            text="Temukan file berukuran besar atau file dengan format tidak umum",
            font=("Arial", self.fonts['subtitle'])
        )
        subtitle_label.grid(row=1, column=0, pady=(0, 20))
        
        # Frame untuk mode scan dengan padding responsif
        frame_padding = max(10, self.padding - 5)
        mode_frame = ttk.LabelFrame(main_frame, text="Mode Scan", padding=str(frame_padding))
        mode_frame.grid(row=2, column=0, sticky=(tk.W, tk.E), pady=(0, 15))
        mode_frame.columnconfigure(0, weight=1)
        
        # Radio buttons untuk memilih mode
        self.scan_mode = tk.StringVar(value="size")
        
        mode_size_rb = ttk.Radiobutton(
            mode_frame,
            text="📏 File Besar (berdasarkan ukuran minimum)",
            variable=self.scan_mode,
            value="size",
            command=self.on_mode_change
        )
        mode_size_rb.grid(row=0, column=0, sticky=tk.W, pady=(0, 5))
        
        mode_format_rb = ttk.Radiobutton(
            mode_frame,
            text="📄 Format Non-Dokumen (selain Office, Image, PDF, TXT)",
            variable=self.scan_mode,
            value="format",
            command=self.on_mode_change
        )
        mode_format_rb.grid(row=1, column=0, sticky=tk.W, pady=(5, 0))
        
        # Frame untuk ukuran minimum (hanya aktif jika mode = size)
        self.size_frame = ttk.LabelFrame(main_frame, text="Pengaturan Ukuran", padding=str(frame_padding))
        self.size_frame.grid(row=3, column=0, sticky=(tk.W, tk.E), pady=(0, 15))
        self.size_frame.columnconfigure(1, weight=1)
        
        # Label dan input untuk ukuran minimum
        ttk.Label(self.size_frame, text="Ukuran Minimum (MB):").grid(row=0, column=0, sticky=tk.W, padx=(0, 10))
        
        self.size_var = tk.StringVar(value="10")
        self.size_entry = ttk.Entry(self.size_frame, textvariable=self.size_var, width=10)
        self.size_entry.grid(row=0, column=1, sticky=tk.W)
        
        ttk.Label(self.size_frame, text="(File yang lebih kecil akan diabaikan)", 
                 font=("Arial", self.fonts['small']), foreground="gray").grid(row=0, column=2, sticky=tk.W, padx=(10, 0))
        
        # Frame untuk folder selection
        folder_frame = ttk.LabelFrame(main_frame, text="Pilih Folder Arsip Digital Owncloud", padding=str(frame_padding))
        folder_frame.grid(row=4, column=0, sticky=(tk.W, tk.E), pady=(0, 15))
        folder_frame.columnconfigure(0, weight=1)
        
        # Folder path display dengan wraplength responsif
        wrap_length = max(500, int(self.root.winfo_screenwidth() * 0.7))
        self.folder_var = tk.StringVar(value="Belum ada folder yang dipilih...")
        folder_path_label = ttk.Label(
            folder_frame, 
            textvariable=self.folder_var,
            foreground="gray",
            wraplength=wrap_length
        )
        folder_path_label.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        
        # Folder browse button
        folder_btn = ttk.Button(
            folder_frame, 
            text="📂 Browse Folder", 
            command=self.browse_folder
        )
        folder_btn.grid(row=1, column=0)
        
        # Frame untuk hasil scan
        result_frame = ttk.LabelFrame(main_frame, text="Hasil Scan", padding="15")
        result_frame.grid(row=5, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(0, 15))
        result_frame.columnconfigure(0, weight=1)
        result_frame.rowconfigure(0, weight=1)
        
        # Treeview untuk menampilkan hasil
        columns = ("No", "Nama File", "Ekstensi", "Ukuran", "Path")
        self.tree = ttk.Treeview(result_frame, columns=columns, show="headings", height=15)
        
        # Define headings
        self.tree.heading("No", text="No")
        self.tree.heading("Nama File", text="Nama File")
        self.tree.heading("Ekstensi", text="Ekstensi")
        self.tree.heading("Ukuran", text="Ukuran")
        self.tree.heading("Path", text="Path Lengkap")
        
        # Define column widths
        self.tree.column("No", width=50, anchor=tk.CENTER)
        self.tree.column("Nama File", width=200)
        self.tree.column("Ekstensi", width=80, anchor=tk.CENTER)
        self.tree.column("Ukuran", width=100, anchor=tk.E)
        self.tree.column("Path", width=400)
        
        # Scrollbars
        vsb = ttk.Scrollbar(result_frame, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(result_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        
        # Grid
        self.tree.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        vsb.grid(row=0, column=1, sticky=(tk.N, tk.S))
        hsb.grid(row=1, column=0, sticky=(tk.W, tk.E))
        
        # Info label
        self.info_var = tk.StringVar(value="Pilih folder untuk memulai scan")
        info_label = ttk.Label(main_frame, textvariable=self.info_var, font=("Arial", 9), foreground="blue")
        info_label.grid(row=6, column=0, pady=(0, 10))
        
        # Action buttons frame
        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=7, column=0, pady=(10, 0))
        
        # Scan button
        self.scan_btn = ttk.Button(
            button_frame, 
            text="🔍 Mulai Scan", 
            command=self.start_scan,
            state="disabled"
        )
        self.scan_btn.grid(row=0, column=0, padx=(0, 10))
        
        # Export button
        self.export_btn = ttk.Button(
            button_frame, 
            text="📊 Export ke Excel", 
            command=self.export_to_excel,
            state="disabled"
        )
        self.export_btn.grid(row=0, column=1, padx=(10, 10))
        
        # Clear button
        self.clear_btn = ttk.Button(
            button_frame, 
            text="🗑️ Clear", 
            command=self.clear_results,
            state="disabled"
        )
        self.clear_btn.grid(row=0, column=2, padx=(10, 10))
        
        # Back button
        if self.parent_window:
            back_btn = ttk.Button(
                button_frame, 
                text="⬅️ Kembali", 
                command=self.back_to_menu
            )
            back_btn.grid(row=0, column=3, padx=(10, 0))
        
        # Update canvas width to match window
        def _configure_canvas(event):
            canvas.itemconfig(canvas.find_withtag("all")[0], width=event.width)
        
        canvas.bind("<Configure>", _configure_canvas)
    
    def on_mode_change(self):
        """Handle perubahan mode scan"""
        mode = self.scan_mode.get()
        
        if mode == "size":
            # Enable size input
            self.size_entry.config(state="normal")
            self.info_var.set("Mode: File Besar - Pilih folder untuk memulai scan")
        else:  # format
            # Disable size input
            self.size_entry.config(state="disabled")
            self.info_var.set("Mode: Format Non-Dokumen - Pilih folder untuk memulai scan")
    
    def browse_folder(self):
        """Fungsi untuk memilih folder"""
        # Gunakan default folder jika ada
        default_folder = config_manager.get_default_folder()
        initial_dir = default_folder if default_folder and os.path.exists(default_folder) else os.getcwd()
        
        folder_path = filedialog.askdirectory(
            title="Pilih Folder Arsip Digital Owncloud",
            initialdir=initial_dir
        )
        
        if folder_path:
            if os.path.exists(folder_path) and os.path.isdir(folder_path):
                self.selected_folder = folder_path
                self.folder_var.set(folder_path)
                self.scan_btn.config(state="normal")
                self.info_var.set(f"Folder dipilih: {os.path.basename(folder_path)}")
            else:
                messagebox.showerror("Error", "Folder yang dipilih tidak valid!")
                self.selected_folder = ""
    
    def start_scan(self):
        """Mulai scan file besar atau format non-dokumen"""
        if not self.selected_folder:
            messagebox.showwarning("Peringatan", "Silakan pilih folder terlebih dahulu!")
            return
        
        mode = self.scan_mode.get()
        
        # Validasi input ukuran minimum jika mode = size
        if mode == "size":
            try:
                self.min_size_mb = float(self.size_var.get())
                if self.min_size_mb <= 0:
                    messagebox.showerror("Error", "Ukuran minimum harus lebih dari 0 MB!")
                    return
            except ValueError:
                messagebox.showerror("Error", "Ukuran minimum harus berupa angka!")
                return
        
        # Clear previous results
        self.clear_results()
        
        # Progress dialog
        if mode == "size":
            progress_msg = f"Scanning file lebih dari {self.min_size_mb} MB..."
        else:
            progress_msg = "Scanning file dengan format non-dokumen..."
        
        progress_window = self.show_progress_dialog(progress_msg)
        
        try:
            # Scan folder
            self.scan_results = []
            self.scan_folder_recursive(self.selected_folder)
            
            progress_window.destroy()
            
            # Sort by size (descending)
            self.scan_results.sort(key=lambda x: x['size_bytes'], reverse=True)
            
            # Display results
            self.display_results()
            
            # Update info
            total_files = len(self.scan_results)
            total_size = sum(f['size_bytes'] for f in self.scan_results)
            total_size_mb = total_size / (1024 * 1024)
            
            if mode == "size":
                self.info_var.set(
                    f"✅ Scan selesai! Ditemukan {total_files} file >{self.min_size_mb}MB (Total: {total_size_mb:.2f} MB)"
                )
            else:
                self.info_var.set(
                    f"✅ Scan selesai! Ditemukan {total_files} file format non-dokumen (Total: {total_size_mb:.2f} MB)"
                )
            
            # Enable buttons
            if total_files > 0:
                self.export_btn.config(state="normal")
                self.clear_btn.config(state="normal")
            
        except Exception as e:
            progress_window.destroy()
            messagebox.showerror("Error", f"Terjadi kesalahan saat scan:\n{str(e)}")
    
    def scan_folder_recursive(self, folder_path):
        """Scan folder secara recursive berdasarkan mode yang dipilih"""
        try:
            mode = self.scan_mode.get()
            
            if mode == "size":
                min_size_bytes = self.min_size_mb * 1024 * 1024  # Convert MB to bytes
            
            for root, dirs, files in os.walk(folder_path):
                for file in files:
                    try:
                        # Skip ignored files (owncloud sync files)
                        if file in self.ignored_files:
                            continue
                        
                        file_path = os.path.join(root, file)
                        file_size = os.path.getsize(file_path)
                        
                        # Get file extension
                        _, ext = os.path.splitext(file)
                        ext = ext.lower()
                        
                        # Check berdasarkan mode
                        if mode == "size":
                            # Mode: File Besar - Check if file >= min_size_mb
                            if file_size >= min_size_bytes:
                                self.scan_results.append({
                                    'name': file,
                                    'size_bytes': file_size,
                                    'size_mb': file_size / (1024 * 1024),
                                    'path': file_path,
                                    'extension': ext if ext else '(no ext)'
                                })
                        else:  # mode == "format"
                            # Mode: Format Non-Dokumen - Check if extension NOT in allowed list
                            if ext not in self.allowed_extensions:
                                self.scan_results.append({
                                    'name': file,
                                    'size_bytes': file_size,
                                    'size_mb': file_size / (1024 * 1024),
                                    'path': file_path,
                                    'extension': ext if ext else '(no ext)'
                                })
                    except Exception as e:
                        # Skip files that can't be accessed
                        continue
        except Exception as e:
            print(f"Error scanning folder: {str(e)}")
    
    def display_results(self):
        """Menampilkan hasil scan di treeview"""
        # Clear existing items
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        # Add results
        for idx, file_info in enumerate(self.scan_results, 1):
            self.tree.insert("", tk.END, values=(
                idx,
                file_info['name'],
                file_info['extension'],
                f"{file_info['size_mb']:.2f} MB",
                file_info['path']
            ))
    
    def export_to_excel(self):
        """Export hasil scan ke Excel"""
        if not self.scan_results:
            messagebox.showwarning("Peringatan", "Tidak ada data untuk di-export!")
            return
        
        try:
            file_path = filedialog.asksaveasfilename(
                title="Export ke Excel",
                defaultextension=".xlsx",
                initialfile=f"file_besar_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                filetypes=[
                    ("Excel Files", "*.xlsx"),
                    ("All Files", "*.*")
                ]
            )
            
            if file_path:
                # Create DataFrame
                data = []
                for idx, file_info in enumerate(self.scan_results, 1):
                    data.append({
                        'No': idx,
                        'Nama File': file_info['name'],
                        'Ekstensi': file_info['extension'],
                        'Ukuran (MB)': round(file_info['size_mb'], 2),
                        'Ukuran (Bytes)': file_info['size_bytes'],
                        'Path Lengkap': file_info['path']
                    })
                
                df = pd.DataFrame(data)
                
                # Export to Excel
                df.to_excel(file_path, index=False, sheet_name="File_Besar")
                
                messagebox.showinfo(
                    "Export Berhasil",
                    f"Data berhasil di-export!\n\n"
                    f"File: {file_path}\n"
                    f"Total: {len(self.scan_results)} file"
                )
        except Exception as e:
            messagebox.showerror("Error", f"Gagal export ke Excel:\n{str(e)}")
    
    def clear_results(self):
        """Clear hasil scan"""
        # Clear treeview
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        # Clear results
        self.scan_results = []
        
        # Disable buttons
        self.export_btn.config(state="disabled")
        self.clear_btn.config(state="disabled")
        
        # Update info
        self.info_var.set("Hasil scan telah dibersihkan. Pilih folder untuk scan ulang.")
    
    def show_progress_dialog(self, message):
        """Menampilkan dialog progress"""
        progress_window = tk.Toplevel(self.root)
        progress_window.title("Processing...")
        progress_window.geometry("350x100")
        progress_window.resizable(False, False)
        
        # Center window
        progress_window.update_idletasks()
        x = (progress_window.winfo_screenwidth() // 2) - (175)
        y = (progress_window.winfo_screenheight() // 2) - (50)
        progress_window.geometry(f'350x100+{x}+{y}')
        
        # Make it modal
        progress_window.transient(self.root)
        progress_window.grab_set()
        
        # Progress content
        frame = ttk.Frame(progress_window, padding="20")
        frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        ttk.Label(frame, text=message, font=("Arial", 10)).grid(row=0, column=0, pady=(0, 10))
        
        progress_bar = ttk.Progressbar(frame, mode='indeterminate')
        progress_bar.grid(row=1, column=0, sticky=(tk.W, tk.E))
        progress_bar.start()
        
        progress_window.update()
        return progress_window
    
    def back_to_menu(self):
        """Kembali ke menu utama"""
        if self.parent_window:
            self.root.destroy()
            self.parent_window.deiconify()
